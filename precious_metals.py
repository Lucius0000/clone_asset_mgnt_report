"""
商品与贵金属，输出 commodity_indicators_summary.xlsx
"""

import os
from datetime import datetime, timedelta
import yfinance as yf
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment
import math
import logging
import akshare as ak

# ====== 调试打印与日志 ======
DEBUG = False  # 控制是否在控制台打印；日志文件始终写入
output_folder = "output"
raw_data_folder = os.path.join(output_folder, "raw_data")
os.makedirs(raw_data_folder, exist_ok=True)

LOG_PATH = os.path.join(raw_data_folder, "precious_metals_calculation_steps.log")

logger = logging.getLogger("pm_calc")
logger.setLevel(logging.DEBUG)
# 文件日志
_file_handler = logging.FileHandler(LOG_PATH, encoding="utf-8")
_file_handler.setLevel(logging.DEBUG)
_formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
_file_handler.setFormatter(_formatter)
# 避免重复添加 handler（例如在交互式环境多次运行）
if not any(isinstance(h, logging.FileHandler) and getattr(h, 'baseFilename', '') == _file_handler.baseFilename
           for h in logger.handlers):
    logger.addHandler(_file_handler)
# 控制台日志（可选）
if DEBUG and not any(isinstance(h, logging.StreamHandler) for h in logger.handlers):
    _console = logging.StreamHandler()
    _console.setLevel(logging.DEBUG)
    _console.setFormatter(_formatter)
    logger.addHandler(_console)

def dprint(*args):
    msg = " ".join(str(a) for a in args)
    logger.debug(msg)
    if DEBUG:
        print(msg)

# 你原有的代理设置（如不需要可移除）
os.environ['http_proxy'] = 'http://127.0.0.1:7890'
os.environ['https_proxy'] = 'http://127.0.0.1:7890'

# ====== 基础工具 ======
def get_period_data(df, offset: pd.DateOffset):
    df = df.sort_index()
    end_date = df.index[-1]
    start_date = end_date - offset
    return df[df.index >= start_date]

def find_nearest_index_within_window(df, target_date, window=5):
    date_diffs = np.abs(df.index - target_date)
    in_window = date_diffs[date_diffs <= pd.Timedelta(days=window)]
    return df.index[np.argmin(date_diffs)] if not in_window.empty else None

# yfinance 标的（保持不变）
tickers = ['GLD', 'CL=F', 'HG=F']
data_dict = {}

for symbol in tickers:
    ticker = yf.Ticker(symbol)
    hist = ticker.history(period='6y')
    # 统一：确保索引为升序日期
    hist = hist.sort_index()
    hist.to_csv(os.path.join(raw_data_folder, f"{symbol}.csv"), encoding='utf-8-sig')
    data_dict[symbol] = hist

# ====== 新增：接入 “上海金基准价” ======
# ak.spot_golden_benchmark_sge() 返回列：交易时间、晚盘价、早盘价（附件验证）
try:
    sge_df = ak.spot_golden_benchmark_sge()
    # 兼容列名：
    # 去掉潜在的多余列（如 Unnamed: 0），标准化索引为日期
    sge_df = sge_df.copy()
    if 'Unnamed: 0' in sge_df.columns:
        sge_df = sge_df.drop(columns=['Unnamed: 0'])
    # 转为日期索引
    if '交易时间' in sge_df.columns:
        sge_df['交易时间'] = pd.to_datetime(sge_df['交易时间'])
        sge_df = sge_df.set_index('交易时间')
    sge_df = sge_df.sort_index()
    # 原样保存原始数据
    sge_df.to_csv(os.path.join(raw_data_folder, "SGE_GOLD.csv"), encoding='utf-8-sig')
except Exception as e:
    dprint(f"[SGE] 下载失败：{e}")
    sge_df = None

# 显示名称映射（仅供 yfinance 行使用；SGE 行走自定义覆盖）
category_map = {
    'GLD': '黄金ETF',
    'CL=F': '原油期货',
    'HG=F': '铜期货'
}

# ====== 年化/波动率/Sharpe ======
def calculate_annualized_return(data):
    closes = data.dropna()
    if len(closes) < 2:
        return np.nan
    total_return = closes.iloc[-1] / closes.iloc[0] - 1.0
    returns = closes.pct_change().dropna()
    N = len(returns)
    if N <= 0:
        return np.nan
    k = 252.0 / N
    return (1.0 + total_return) ** k - 1.0

def calculate_annualized_volatility(returns):
    std_daily = returns.dropna().std()
    return std_daily * math.sqrt(252.0)

def calculate_sharpe_ratio(ann_return, ann_vol, rf=0.045):
    if ann_vol is None or pd.isna(ann_vol) or ann_vol == 0:
        return np.nan
    return (ann_return - rf) / ann_vol

# ====== 通用指标计算（适配 price_col / volume_col / 类别与市值覆盖） ======
def compute_indicators(
    df,
    ticker_symbol,
    *,
    volume_display_override=None,
    price_col='Close',
    volume_col='Volume',
    category_override=None,
    market_cap_override=None
):
    df = df.copy().sort_index()
    
    # 成交量（若无该列则留空）
    if volume_col and (volume_col in df.columns):
        current_volume = df.iloc[-1][volume_col]
    else:
        current_volume = np.nan

    # 价格序列（核心适配处）
    if price_col not in df.columns:
        raise KeyError(f"价格列 {price_col} 不存在，现有列：{list(df.columns)}")
    price = df[price_col].astype(float)
    returns = price.pct_change()

    end_date = df.index[-1]
    current_close = float(price.iloc[-1])

    # 月中值（仍按“接近每月15日”的原则，用 price_col）
    if end_date.day >= 15:
        monthly_date = end_date.replace(day=15)
    else:
        prev_month_for_mid = end_date - pd.DateOffset(months=1)
        monthly_date = prev_month_for_mid.replace(day=15)
    idx_mid = find_nearest_index_within_window(df, monthly_date)
    monthly_close = float(price.loc[idx_mid]) if idx_mid is not None else np.nan

    # 百分点位：近 252 个交易日内的位置
    close_series_1y = price.dropna().tail(252)
    percentile = round((close_series_1y <= current_close).sum() / len(close_series_1y) * 100, 2) if len(close_series_1y) > 0 else np.nan

    # 成交量（若无该列则留空）
    if volume_col and (volume_col in df.columns):
        current_volume = df.iloc[-1][volume_col]
    else:
        current_volume = np.nan

    # 市值（yfinance 的 GLD 自动取；若指定覆盖则用覆盖值）
    if market_cap_override is not None:
        market_cap = market_cap_override
    else:
        if ticker_symbol == 'GLD':
            try:
                raw_cap = yf.Ticker('GLD').info.get('marketCap', None)
            except Exception:
                raw_cap = None
            market_cap = round(raw_cap / 1e9, 2) if raw_cap else 'n/a'
        else:
            market_cap = 'n/a'

    # 窗口设定：短（1月）、中（1年）、长（5年）
    periods = {
        "短期": pd.DateOffset(months=1),
        "中期": pd.DateOffset(years=1),
        "长期": pd.DateOffset(years=5)
    }
    period_dfs = {label: get_period_data(df, offset) for label, offset in periods.items()}

    volatility = {}
    sharpe = {}
    annualized_return_map = {}

    dprint(f"\n===== {ticker_symbol} | {category_override or category_map.get(ticker_symbol, '未知')} =====")
    dprint(f"[{price_col}] 结束日期: {end_date.date()}  当前值: {current_close:.6f}")

    for label, period_df in period_dfs.items():
        closes = period_df[price_col].dropna()
        rets = closes.pct_change().dropna()
        if len(closes) < 2 or len(rets) < 1:
            annualized_return_map[label] = np.nan
            volatility[label] = np.nan
            sharpe[label] = np.nan
            dprint(f"[{label}] 数据不足，跳过。")
            continue

        start_price = float(closes.iloc[0])
        end_price_p = float(closes.iloc[-1])
        total_R = end_price_p / start_price - 1.0
        N = len(rets)
        k = 252.0 / N
        ann_ret = (1.0 + total_R) ** k - 1.0
        std_daily = rets.std()
        ann_vol = std_daily * math.sqrt(252.0)
        sr = calculate_sharpe_ratio(ann_ret, ann_vol, rf=0.045)

        annualized_return_map[label] = ann_ret
        volatility[label] = ann_vol
        sharpe[label] = sr

        dprint(f"[{label}] 起止: {period_df.index[0].date()} -> {period_df.index[-1].date()}  首末: {start_price:.6f}->{end_price_p:.6f}  R={total_R:.6f}")
        dprint(f"[{label}] N={N} k={k:.6f}  ann_ret={ann_ret:.6f}  ann_vol={ann_vol:.6f}  sharpe={sr:.6f}")

    # 统一口径 MoM（短期窗口）
    short_df = period_dfs.get("短期")
    if short_df is not None and len(short_df) >= 2:
        mom_start_close = float(short_df[price_col].iloc[0])
        mom_end_close = float(short_df[price_col].iloc[-1])
        mom = (mom_end_close / mom_start_close - 1.0) * 100.0
    else:
        mom = np.nan

    # 统一口径 YoY（中期窗口）
    mid_df = period_dfs.get("中期")
    if mid_df is not None and len(mid_df) >= 2:
        yoy_start_close = float(mid_df[price_col].iloc[0])
        yoy_end_close = float(mid_df[price_col].iloc[-1])
        yoy = (yoy_end_close / yoy_start_close - 1.0) * 100.0
    else:
        yoy = np.nan

    category = category_override if category_override is not None else category_map.get(ticker_symbol, '未知')

    def pct_format(x):
        return f"{round(x, 2)}%" if pd.notnull(x) else ""
    
    volume_display = (
    volume_display_override
    if volume_display_override is not None
    else (f"{current_volume:,.0f}" if pd.notnull(current_volume) else '')
)

    return [
        ticker_symbol,  # “指标”
        category,       # “类别”
        f"{current_close:,.2f}",
        f"{monthly_close:,.2f}" if pd.notnull(monthly_close) else '',
        (f"{market_cap:,.0f} B USD" if isinstance(market_cap, (int, float)) else (market_cap if market_cap is not None else '')),
        volume_display, 
        #f"{current_volume:,.0f}" if pd.notnull(current_volume) else '',
        pct_format(mom),
        pct_format(yoy),
        pct_format(percentile),
        pct_format(volatility['短期'] * 100 if pd.notnull(volatility.get('短期')) else np.nan),
        pct_format(volatility['中期'] * 100 if pd.notnull(volatility.get('中期')) else np.nan),
        pct_format(volatility['长期'] * 100 if pd.notnull(volatility.get('长期')) else np.nan),
        f"{sharpe['短期']:,.4f}" if pd.notnull(sharpe.get('短期')) else "",
        f"{sharpe['中期']:,.4f}" if pd.notnull(sharpe.get('中期')) else "",
        f"{sharpe['长期']:,.4f}" if pd.notnull(sharpe.get('长期')) else "",
        pct_format(annualized_return_map["短期"] * 100 if pd.notnull(annualized_return_map.get("短期")) else np.nan),
        pct_format(annualized_return_map["中期"] * 100 if pd.notnull(annualized_return_map.get("中期")) else np.nan),
        pct_format(annualized_return_map["长期"] * 100 if pd.notnull(annualized_return_map.get("长期")) else np.nan)
    ]

# ====== 生成各行 ======
data_rows = []

# 先处理 yfinance 三个标的（按原逻辑的 Close/Volume）
for symbol, df in data_dict.items():
    row = compute_indicators(
        df,
        symbol,
        price_col='Close',
        volume_col='Volume',
        category_override=None,
        market_cap_override=None  # GLD 会自动取市值，其它为 'n/a'
    )
    data_rows.append(row)

# 再追加“上海金基准价”一行（关键：price_col='晚盘价'，类别&市值覆盖，指标=GLD）
if sge_df is not None and '晚盘价' in sge_df.columns:
    sge_row = compute_indicators(
        sge_df,
        'GLD',  # —— 按你的要求，“指标”为 GLD
        price_col='晚盘价',      # —— 所有计算用“晚盘价”
        volume_col=None,        # —— 无成交量则留空
        category_override='上海金基准价（晚盘价）',  # —— 类别覆盖
        market_cap_override='',   # —— 总市值留空
        volume_display_override='n/a'  
    )
    data_rows.append(sge_row)
else:
    dprint("[SGE] 数据为空或缺少列 '晚盘价'，未追加该行。")

# ====== 写入 Excel（保持你原有表头与合并设置） ======
header_top = [
    "指标", "类别", "当前收盘价 ($)", "月中收盘值 ($)", "总市值 (B $)", "交易量（股\\合约数）", "环比 MoM(%)", "同比 YoY(%)", "百分点位",
    "波动率(%)", "波动率(%)", "波动率(%)",
    "Sharpe Ratio", "Sharpe Ratio", "Sharpe Ratio",
    "收益率年化(%)", "收益率年化(%)", "收益率年化(%)"
]

header_bottom = [
    "", "", "", "", "", "", "", "", "",
    "短期（月）", "中期（年）", "长期（5年）",
    "短期（月）", "中期（年）", "长期（5年）",
    "短期（月）", "中期（年）", "长期（5年）"
]

wb = Workbook()
ws = wb.active
ws.title = "Commodities"

ws.append(header_top)
ws.append(header_bottom)

merge_config = {
    (1, 1): (1, 2),
    (2, 2): (2, 2),
    (3, 3): (3, 2),
    (4, 4): (4, 2),
    (5, 5): (5, 2),
    (6, 6): (6, 2),
    (7, 7): (7, 2),
    (8, 8): (8, 2),
    (9, 9): (9, 2),
    (10, 12): (10, 1),
    (13, 15): (13, 1),
    (16, 18): (16, 1)
}

for (col_start, col_end), (col, row_span) in merge_config.items():
    cell = ws.cell(row=1, column=col_start)
    ws.merge_cells(start_row=1, start_column=col_start, end_row=1 + row_span - 1, end_column=col_end)
    cell.alignment = Alignment(horizontal='center', vertical='center')

for row in data_rows:
    ws.append(row)

for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

output_path = os.path.join(output_folder, "commodity_indicators_summary.xlsx")
os.makedirs(output_folder, exist_ok=True)
wb.save(output_path)

# ====== 关闭日志 handler，确保写盘 ======
for h in list(logger.handlers):
    try:
        h.flush()
    except Exception:
        pass
    try:
        h.close()
    except Exception:
        pass
    logger.removeHandler(h)
