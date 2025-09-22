'''
股指分析
输出股指表：stock_weekly_report.xlsx
'''

import akshare as ak
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import logging
from typing import Dict, Optional, Tuple, Any
from dataclasses import dataclass
import os
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from stock_cap import get_all_index_caps
import time
        

# 设置日志
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@dataclass
class MarketConfig:
    """市场配置数据类"""
    name: str
    symbol: str
    risk_free_rate: float
    data_source: str
    api_function: str

@dataclass
class TimePeriod:
    """时间周期数据类"""
    name: str
    days: int
    is_long_term: bool


class StockIndexAnalyzer:
    """股票指数分析器"""
    
    # 市场配置
    MARKETS = {
        'US': MarketConfig('S&P 500', '.inx', 0.045, 'index_us_stock_sina', 'index_us_stock_sina'),
        'CN': MarketConfig('CSI 300', 'sh000300', 0.017, 'stock_zh_index_daily', 'stock_zh_index_daily'),
        'HK': MarketConfig('Hang Seng', 'HSI', 0.0062, 'stock_hk_index_daily_sina', 'stock_hk_index_daily_sina')
    }
    
    # 时间周期配置
    # 注意：使用纯日历方法，所有时间周期都使用实际日期范围
    # Monthly: 使用实际日期范围 (30天) - 算术平均
    # Semi-Annual: 使用实际日期范围 (180天) - 几何平均 (长期业绩)
    # Annual: 使用实际日期范围 (365天) - 几何平均 (长期业绩)  
    # 5-Year: 使用实际日期范围 (1825天) - 几何平均 (长期业绩)
    TIME_PERIODS = [
        TimePeriod('Monthly', 30, False),      # 30天 (约1个月) - 算术平均
        TimePeriod('Semi-Annual', 180, True),  # 180天 (约6个月) - 几何平均
        TimePeriod('Annual', 365, True),       # 365天 (约1年) - 几何平均
        TimePeriod('5-Year', 1825, True)       # 1825天 (约5年) - 几何平均
    ]
    
    def __init__(self, debug: bool = False):
        self.debug = debug
        # 验证时间周期配置
        self._validate_time_periods()
    
    def _validate_time_periods(self):
        """验证时间周期配置的合理性"""
        if self.debug:
            logger.info("Validating time periods configuration (Pure Calendar Approach):")
            for period in self.TIME_PERIODS:
                method = "Geometric" if period.is_long_term else "Arithmetic"
                logger.info(f"  {period.name}: {period.days} calendar days, Method: {method}")
        
        # 检查时间周期的逻辑顺序
        for i in range(len(self.TIME_PERIODS) - 1):
            if self.TIME_PERIODS[i].days >= self.TIME_PERIODS[i + 1].days:
                logger.warning(f"Time period {self.TIME_PERIODS[i].name} ({self.TIME_PERIODS[i].days} days) is not shorter than {self.TIME_PERIODS[i + 1].name} ({self.TIME_PERIODS[i + 1].days} days)")
        
        # 检查长期/短期分类的合理性
        for period in self.TIME_PERIODS:
            if period.days >= 180 and not period.is_long_term:
                logger.warning(f"Period {period.name} ({period.days} days) should be classified as long-term for geometric calculation")
            elif period.days < 180 and period.is_long_term:
                logger.warning(f"Period {period.name} ({period.days} days) should be classified as short-term for arithmetic calculation")
    
    def normalize_stock_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """标准化股票数据格式"""
        if df is None or df.empty:
            return pd.DataFrame(columns=['date', 'open(point)', 'high(point)', 'low(point)', 'close(point)', 'volume(share)'])
        
        try:
            # 列名映射
            column_mappings = {
                'date': ['date', '日期', '报告日'],
                'open(point)': ['open(point)', 'open (point)', 'open', '开盘', '开盘价'],
                'high(point)': ['high(point)', 'high (point)', 'high', '最高', '最高价'],
                'low(point)': ['low(point)', 'low (point)', 'low', '最低', '最低价'],
                'close(point)': ['close(point)', 'close (point)', 'close', '收盘', '收盘价'],
                'volume(share)': ['volume(share)', 'volume (share)', 'volume', '成交量', '成交股数']
            }
            
            normalized_df = pd.DataFrame()
            
            for std_col, possible_names in column_mappings.items():
                found_col = next((col for col in possible_names if col in df.columns), None)
                if found_col:
                    normalized_df[std_col] = df[found_col]
                else:
                    normalized_df[std_col] = 0
                    if self.debug:
                        logger.warning(f"Column {std_col} not found in data, filled with 0")
            
            # 确保日期列是datetime类型并排序
            normalized_df['date'] = pd.to_datetime(normalized_df['date'])
            normalized_df = normalized_df.sort_values('date', ascending=False)
            
            if self.debug:
                logger.info(f"Normalized data shape: {normalized_df.shape}")
            
            return normalized_df
            
        except Exception as e:
            logger.error(f"Error normalizing stock data: {str(e)}")
            return pd.DataFrame(columns=['date', 'open(point)', 'high(point)', 'low(point)', 'close(point)', 'volume(share)'])
    
    def fetch_market_data(self, market: str, time_range: int) -> pd.DataFrame:
        """获取指定市场的数据"""
        config = self.MARKETS[market]
        
        try:
            if self.debug:
                logger.info(f"Fetching {config.name} data...")
            
            # 根据市场选择API函数
            if market == 'US':
                data = ak.index_us_stock_sina(symbol=config.symbol)
            elif market == 'CN':
                data = ak.stock_zh_index_daily(symbol=config.symbol)
            elif market == 'HK':
                data = ak.stock_hk_index_daily_sina(symbol=config.symbol)
            else:
                raise ValueError(f"Unknown market: {market}")
            
            if data is None or data.empty:
                logger.error(f"{config.name} data is empty or None")
                return pd.DataFrame()
            
            if self.debug:
                logger.info(f"Successfully fetched {config.name} data, shape: {data.shape}")
            
            # 标准化数据
            normalized_data = self.normalize_stock_data(data)
            
            # 过滤日期范围
            start_date = datetime.now() - timedelta(days=time_range)
            filtered_data = normalized_data[normalized_data['date'] >= start_date]
            
            if self.debug:
                logger.info(f"{config.name} data after filtering: {filtered_data.shape[0]} rows")
            
            return filtered_data
            
        except Exception as e:
            logger.error(f"Error fetching {config.name} data: {str(e)}")
            return pd.DataFrame()
    
    def calculate_daily_returns(self, df: pd.DataFrame) -> pd.DataFrame:
        """计算每日收益率"""
        # 按时间顺序计算收益率
        df_asc = df.sort_values('date', ascending=True)
        df_asc['daily_return'] = df_asc['close(point)'].pct_change()
        
        # 重新排序为降序
        return df_asc.sort_values('date', ascending=False)
    
    def get_period_data(self, df: pd.DataFrame, period: TimePeriod, current_date: datetime) -> pd.DataFrame:
        """获取指定时间周期的数据"""
        # 所有时间周期都使用实际日期范围 (纯日历方法)
        start_date = current_date - timedelta(days=period.days)
        return df[df['date'] >= start_date]
    
    def calculate_period_change(self, df: pd.DataFrame, target_days: int, current_date: datetime, current_close: float) -> Optional[float]:
        """计算指定时间周期的价格变化"""
        target_date = current_date - timedelta(days=target_days)
        
        # 在前后范围内寻找最接近的日期
        range_days = 10 if target_days > 100 else 3
        range_start = target_date - timedelta(days=range_days)
        range_end = target_date + timedelta(days=range_days)
        
        period_data = df[(df['date'] >= range_start) & (df['date'] <= range_end)]
        
        if not period_data.empty:
            # 找到最接近目标日期的数据点
            date_diffs = (period_data['date'] - target_date).abs()
            closest_idx = date_diffs.idxmin()
            closest_data = period_data.loc[closest_idx]
            
            if self.debug:
                logger.info(f"Target: {target_date.strftime('%Y-%m-%d')}, Found: {closest_data['date'].strftime('%Y-%m-%d')}")
            
            return ((current_close - closest_data['close(point)']) / closest_data['close(point)'] * 100)
        
        return None
    
    def calculate_annualized_return(self, data: pd.DataFrame, period: TimePeriod) -> Optional[float]:
        """计算年化收益率"""
        if data.empty or len(data) < 2:
            return None
        
        if period.is_long_term:
            # 长期业绩：使用几何平均（尾减头）
            start_price = data['close(point)'].iloc[-1]
            end_price = data['close(point)'].iloc[0]
            total_return = (end_price - start_price) / start_price
            date_range = (data['date'].max() - data['date'].min()).days
            years = date_range / 365.25
            return (1 + total_return) ** (1 / years) - 1
        else:
            # 短期业绩：使用算术平均
            daily_return_mean = data['daily_return'].mean()
            return daily_return_mean * 252
    
    def calculate_volatility(self, data: pd.DataFrame) -> Optional[float]:
        """计算年化波动率"""
        if data.empty or len(data) < 5:
            return None
        return data['daily_return'].std() * np.sqrt(252)
    
    def calculate_sharpe_ratio(self, data: pd.DataFrame, annualized_return: Optional[float], risk_free_rate: float) -> Optional[float]:
        """计算夏普比率"""
        if annualized_return is None:
            return None
        
        annualized_volatility = self.calculate_volatility(data)
        if annualized_volatility is None or annualized_volatility == 0:
            return None
        
        return (annualized_return - risk_free_rate) / annualized_volatility
    
    def calculate_metrics(self, df: pd.DataFrame, market: str) -> Optional[Dict[str, Any]]:
        """计算所有指标"""
        if df is None or df.empty:
            logger.error(f"Empty data provided for {market} market")
            return None
        
        try:
            config = self.MARKETS[market]
            
            # 计算每日收益率
            df = self.calculate_daily_returns(df)
            
            # 获取最新数据
            latest_data = df.iloc[0]
            current_date = latest_data['date']
            current_close = latest_data['close(point)']
            current_volume = latest_data['volume(share)']
            
            if self.debug:
                logger.info(f"\nCalculating metrics for {config.name} ({market})")
                logger.info(f"Current Date: {current_date.strftime('%Y-%m-%d')}")
                logger.info(f"Current Close: {current_close:.2f}")
                logger.info(f"Current Volume: {current_volume:,.0f}")
            
            # 计算MoM和YoY变化
            mom_change = self.calculate_period_change(df, 30, current_date, current_close)
            yoy_change = self.calculate_period_change(df, 365, current_date, current_close)
            
            if self.debug:
                logger.info(f"MoM Change: {mom_change:.2f}%" if mom_change is not None else "MoM Change: Insufficient data")
                logger.info(f"YoY Change: {yoy_change:.2f}%" if yoy_change is not None else "YoY Change: Insufficient data")
            
            # 计算5年平均
            five_year_data = self.get_period_data(df, TimePeriod('5-Year', 1825, True), current_date)
            five_year_avg = five_year_data['close(point)'].mean() if not five_year_data.empty else None
            five_year_start_date = five_year_data['date'].min() if not five_year_data.empty else None
            five_year_end_date = five_year_data['date'].max() if not five_year_data.empty else None
            
            if self.debug and not five_year_data.empty:
                logger.info(f"5-Year Average: {five_year_avg:.2f}")
                logger.info(f"5-Year Date Range: {five_year_start_date.strftime('%Y-%m-%d')} to {five_year_end_date.strftime('%Y-%m-%d')}")
                logger.info(f"5-Year Data Points: {len(five_year_data)}")
            
            # 计算各时间周期的指标
            metrics = {
                'current_data': {
                    'current_date': current_date.strftime('%Y-%m-%d'),
                    'close': current_close,
                    'volume': current_volume
                },
                'change_overtime': {
                    'MoM (%)': mom_change,
                    'YoY (%)': yoy_change,
                    '5_year_average': five_year_avg,
                    '5_year_average_dates': f"{five_year_start_date.strftime('%Y-%m-%d')} to {five_year_end_date.strftime('%Y-%m-%d')}" if five_year_start_date and five_year_end_date else None
                },
                'volatility': {},
                'annualized_return': {},
                'sharpe_ratio': {}
            }
            
            # 计算各时间周期的波动率、年化收益率和夏普比率
            for period in self.TIME_PERIODS:
                period_data = self.get_period_data(df, period, current_date)
                
                if self.debug:
                    logger.info(f"\n{period.name} Period Analysis:")
                    logger.info(f"Period Days: {period.days}")
                    logger.info(f"Method: Date Range (Pure Calendar)")
                    logger.info(f"Target Start Date: {(current_date - timedelta(days=period.days)).strftime('%Y-%m-%d')}")
                    logger.info(f"Actual Date Range: {period_data['date'].min().strftime('%Y-%m-%d')} to {period_data['date'].max().strftime('%Y-%m-%d')}")
                    logger.info(f"Actual Days: {(period_data['date'].max() - period_data['date'].min()).days}")
                    logger.info(f"Data Points: {len(period_data)}")
                    logger.info(f"Is Long Term: {period.is_long_term}")
                    logger.info(f"Calculation Method: {'Geometric (End-to-End)' if period.is_long_term else 'Arithmetic (Daily Average)'}")
                
                # 波动率
                volatility = self.calculate_volatility(period_data)
                metrics['volatility'][f'{period.name.lower()}_volatility'] = volatility
                
                if self.debug:
                    logger.info(f"Volatility: {volatility:.2%}" if volatility is not None else "Volatility: Insufficient data")
                
                # 年化收益率
                annualized_return = self.calculate_annualized_return(period_data, period)
                metrics['annualized_return'][f'{period.name.lower()}_annualized_return'] = annualized_return
                
                if self.debug:
                    method = "Geometric" if period.is_long_term else "Arithmetic"
                    logger.info(f"Annualized Return ({method}): {annualized_return:.2%}" if annualized_return is not None else f"Annualized Return ({method}): Insufficient data")
                
                # 夏普比率
                sharpe_ratio = self.calculate_sharpe_ratio(period_data, annualized_return, config.risk_free_rate)
                metrics['sharpe_ratio'][f'{period.name.lower()}_sharpe_ratio'] = sharpe_ratio
                
                if self.debug:
                    logger.info(f"Sharpe Ratio: {sharpe_ratio:.2f}" if sharpe_ratio is not None else "Sharpe Ratio: Insufficient data")
                
                # 显示样本数据
                if self.debug and not period_data.empty:
                    logger.info("Sample Data (5 rows):")
                    sample_indices = np.linspace(0, len(period_data)-1, 5, dtype=int)
                    for idx in sample_indices:
                        row = period_data.iloc[idx]
                        logger.info(f"  {row['date'].strftime('%Y-%m-%d')}: Close={row['close(point)']:.2f}, Return={row['daily_return']:.4f}")
            
            if self.debug:
                logger.info(f"\nFinal metrics for {config.name}:")
                logger.info(f"Current data: {metrics['current_data']}")
                logger.info(f"Change overtime: {metrics['change_overtime']}")
                logger.info(f"Volatility: {metrics['volatility']}")
                logger.info(f"Annualized return: {metrics['annualized_return']}")
                logger.info(f"Sharpe ratio: {metrics['sharpe_ratio']}")
            
            return metrics
            
        except Exception as e:
            logger.error(f"Error calculating metrics for {market} market: {str(e)}")
            return None
    
    def get_all_market_data(self, time_range: int) -> Dict[str, pd.DataFrame]:
        """获取所有市场的数据"""
        if self.debug:
            logger.info(f"Fetching stock index data for {time_range} days")
        
        market_data = {}
        
        for market in self.MARKETS.keys():
            data = self.fetch_market_data(market, time_range)
            market_data[market] = data
        
        # 检查数据可用性
        available_markets = {k: not v.empty for k, v in market_data.items()}
        
        if self.debug:
            logger.info("Data availability:")
            for market, available in available_markets.items():
                logger.info(f"{market}: {'Available' if available else 'Not available'}")
        
        if not any(available_markets.values()):
            logger.error("No valid data was fetched for any index")
            return {}
        
        return market_data
    
    def format_value(self, value: Optional[float], is_percent: bool = False, is_ratio: bool = False) -> str:
        """格式化数值显示"""
        if value is None:
            return "N/A"
        if is_percent:
            return f"{value:.2f}%"
        if is_ratio:
            return f"{value:.2f}"
        return f"{value:.2f}"
    
    def print_summary_table(self, market_data: Dict[str, pd.DataFrame], time_range: int):
        """打印汇总表格"""
        # 临时禁用日志输出，避免在表格中显示日志信息
        original_level = logger.level
        logger.setLevel(logging.ERROR)
        
        try:
            print("\n=== Stock Index Data Summary ===")
            print(f"Data Retrieval Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
            print(f"Data Range: Last {time_range} days")
            
            # 定义表格头部
            headers = [
                "Index", "Latest Date", "Close", "Volume", "MoM (%)", "YoY (%)", 
                "5Y Avg", "5Y Avg Range", "Monthly Vol (Ann.)", "Semi-Annual Vol (Ann.)", 
                "Annual Vol (Ann.)", "5Y Vol (Ann.)", "Monthly Return*", "Semi-Annual Return", 
                "Annual Return", "5Y Return", "Monthly Sharpe", "Semi-Annual Sharpe", 
                "Annual Sharpe", "5Y Sharpe"
            ]
            
            # 打印表头
            header_format = "{:<10} {:<12} {:<10} {:<12} {:<10} {:<10} {:<10} {:<20} {:<12} {:<12} {:<12} {:<12} {:<12} {:<12} {:<12} {:<12} {:<12} {:<12} {:<12}"
            print(header_format.format(*headers))
            print("-" * 240)
            print("Note: Monthly returns use arithmetic mean; Semi-Annual, Annual and 5Y returns use geometric mean for accurate compound effect")
            print()
            
            # 打印每个指数的数据
            for market, data in market_data.items():
                if not data.empty:
                    config = self.MARKETS[market]
                    metrics = self.calculate_metrics(data, market)
                    
                    if metrics:
                        row = [
                            config.name,
                            metrics['current_data']['current_date'],
                            self.format_value(metrics['current_data']['close']),
                            f"{metrics['current_data']['volume']:,.0f}",
                            self.format_value(metrics['change_overtime']['MoM (%)'], is_percent=True),
                            self.format_value(metrics['change_overtime']['YoY (%)'], is_percent=True),
                            self.format_value(metrics['change_overtime']['5_year_average']),
                            metrics['change_overtime']['5_year_average_dates'] or "N/A",
                            self.format_value(metrics['volatility']['monthly_volatility'] * 100, is_percent=True),
                            self.format_value(metrics['volatility']['semi-annual_volatility'] * 100, is_percent=True),
                            self.format_value(metrics['volatility']['annual_volatility'] * 100, is_percent=True),
                            self.format_value(metrics['volatility']['5-year_volatility'] * 100, is_percent=True),
                            self.format_value(metrics['annualized_return']['monthly_annualized_return'], is_percent=True),
                            self.format_value(metrics['annualized_return']['semi-annual_annualized_return'], is_percent=True),
                            self.format_value(metrics['annualized_return']['annual_annualized_return'], is_percent=True),
                            self.format_value(metrics['annualized_return']['5-year_annualized_return'], is_percent=True),
                            self.format_value(metrics['sharpe_ratio']['monthly_sharpe_ratio'], is_ratio=True),
                            self.format_value(metrics['sharpe_ratio']['semi-annual_sharpe_ratio'], is_ratio=True),
                            self.format_value(metrics['sharpe_ratio']['annual_sharpe_ratio'], is_ratio=True),
                            self.format_value(metrics['sharpe_ratio']['5-year_sharpe_ratio'], is_ratio=True)
                        ]
                        print(header_format.format(*row))
        finally:
            # 恢复原始日志级别
            logger.setLevel(original_level)
            
    def export_weekly_report_table_to_csv(self, market_data: Dict[str, pd.DataFrame], filename: str = "output/stock_weekly_report.csv"):
        """ 导出周报表格 """
        from collections import defaultdict
        output = defaultdict(dict)
        
        # 获取指数总市值
        time.sleep(20)
        index_caps = get_all_index_caps()
    
        ordered_markets = ['US', 'CN', 'HK']
    
        market_display = {
            'US': '美国',
            'CN': '中国',
            'HK': '香港',
        }
    
        index_display = {
            'US': 'S&P500',
            'CN': '沪深300',
            'HK': '恒生',
        }
    
        for market_key in ordered_markets:
            stock_metrics_show = market_data.get(market_key)
            if stock_metrics_show is None or stock_metrics_show.empty:
                continue
    
            metrics = self.calculate_metrics(stock_metrics_show, market_key)
            display_name = market_display[market_key]

    
            # 当前数据
            output[("", "收盘值（点）")][display_name] = metrics['current_data']['close']
            cap = index_caps.get(market_key, "-")
            output[("", "总市值")][display_name] = cap
            output[("", "交易量（股）")][display_name] = f"{metrics['current_data']['volume']:,}"
            output[("", "当前时间")][display_name] = metrics['current_data']['current_date']
    
            # 涨跌幅
            output[("", "MoM (%)")][display_name] = self.format_value(metrics['change_overtime']['MoM (%)'], is_percent=True)
            output[("", "YoY (%)")][display_name] = self.format_value(metrics['change_overtime']['YoY (%)'], is_percent=True)
            output[("", "5年均值")][display_name] = self.format_value(metrics['change_overtime']['5_year_average'])
            output[("", "5年均值日期")][display_name] = metrics['change_overtime']['5_year_average_dates']
    
            # 年化
            output[("年化波动率", "短期（月, Ann.）(%)")][display_name] = self.format_value(metrics['volatility']['monthly_volatility'] * 100, is_percent=True)
            output[("年化波动率", "中期（半年, Ann.）(%)")][display_name] = self.format_value(metrics['volatility']['semi-annual_volatility'] * 100, is_percent=True)
            output[("年化波动率", "年度 (%)")][display_name] = self.format_value(metrics['volatility']['annual_volatility'] * 100, is_percent=True)
            output[("年化波动率", "长期（5年, Ann.）(%)")][display_name] = self.format_value(metrics['volatility']['5-year_volatility'] * 100, is_percent=True)
    
            # 年化增长率
            output[("年化增长率", "短期（月, Ann.）(%)")][display_name] = self.format_value(metrics['annualized_return']['monthly_annualized_return'] * 100, is_percent=True)
            output[("年化增长率", "中期（半年, Ann.）(%)")][display_name] = self.format_value(metrics['annualized_return']['semi-annual_annualized_return'] * 100, is_percent=True)
            output[("年化增长率", "年度 (%)")][display_name] = self.format_value(metrics['annualized_return']['annual_annualized_return'] * 100, is_percent=True)
            output[("年化增长率", "长期（5年, Ann.）(%)")][display_name] = self.format_value(metrics['annualized_return']['5-year_annualized_return'] * 100, is_percent=True)
    
            # Sharpe Ratio
            output[("Sharpe Ratio", "短期（月）")][display_name] = self.format_value(metrics['sharpe_ratio']['monthly_sharpe_ratio'], is_ratio=True)
            output[("Sharpe Ratio", "中期（半年）")][display_name] = self.format_value(metrics['sharpe_ratio']['semi-annual_sharpe_ratio'], is_ratio=True)
            output[("Sharpe Ratio", "年度")][display_name] = self.format_value(metrics['sharpe_ratio']['annual_sharpe_ratio'], is_ratio=True)
            output[("Sharpe Ratio", "长期（5年）")][display_name] = self.format_value(metrics['sharpe_ratio']['5-year_sharpe_ratio'], is_ratio=True)
    
            output[("", "PE Ratio")][display_name] = "-"
    
        # 注意这里的 .T 必须加：把指标放到行索引
        stock_metrics_show = pd.DataFrame(output).T
    
        # 拆分多级索引
        stock_metrics_show.index.names = ["类别", "指标"]
        stock_metrics_show.reset_index(inplace=True)
    
        # 添加第一行（指数名）作为标题下第二行
        index_row = ["", "指标说明"] + [index_display[k] for k in ordered_markets if market_display[k] in stock_metrics_show.columns]
        stock_metrics_show.loc[-1] = index_row
        stock_metrics_show.index = stock_metrics_show.index + 1
        stock_metrics_show = stock_metrics_show.sort_index()
        

        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "Weekly Report"
        
        # 写入 DataFrame 到表格
        for r in dataframe_to_rows(stock_metrics_show, index=False, header=True):
            ws.append(r)
        
        # 设置自动换行和居中
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
        
        # 合并“类别”列中相邻的相同值
        category_col = 1  # 第一列是“类别”
        merge_start = 2   # 从第二行开始（跳过标题）
        current_val = ws.cell(row=merge_start, column=category_col).value
        
        for row in range(merge_start + 1, ws.max_row + 2):  # 多一行处理末尾
            cell_val = ws.cell(row=row, column=category_col).value
            if cell_val != current_val:
                if row - merge_start > 1:
                    ws.merge_cells(start_row=merge_start, start_column=category_col,
                                   end_row=row - 1, end_column=category_col)
                merge_start = row
                current_val = cell_val
        
        os.makedirs("output", exist_ok=True)
        output_path = "output/stock_weekly_report.xlsx"
        wb.save(output_path)
        logger.info(f"Weekly report exported to {output_path}")



    def export_summary_table_to_csv(self, market_data: Dict[str, pd.DataFrame], time_range: int, filename: str = "summary.csv"):
        """导出汇总表格为 CSV 文件"""
        rows = []
    
        for market, data in market_data.items():
            if not data.empty:
                config = self.MARKETS[market]
                metrics = self.calculate_metrics(data, market)
    
                if metrics:
                    row = {
                        "Index": config.name,
                        "Latest Date": metrics['current_data']['current_date'],
                        "Close": metrics['current_data']['close'],
                        "Volume": metrics['current_data']['volume'],
                        "MoM (%)": metrics['change_overtime']['MoM (%)'],
                        "YoY (%)": metrics['change_overtime']['YoY (%)'],
                        "5Y Avg": metrics['change_overtime']['5_year_average'],
                        "5Y Avg Range": metrics['change_overtime']['5_year_average_dates'],
                        "Monthly Vol (Ann.)": metrics['volatility']['monthly_volatility'],
                        "Semi-Annual Vol (Ann.)": metrics['volatility']['semi-annual_volatility'],
                        "Annual Vol (Ann.)": metrics['volatility']['annual_volatility'],
                        "5Y Vol (Ann.)": metrics['volatility']['5-year_volatility'],
                        "Monthly Return": metrics['annualized_return']['monthly_annualized_return'],
                        "Semi-Annual Return": metrics['annualized_return']['semi-annual_annualized_return'],
                        "Annual Return": metrics['annualized_return']['annual_annualized_return'],
                        "5Y Return": metrics['annualized_return']['5-year_annualized_return'],
                        "Monthly Sharpe": metrics['sharpe_ratio']['monthly_sharpe_ratio'],
                        "Semi-Annual Sharpe": metrics['sharpe_ratio']['semi-annual_sharpe_ratio'],
                        "Annual Sharpe": metrics['sharpe_ratio']['annual_sharpe_ratio'],
                        "5Y Sharpe": metrics['sharpe_ratio']['5-year_sharpe_ratio']
                    }
                    rows.append(row)
    
        raw_path = 'output/raw_data'
        os.makedirs(raw_path, exist_ok = True)
        
        stock_metrics = pd.DataFrame(rows)
        stock_metrics.to_excel(f'{raw_path}/stock_metrics.xlsx', index=False)     
        
    
    def run_analysis(self, time_range: int = 2920):
        """运行完整分析"""
        try:
            logger.info(f"Starting stock index analysis... (time_range={time_range} days, debug={self.debug})")
            
            # 获取所有市场数据
            market_data = self.get_all_market_data(time_range)
            
            if not market_data:
                logger.error("Failed to fetch stock index data")
                return
            
            # 打印详细分析（如果启用调试模式）
            if self.debug:
                self.print_detailed_analysis(market_data, time_range)
            
            # 打印汇总表格
            self.print_summary_table(market_data, time_range)
            
            # 输出汇总表格
            self.export_summary_table_to_csv(market_data, time_range)
            self.export_weekly_report_table_to_csv(market_data)
            
        except Exception as e:
            logger.error(f"Analysis failed: {str(e)}")

    
    def print_detailed_analysis(self, market_data: Dict[str, pd.DataFrame], time_range: int):
        """打印详细分析（调试模式）"""
        print("\n=== Detailed Stock Index Analysis ===")
        print(f"Data Retrieval Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Data Range: Last {time_range} days")
        
        for market, data in market_data.items():
            if not data.empty:
                config = self.MARKETS[market]
                metrics = self.calculate_metrics(data, market)
                
                if metrics:
                    print(f"\n{config.name} ({market})")
                    print("-" * 30)
                    print(f"Data Range: {data['date'].min().strftime('%Y-%m-%d')} to {data['date'].max().strftime('%Y-%m-%d')}")
                    print(f"Data Points: {len(data)}")
                    
                    print("\nCurrent Data:")
                    print(f"Date: {metrics['current_data']['current_date']}")
                    print(f"Close: {metrics['current_data']['close']:.2f}")
                    print(f"Volume: {metrics['current_data']['volume']:,.0f}")
                    
                    print("\nChanges:")
                    print(f"MoM: {metrics['change_overtime']['MoM (%)']:.2f}%" if metrics['change_overtime']['MoM (%)'] is not None else "MoM: Insufficient data")
                    print(f"YoY: {metrics['change_overtime']['YoY (%)']:.2f}%" if metrics['change_overtime']['YoY (%)'] is not None else "YoY: Insufficient data")
                    print(f"5Y Avg: {metrics['change_overtime']['5_year_average']:.2f}" if metrics['change_overtime']['5_year_average'] is not None else "5Y Avg: Insufficient data")
                    
                    print("\nVolatility (Annualized):")
                    for period in self.TIME_PERIODS:
                        vol_key = f'{period.name.lower()}_volatility'
                        vol_value = metrics['volatility'].get(vol_key)
                        print(f"{period.name}: {vol_value:.2%}" if vol_value is not None else f"{period.name}: Insufficient data")
                    
                    print("\nAnnualized Returns:")
                    for period in self.TIME_PERIODS:
                        ret_key = f'{period.name.lower()}_annualized_return'
                        ret_value = metrics['annualized_return'].get(ret_key)
                        print(f"{period.name}: {ret_value:.2%}" if ret_value is not None else f"{period.name}: Insufficient data")
                    
                    print("\nSharpe Ratios:")
                    for period in self.TIME_PERIODS:
                        sharpe_key = f'{period.name.lower()}_sharpe_ratio'
                        sharpe_value = metrics['sharpe_ratio'].get(sharpe_key)
                        print(f"{period.name}: {sharpe_value:.2f}" if sharpe_value is not None else f"{period.name}: Insufficient data")
            else:
                print(f"\n{config.name} ({market}) - Data fetch failed")

def main(time_range: int = 2920, debug: bool = False):
    """主函数"""
    analyzer = StockIndexAnalyzer(debug=debug)
    analyzer.run_analysis(time_range=time_range)

if __name__ == "__main__":
    # 可以通过修改这里的参数来调整数据范围和调试模式
    main(time_range=2920, debug=False)  # 2920 days ≈ 8 years 