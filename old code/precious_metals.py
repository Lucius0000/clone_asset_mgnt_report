"""
商品与贵金属，输出commodity_indicators_summary.xlsx
"""

import os
from datetime import datetime, timedelta
import yfinance as yf
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment
import math
import logging

# ====== 调试打印与日志 ======
DEBUG = False  # 控制是否在控制台打印；日志文件始终写入
output_folder = "output"
raw_data_folder = os.path.join(output_folder, "raw_data")
os.makedirs(raw_data_folder, exist_ok=True)

LOG_PATH = os.path.join(raw_data_folder, "precious_metals_calculation_steps.log")

logger = logging.getLogger("pm_calc")
logger.setLevel(logging.DEBUG)
# 文件日志
_file_handler = logging.FileHandler(LOG_PATH, encoding="utf-8")
_file_handler.setLevel(logging.DEBUG)
_formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
_file_handler.setFormatter(_formatter)
# 避免重复添加 handler（例如在交互式环境多次运行）
if not any(isinstance(h, logging.FileHandler) and getattr(h, 'baseFilename', '') == _file_handler.baseFilename
           for h in logger.handlers):
    logger.addHandler(_file_handler)
# 控制台日志（可选）
if DEBUG and not any(isinstance(h, logging.StreamHandler) for h in logger.handlers):
    _console = logging.StreamHandler()
    _console.setLevel(logging.DEBUG)
    _console.setFormatter(_formatter)
    logger.addHandler(_console)

def dprint(*args):
    msg = " ".join(str(a) for a in args)
    logger.debug(msg)           # 一律写入日志文件
    if DEBUG:                   # 视需要在控制台打印
        print(msg)

os.environ['http_proxy'] = 'http://127.0.0.1:7890'
os.environ['https_proxy'] = 'http://127.0.0.1:7890'

tickers = ['GLD', 'CL=F', 'HG=F']
data_dict = {}

for symbol in tickers:
    ticker = yf.Ticker(symbol)
    hist = ticker.history(period='6y')
    hist.to_csv(os.path.join(raw_data_folder, f"{symbol}.csv"), encoding='utf-8-sig')
    data_dict[symbol] = hist

def get_period_data(df, offset: pd.DateOffset):
    end_date = df.index[-1]
    start_date = end_date - offset
    return df[df.index >= start_date]

def find_nearest_index_within_window(df, target_date, window=5):
    date_diffs = np.abs(df.index - target_date)
    in_window = date_diffs[date_diffs <= pd.Timedelta(days=window)]
    return df.index[np.argmin(date_diffs)] if not in_window.empty else None

category_map = {
    'GLD': '黄金ETF',
    'CL=F': '原油期货',
    'HG=F': '铜期货'
}

# ====== 复利年化收益率：用首末价区间总收益R，按交易日数N年化 ======
def calculate_annualized_return(data):
    closes = data['Close'].dropna()
    if len(closes) < 2:
        return np.nan
    total_return = closes.iloc[-1] / closes.iloc[0] - 1.0
    returns = closes.pct_change().dropna()
    N = len(returns)               # 交易日数
    if N <= 0:
        return np.nan
    k = 252.0 / N                  # 年化指数
    return (1.0 + total_return) ** k - 1.0

# ====== 年化波动率：std(daily) * sqrt(252) ======
def calculate_annualized_volatility(returns):
    std_daily = returns.dropna().std()
    return std_daily * math.sqrt(252.0)

# ====== Sharpe： (ann_return - 0.045) / ann_vol ======
def calculate_sharpe_ratio(ann_return, ann_vol, rf=0.045):
    if ann_vol is None or pd.isna(ann_vol) or ann_vol == 0:
        return np.nan
    return (ann_return - rf) / ann_vol

def compute_indicators(df, ticker_symbol):
    df = df.copy()
    df['Return'] = df['Close'].pct_change()

    end_date = df.index[-1]
    current_close = float(df.iloc[-1]['Close'])

    # 月中收盘（原逻辑保留）
    if end_date.day >= 15:
        monthly_date = end_date.replace(day=15)
    else:
        prev_month_for_mid = end_date - pd.DateOffset(months=1)
        monthly_date = prev_month_for_mid.replace(day=15)
    idx_mid = find_nearest_index_within_window(df, monthly_date)
    monthly_close = df.loc[idx_mid]['Close'] if idx_mid else np.nan

    close_series_1y = df['Close'].dropna().tail(252)
    percentile = round((close_series_1y <= current_close).sum() / len(close_series_1y) * 100, 2) if len(close_series_1y) > 0 else np.nan
    current_volume = df.iloc[-1]['Volume']

    if ticker_symbol == 'GLD':
        raw_cap = yf.Ticker(ticker_symbol).info.get('marketCap', None)
        market_cap = round(raw_cap / 1e9, 2) if raw_cap else 'n/a'
    else:
        market_cap = 'n/a'

    periods = {
        "短期": pd.DateOffset(months=1),   # 近1个日历月
        "中期": pd.DateOffset(years=1),    # 近1个日历年
        "长期": pd.DateOffset(years=5)     # 近5个日历年
    }

    # 统一窗口
    period_dfs = {label: get_period_data(df, offset) for label, offset in periods.items()}

    volatility = {}
    sharpe = {}
    annualized_return_map = {}

    dprint(f"\n===== {ticker_symbol} | {category_map.get(ticker_symbol, '未知')} =====")
    dprint(f"[收盘] 结束日期: {end_date.date()}  当前收盘: {current_close:.6f}")

    # —— 逐段计算：年化收益/波动率/Sharpe，并打印细节
    for label, period_df in period_dfs.items():
        closes = period_df['Close'].dropna()
        rets = period_df['Return'].dropna()
        if len(closes) < 2 or len(rets) < 1:
            annualized_return_map[label] = np.nan
            volatility[label] = np.nan
            sharpe[label] = np.nan
            dprint(f"[{label}] 数据不足，跳过。")
            continue

        start_price = float(closes.iloc[0])
        end_price_p = float(closes.iloc[-1])
        total_R = end_price_p / start_price - 1.0
        N = len(rets)
        k = 252.0 / N
        ann_ret = (1.0 + total_R) ** k - 1.0
        std_daily = rets.std()
        ann_vol = std_daily * math.sqrt(252.0)
        sr = calculate_sharpe_ratio(ann_ret, ann_vol, rf=0.045)

        annualized_return_map[label] = ann_ret
        volatility[label] = ann_vol
        sharpe[label] = sr

        dprint(f"[{label}] 起止日期: {period_df.index[0].date()} -> {period_df.index[-1].date()}")
        dprint(f"[{label}] 首末价: {start_price:.6f} -> {end_price_p:.6f}  总收益R = {total_R:.6f}")
        dprint(f"[{label}] 交易日数N = {N}  年化指数k = 252/N = {k:.6f}")
        dprint(f"[{label}] 年化收益率 = (1+R)^k - 1 = (1+{total_R:.6f})^{k:.6f} - 1 = {ann_ret:.6f}  -> {ann_ret*100:.4f}%")
        dprint(f"[{label}] 日波动率 = std(daily returns) = {std_daily:.6f}")
        dprint(f"[{label}] 年化波动率 = 日波动率 * sqrt(252) = {std_daily:.6f} * {math.sqrt(252.0):.6f} = {ann_vol:.6f}  -> {ann_vol*100:.4f}%")
        dprint(f"[{label}] Sharpe = (年化收益率 - 0.045) / 年化波动率 = ({ann_ret:.6f} - 0.045) / {ann_vol:.6f} = {sr:.6f}")

    # —— 统一口径的 MoM（短期窗口首末价）
    short_df = period_dfs.get("短期")
    if short_df is not None and len(short_df) >= 2:
        mom_start_date = short_df.index[0]
        mom_end_date = short_df.index[-1]
        mom_start_close = float(short_df['Close'].iloc[0])
        mom_end_close = float(short_df['Close'].iloc[-1])  # == current_close
        mom = (mom_end_close / mom_start_close - 1.0) * 100.0
        dprint(f"[MoM-统一口径] 窗口: {mom_start_date.date()} -> {mom_end_date.date()}")
        dprint(f"[MoM-统一口径] 首末价: {mom_start_close:.6f} -> {mom_end_close:.6f}")
        dprint(f"[MoM-统一口径] = (末/首 - 1)*100 = ({mom_end_close:.6f}/{mom_start_close:.6f} - 1)*100 = {mom:.4f}%")
    else:
        mom = np.nan
        dprint("[MoM-统一口径] 短期窗口数据不足，MoM=NaN")

    # —— 统一口径的 YoY（中期窗口首末价）
    mid_df = period_dfs.get("中期")
    if mid_df is not None and len(mid_df) >= 2:
        yoy_start_date = mid_df.index[0]
        yoy_end_date = mid_df.index[-1]
        yoy_start_close = float(mid_df['Close'].iloc[0])
        yoy_end_close = float(mid_df['Close'].iloc[-1])    # == current_close
        yoy = (yoy_end_close / yoy_start_close - 1.0) * 100.0
        dprint(f"[YoY-统一口径] 窗口: {yoy_start_date.date()} -> {yoy_end_date.date()}")
        dprint(f"[YoY-统一口径] 首末价: {yoy_start_close:.6f} -> {yoy_end_close:.6f}")
        dprint(f"[YoY-统一口径] = (末/首 - 1)*100 = ({yoy_end_close:.6f}/{yoy_start_close:.6f} - 1)*100 = {yoy:.4f}%")
    else:
        yoy = np.nan
        dprint("[YoY-统一口径] 中期窗口数据不足，YoY=NaN")

    category = category_map.get(ticker_symbol, '未知')

    def pct_format(x):
        return f"{round(x, 2)}%" if pd.notnull(x) else ""

    return [
        ticker_symbol,
        category,
        f"{current_close:,.2f}",
        f"{monthly_close:,.2f}" if pd.notnull(monthly_close) else '',
        f"{market_cap:,.0f} B USD" if isinstance(market_cap, (int, float)) else market_cap,
        f"{current_volume:,.0f}" if pd.notnull(current_volume) else '',
        pct_format(mom),
        pct_format(yoy),
        pct_format(percentile),
        pct_format(volatility['短期'] * 100 if pd.notnull(volatility.get('短期')) else np.nan),
        pct_format(volatility['中期'] * 100 if pd.notnull(volatility.get('中期')) else np.nan),
        pct_format(volatility['长期'] * 100 if pd.notnull(volatility.get('长期')) else np.nan),
        f"{sharpe['短期']:,.4f}" if pd.notnull(sharpe.get('短期')) else "",
        f"{sharpe['中期']:,.4f}" if pd.notnull(sharpe.get('中期')) else "",
        f"{sharpe['长期']:,.4f}" if pd.notnull(sharpe.get('长期')) else "",
        pct_format(annualized_return_map["短期"] * 100 if pd.notnull(annualized_return_map.get("短期")) else np.nan),
        pct_format(annualized_return_map["中期"] * 100 if pd.notnull(annualized_return_map.get("中期")) else np.nan),
        pct_format(annualized_return_map["长期"] * 100 if pd.notnull(annualized_return_map.get("长期")) else np.nan)
    ]

data_rows = [compute_indicators(df, symbol) for symbol, df in data_dict.items()]

header_top = [
    "指标", "类别", "当前收盘价 ($)", "月中收盘值 ($)", "总市值 (B $)", "交易量（股\\合约数）", "环比 MoM(%)", "同比 YoY(%)", "百分点位",
    "波动率(%)", "波动率(%)", "波动率(%)",
    "Sharpe Ratio", "Sharpe Ratio", "Sharpe Ratio",
    "收益率年化(%)", "收益率年化(%)", "收益率年化(%)"
]

header_bottom = [
    "", "", "", "", "", "", "", "", "",
    "短期（月）", "中期（年）", "长期（5年）",
    "短期（月）", "中期（年）", "长期（5年）",
    "短期（月）", "中期（年）", "长期（5年）"
]

wb = Workbook()
ws = wb.active
ws.title = "Commodities"

ws.append(header_top)
ws.append(header_bottom)

merge_config = {
    (1, 1): (1, 2),
    (2, 2): (2, 2),
    (3, 3): (3, 2),
    (4, 4): (4, 2),
    (5, 5): (5, 2),
    (6, 6): (6, 2),
    (7, 7): (7, 2),
    (8, 8): (8, 2),
    (9, 9): (9, 2),
    (10, 12): (10, 1),
    (13, 15): (13, 1),
    (16, 18): (16, 1)
}

for (col_start, col_end), (col, row_span) in merge_config.items():
    cell = ws.cell(row=1, column=col_start)
    ws.merge_cells(start_row=1, start_column=col_start, end_row=1 + row_span - 1, end_column=col_end)
    cell.alignment = Alignment(horizontal='center', vertical='center')

for row in data_rows:
    ws.append(row)

for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

output_path = os.path.join(output_folder, "commodity_indicators_summary.xlsx")
wb.save(output_path)

# ====== 关闭日志 handler，确保写盘 ======
for h in list(logger.handlers):
    try:
        h.flush()
    except Exception:
        pass
    try:
        h.close()
    except Exception:
        pass
    logger.removeHandler(h)

