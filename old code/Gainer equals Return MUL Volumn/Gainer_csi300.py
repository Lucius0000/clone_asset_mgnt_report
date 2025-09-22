#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CSI 300 Gainer calculator using ak.index_stock_cons_sina() — supports Config + CLI end-date

流程:
1) 使用 akshare 获取沪深300成分股: ak.index_stock_cons_sina()
   - 仅保留股票名录信息: 'symbol' 列 (如 'sh600000', 'sz001391')
2) 将 symbol 严格映射到 yfinance 代码:
   - 'sh######' -> '######.SS'
   - 'sz######' -> '######.SZ'
3) 为每只股票从 yfinance 下载日线，按自然月口径计算 Gainer:
   - 若指定 --end-date 或在 Config 设置 END_DATE_STR:
       * 下载窗口: start = end_date - WINDOW_DAYS 天, end = end_date + 1 天（包含结束日）
       * 仅使用 <= end_date 的数据计算
   - 未指定 end-date 时使用 period=PERIOD（默认 90d）
   - Return = 最新收盘 - 一个月前收盘
   - Volumn = (month_ago_date, as_of_date] 区间的成交量之和
   - Gainer = Return * Volumn
   - 展示格式:
       * Return、Volumn：千分位
       * Gainer：以十亿为单位显示为 "<#,###> B CNY"；若 |Gainer| < 1 B，则保留两位小数，否则为整数位
4) 汇总所有个股 Gainer，得到“沪深300总 Gainer”
5) 输出:
   - 每股原始日线 CSV: output/raw_data/Gainer_csi300/<yfinance_ticker>.csv
   - 计算结果: output/csi300_gainer_results.xlsx（含 gainer_numeric 数值列）
   - 日志: output/raw_data/csi300_gainer.log  （--debug/-d 或 Config 中 DEBUG_DEFAULT 为 True 时同步打印）
   - 失败清单: output/raw_data/csi300_failures.csv
"""

# ---- 全局配置：代理 & 关闭警告 ----
import os
import warnings
os.environ['http_proxy'] = 'http://127.0.0.1:7890'
os.environ['https_proxy'] = 'http://127.0.0.1:7890'
warnings.filterwarnings("ignore")

# ============================= Config ================================
# 在 Spyder/IDE 中可直接修改；命令行传参会覆盖这里的设置
END_DATE_STR = None          # 例如 "2025-08-15"；None 表示用最新交易日
DEBUG_DEFAULT = False        # True 时在控制台同步打印日志
WINDOW_DAYS = 120            # 指定 end-date 时向前抓取的最小天数
PERIOD = "90d"               # 未指定 end-date 时 yfinance 的 period
INTERVAL = "1d"              # yfinance bar 间隔
DEFAULT_CURRENCY = "CNY"     # 展示用货币单位
# ====================================================================

# ---- 标准库 & 三方库 ----
import re
import time
import argparse
import logging
from typing import Optional, Tuple, Dict, Any, List

import pandas as pd
import akshare as ak
import yfinance as yf
import openpyxl  # noqa: F401
from tqdm import tqdm

# ----------------------------- Config (paths/const) -------------------
OUTPUT_DIR = "output"
RAW_DIR = os.path.join(OUTPUT_DIR, "raw_data")
RAW_SUBDIR = os.path.join(RAW_DIR, "Gainer_csi300")  # 原始CSV保存目录
LOG_PATH = os.path.join(RAW_DIR, "csi300_gainer.log")  # 日志放在 output/raw_data
FAIL_CSV = os.path.join(RAW_DIR, "csi300_failures.csv")  # 失败清单也在 output/raw_data

BILLION = 1_000_000_000
RETURN_DECIMALS = 2        # Return 保留两位小数
RETRY_SLEEP_SECONDS = 3    # 重试前暂停秒数

# -------------------------- Utilities ---------------------------------
def ensure_dirs():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(RAW_DIR, exist_ok=True)
    os.makedirs(RAW_SUBDIR, exist_ok=True)

def setup_logging(debug: bool = False) -> None:
    ensure_dirs()
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    for h in list(logger.handlers):
        logger.removeHandler(h)

    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

    fh = logging.FileHandler(LOG_PATH, mode="w", encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    if debug:
        ch = logging.StreamHandler()
        ch.setFormatter(fmt)
        logger.addHandler(ch)

    logging.info("Log file: %s", LOG_PATH)
    logging.info("Console logging enabled: %s", debug)

def format_thousands_number(x: float, decimals: int = 0) -> str:
    if x is None:
        return ""
    return f"{x:,.{decimals}f}"

def format_gainer_billion(value: float, currency: str) -> str:
    if value is None:
        return ""
    b = value / BILLION
    decimals = 2 if abs(value) < BILLION else 0
    return f"{b:,.{decimals}f} B {currency}".strip()

def sina_symbol_to_yf(symbol: str) -> Optional[Tuple[str, str]]:
    """
    'sh600000'/'sz001391' -> ('600000.SS','600000') / ('001391.SZ','001391')
    """
    s = str(symbol).strip().lower()
    m = re.match(r'^(sh|sz)\s*(\d{6})$', s)
    if not m:
        # 兼容不足6位数字，左补零
        m2 = re.match(r'^(sh|sz)\s*(\d{1,6})$', s)
        if not m2:
            return None
        exch, num = m2.group(1), m2.group(2).zfill(6)
    else:
        exch, num = m.group(1), m.group(2)

    suffix = ".SS" if exch == "sh" else ".SZ"
    return f"{num}{suffix}", num

def fetch_daily_yf(ticker: str, end_dt: Optional[pd.Timestamp] = None) -> Optional[pd.DataFrame]:
    """
    若 end_dt 提供：使用 start = end_dt - WINDOW_DAYS, end = end_dt + 1 天（确保包含结束日）
    否则：使用 period=PERIOD
    """
    logging.info("Downloading from yfinance: %s (end=%s)",
                 ticker, end_dt.date().isoformat() if isinstance(end_dt, pd.Timestamp) else "latest")
    if end_dt is None:
        df = yf.download(ticker, period=PERIOD, interval=INTERVAL, auto_adjust=False, progress=False)
    else:
        start_dt = end_dt - pd.Timedelta(days=WINDOW_DAYS)
        end_inclusive = end_dt + pd.Timedelta(days=1)  # yfinance 的 end 为开区间
        df = yf.download(
            ticker,
            start=start_dt.strftime("%Y-%m-%d"),
            end=end_inclusive.strftime("%Y-%m-%d"),
            interval=INTERVAL,
            auto_adjust=False,
            progress=False,
        )
    if df is None or df.empty or "Close" not in df.columns:
        logging.warning("Empty/invalid data: %s", ticker)
        return None
    df.index = pd.to_datetime(df.index)
    df.sort_index(inplace=True)
    logging.info("Got %d rows for %s (%s → %s)", len(df), ticker, df.index.min().date(), df.index.max().date())
    return df

def save_raw_csv(ticker_used: str, df: pd.DataFrame) -> str:
    path = os.path.join(RAW_SUBDIR, f"{ticker_used}.csv")  # 保存到 output/raw_data/Gainer_csi300
    df.to_csv(path)
    logging.info("Raw saved: %s", path)
    return path

def compute_gainer_from_df(df: pd.DataFrame, currency: str = DEFAULT_CURRENCY, end_dt: Optional[pd.Timestamp] = None):
    """
    返回: (ret_num, vol_sum_num, gainer_num, ret_str, vol_str, gainer_str, as_of_str, month_ago_str)
    - 若 end_dt 提供，则仅使用 df.index <= end_dt 的数据计算 as_of 与回溯窗口
    """
    if end_dt is not None:
        df = df.loc[df.index <= end_dt]
        if df.empty:
            raise ValueError("No data on/before specified end-date")

    as_of_date = df["Close"].dropna().index.max()
    close_latest = float(df.loc[as_of_date, "Close"])

    target_month_ago = as_of_date - pd.DateOffset(months=1)
    month_ago_candidates = df.loc[:target_month_ago].index
    if len(month_ago_candidates) == 0:
        raise ValueError("Insufficient history to find month-ago close.")
    month_ago_date = month_ago_candidates.max()
    close_month_ago = float(df.loc[month_ago_date, "Close"])

    ret = close_latest - close_month_ago

    vol_series = df["Volume"] if "Volume" in df.columns else pd.Series(0, index=df.index)
    vol_sum = float(vol_series.loc[(df.index > month_ago_date) & (df.index <= as_of_date)].fillna(0).sum())

    gainer = ret * vol_sum

    ret_str = format_thousands_number(ret, decimals=RETURN_DECIMALS)
    vol_str = format_thousands_number(vol_sum, decimals=0)
    gainer_str = format_gainer_billion(gainer, currency)

    logging.info(
        "Latest(as_of): %s Close=%.6f | Target month-ago: %s -> chosen: %s Close=%.6f",
        as_of_date.date(), close_latest, target_month_ago.date(), month_ago_date.date(), close_month_ago
    )
    logging.info(
        "Return = %.6f - %.6f = %.6f | Volumn(%s, %s] = %.0f | Gainer = %.6f",
        close_latest, close_month_ago, ret, month_ago_date.date(), as_of_date.date(), vol_sum, gainer
    )

    return (
        ret, vol_sum, gainer,
        ret_str, vol_str, gainer_str,
        as_of_date.date().isoformat(), month_ago_date.date().isoformat()
    )

# ----------------------- Processing helper ----------------------------
def process_one(symbol: str, name_map: Dict[str, str], end_dt: Optional[pd.Timestamp]) -> Tuple[Optional[Dict[str, Any]], Optional[Dict[str, Any]]]:
    """
    处理单只股票：
    - 成功时返回 (row_dict, None)
    - 失败时返回 (None, failure_dict)
    failure_dict: {"symbol","ticker","name","reason"}
    """
    mapping = sina_symbol_to_yf(symbol)
    if mapping is None:
        reason = "symbol 格式无法解析"
        logging.error("%s: %s", symbol, reason)
        return None, {"symbol": symbol, "ticker": "", "name": name_map.get(symbol, ""), "reason": reason}

    yf_ticker, code6 = mapping
    stock_name = name_map.get(symbol, "")

    logging.info("==== [%s %s] yfinance ticker = %s | end_date=%s",
                 code6, stock_name, yf_ticker, end_dt.date().isoformat() if isinstance(end_dt, pd.Timestamp) else "latest")

    df = fetch_daily_yf(yf_ticker, end_dt=end_dt)
    if df is None:
        reason = "yfinance 无有效数据"
        logging.error("Skip %s %s (%s): %s", code6, stock_name, yf_ticker, reason)
        return None, {"symbol": symbol, "ticker": yf_ticker, "name": stock_name, "reason": reason}

    save_raw_csv(yf_ticker, df)

    try:
        ret, vol_sum, gainer, ret_str, vol_str, gainer_str, as_of_str, month_ago_str = compute_gainer_from_df(df, DEFAULT_CURRENCY, end_dt=end_dt)
    except Exception as e:
        reason = f"计算失败: {e}"
        logging.exception("计算失败: %s %s (%s): %s", code6, stock_name, yf_ticker, e)
        return None, {"symbol": symbol, "ticker": yf_ticker, "name": stock_name, "reason": reason}

    row = {
        "symbol": symbol,            # 原始 symbol (如 sh600000)
        "ticker": yf_ticker,         # yfinance 代码 (如 600000.SS)
        "name": stock_name,          # 可为空
        "as_of_date": as_of_str,
        "month_ago_date": month_ago_str,
        "Return": ret_str,           # 字符串（千分位）
        "Volumn": vol_str,           # 字符串（千分位）
        "gainer_numeric": gainer,    # 数值型列（未转 B）
        "Gainer": gainer_str,        # 展示列（B + 货币单位）
    }
    return row, None

# ----------------------- Main workflow --------------------------------
def main():
    parser = argparse.ArgumentParser(description="Compute CSI 300 constituents' Gainer and total via ak.index_stock_cons_sina().")
    # CLI 默认值取自 Config（便于 Spyder/IDE 调整）；命令行传参会覆盖
    parser.add_argument("-d", "--debug", action="store_true", default=DEBUG_DEFAULT,
                        help="Enable console logging (logs are always saved to file).")
    parser.add_argument("--end-date", type=str, default=END_DATE_STR,
                        help="结束日期（YYYY-MM-DD），将使用该日或之前最近交易日作为 as_of。")
    args = parser.parse_args()

    setup_logging(debug=args.debug)

    end_dt = None
    if args.end_date:
        try:
            end_dt = pd.to_datetime(args.end_date).normalize()
        except Exception as e:
            raise SystemExit(f"--end-date 解析失败：{args.end_date}，请用 YYYY-MM-DD 格式。错误：{e}")
    logging.info("Configured end-date: %s", end_dt.date().isoformat() if isinstance(end_dt, pd.Timestamp) else "latest")

    # 1) Constituents via akshare（不打印、不保存成分股Excel）
    cons = ak.index_stock_cons_sina()
    if "symbol" not in cons.columns:
        raise RuntimeError("ak.index_stock_cons_sina() 返回不含 'symbol' 列")
    cons_view = cons[["symbol"]].copy()
    logging.info("Constituents loaded from ak.index_stock_cons_sina(): %d rows", len(cons_view))

    # name 映射（若有）
    name_map: Dict[str, str] = {}
    if "name" in cons.columns:
        name_map = dict(zip(cons["symbol"].astype(str), cons["name"].astype(str)))

    rows: List[Dict[str, Any]] = []
    failures: List[Dict[str, Any]] = []
    total_gainer_numeric = 0.0

    # 2) 第一轮遍历
    for r in tqdm(cons_view.itertuples(index=False), total=len(cons_view), desc="Initial pass for CSI300", ncols=100):
        symbol = str(r.symbol).strip()
        row, fail = process_one(symbol, name_map, end_dt)
        if row:
            rows.append(row)
            total_gainer_numeric += row["gainer_numeric"]
        else:
            failures.append(fail)

    logging.info("First pass completed. successes=%d, failures=%d", len(rows), len(failures))

    # 3) 对失败的进行一轮重试
    if failures:
        logging.info("Sleeping %d seconds before retry...", RETRY_SLEEP_SECONDS)
        time.sleep(RETRY_SLEEP_SECONDS)

        retry_symbols = [f["symbol"] for f in failures]
        failures_after_retry: List[Dict[str, Any]] = []
        successes_on_retry = 0

        for symbol in tqdm(retry_symbols, desc="Retrying failed for CSI300", ncols=100):
            row, fail = process_one(symbol, name_map, end_dt)
            if row:
                rows.append(row)
                total_gainer_numeric += row["gainer_numeric"]
                successes_on_retry += 1
            else:
                failures_after_retry.append(fail)

        failures = failures_after_retry  # 覆盖为“重试后仍失败”的清单
        logging.info("Retry completed. successes_on_retry=%d, remaining_failures=%d", successes_on_retry, len(failures))
    else:
        logging.info("No failures in first pass; skip retry.")

    # 4) 汇总与保存结果（Excel）
    total_gainer_str = format_gainer_billion(total_gainer_numeric, DEFAULT_CURRENCY)
    rows.append({
        "symbol": "沪深300合计",
        "ticker": "",
        "name": "",
        "as_of_date": "",
        "month_ago_date": "",
        "Return": "",
        "Volumn": "",
        "gainer_numeric": total_gainer_numeric,
        "Gainer": total_gainer_str,
    })

    result_df = pd.DataFrame(rows)

    # 排序（仅对个股行按 gainer_numeric 排序，合计行置底）
    if not result_df.empty:
        stock_mask = result_df["symbol"] != "沪深300合计"
        sorted_idx = result_df.loc[stock_mask, :].sort_values("gainer_numeric", ascending=False).index.tolist()
        final_rows = pd.concat([result_df.loc[sorted_idx], result_df.loc[~stock_mask]], ignore_index=True)
    else:
        final_rows = result_df

    out_xlsx = os.path.join(RAW_DIR, "csi300_gainer_results.xlsx")
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        final_rows.to_excel(writer, index=False, sheet_name="Gainer")
        ws = writer.book["Gainer"]
        from openpyxl.utils import get_column_letter

        # 自动列宽
        for col_idx, col_name in enumerate(final_rows.columns, start=1):
            max_len = max(len(str(col_name)), *(len(str(v)) for v in final_rows[col_name].astype(str)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

        # 设置 gainer_numeric 数字格式，避免科学计数法
        if "gainer_numeric" in final_rows.columns:
            cidx = final_rows.columns.get_loc("gainer_numeric") + 1
            for r_idx in range(2, ws.max_row + 1):  # 跳过表头
                ws.cell(row=r_idx, column=cidx).number_format = '#,##0.00'

    logging.info("Results saved: %s (rows=%d)", out_xlsx, len(final_rows))

    # 5) 失败清单：写 CSV + 记录日志
    if failures:
        fails_df = pd.DataFrame(failures, columns=["symbol", "ticker", "name", "reason"])
        fails_df.to_csv(FAIL_CSV, index=False, encoding="utf-8-sig")
        logging.info("Failures saved: %s (rows=%d)", FAIL_CSV, len(failures))
        logging.info("==== FAILED SYMBOLS SUMMARY (after retry) ====")
        for f in failures:
            logging.info("FAIL | symbol=%s | ticker=%s | name=%s | reason=%s",
                         f.get("symbol",""), f.get("ticker",""), f.get("name",""), f.get("reason",""))
    else:
        logging.info("No failures remaining after retry.")

    # print("\n已保存结果到：", out_xlsx)
    # print("失败清单（如有）：", FAIL_CSV if failures else "无")
    # print("日志文件：", LOG_PATH)
    # print("原始数据目录：", RAW_SUBDIR)

if __name__ == "__main__":
    main()
