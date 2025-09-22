#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
HSI (Hang Seng Index) Gainer calculator — supports Config + CLI end-date

使用方式
1) 在 Spyder/IDE 里：直接修改 Config 区域的 END_DATE_STR （如 "2025-08-15"）
2) 命令行：python hsi_gainer.py --end-date 2025-08-15
   - 命令行传参优先级高于 Config

输出：
- 结果：output/hsi_gainer_results.xlsx
- 日志：output/raw_data/hsi_gainer.log
- 失败清单：output/raw_data/hsi_failures.csv
- 原始日线：output/raw_data/Gainer_HSI/<TICKER>.csv
- 进度条：tqdm；失败后重试 1 轮（短暂停顿）
"""

# ---- 代理 & 忽略告警（按你的做法） ----
import os
import warnings
os.environ['http_proxy'] = 'http://127.0.0.1:7890'
os.environ['https_proxy'] = 'http://127.0.0.1:7890'
warnings.filterwarnings("ignore")

# ============================= Config ================================
# 在 Spyder 中直接改这里即可；命令行传参（--end-date）会覆盖此处设置
END_DATE_STR = None          # 例如: "2025-08-15"；None 表示用最新交易日
DEBUG_DEFAULT = False        # True 时在控制台同步打印日志
WINDOW_DAYS = 120            # 指定 end-date 时向前抓取的最小天数
RETRY_SLEEP_SECONDS = 3      # 重试前暂停秒数
PERIOD = "90d"               # 未指定 end-date 时使用的 yfinance period
INTERVAL = "1d"              # yfinance bar 间隔
DEFAULT_CURRENCY = "HKD"     # 展示用货币单位
# ====================================================================

import re
import glob
import time
import argparse
import logging
from typing import Optional, Tuple, Dict, Any, List

import pandas as pd
import yfinance as yf
import openpyxl  # noqa: F401
from tqdm import tqdm

# ----------------------------- Paths ---------------------------------
OUTPUT_DIR = "output"
RAW_DIR = os.path.join(OUTPUT_DIR, "raw_data")
RAW_SUBDIR = os.path.join(RAW_DIR, "Gainer_HSI")  # 原始CSV保存目录
LOG_PATH = os.path.join(RAW_DIR, "hsi_gainer.log")  # 日志
FAIL_CSV = os.path.join(RAW_DIR, "hsi_failures.csv")  # 失败清单

DATA_DIR = "data"
AA_PATTERN = os.path.join(DATA_DIR, "AASTOCKS_Export*.xlsx")

BILLION = 1_000_000_000
RETURN_DECIMALS = 2

# -------------------------- Utilities ---------------------------------
def ensure_dirs():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(RAW_DIR, exist_ok=True)
    os.makedirs(RAW_SUBDIR, exist_ok=True)
    os.makedirs(DATA_DIR, exist_ok=True)

def setup_logging(debug: bool = False) -> None:
    ensure_dirs()
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    for h in list(logger.handlers):
        logger.removeHandler(h)

    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    fh = logging.FileHandler(LOG_PATH, mode="w", encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    if debug:
        ch = logging.StreamHandler()
        ch.setFormatter(fmt)
        logger.addHandler(ch)

    logging.info("Log file: %s", LOG_PATH)
    logging.info("Console logging enabled: %s", debug)

def format_thousands_number(x: float, decimals: int = 0) -> str:
    if x is None:
        return ""
    return f"{x:,.{decimals}f}"

def format_gainer_billion(value: float, currency: str) -> str:
    if value is None:
        return ""
    b = value / BILLION
    decimals = 2 if abs(value) < BILLION else 0
    return f"{b:,.{decimals}f} B {currency}".strip()

def pick_latest_aastocks_file(pattern: str) -> str:
    files = glob.glob(pattern)
    if not files:
        raise FileNotFoundError("未找到匹配的 AASTOCKS_Export*.xlsx 文件")
    return max(files, key=os.path.getmtime)

def load_hsi_constituents() -> pd.DataFrame:
    path = pick_latest_aastocks_file(AA_PATTERN)
    df = pd.read_excel(path)
    if "代號" not in df.columns:
        raise RuntimeError("成分股文件缺少『代號』列")
    logging.info("Loaded HSI constituents from %s: %d rows", path, len(df))
    return df

def hk_symbol_to_yf(code: str) -> Optional[str]:
    """
    输入：AASTOCKS『代號』，如 '00001.HK' 或 '01398.HK'
    规则：删除首位数字，并规范成 4 位 + '.HK'
      00001.HK -> 0001.HK
      01398.HK -> 1398.HK
      若不足 4 位则左补零： 83.HK -> 0083.HK
    """
    s = str(code).upper().strip()
    m = re.match(r'^(\d+)\.HK$', s)
    if not m:
        return None
    num = m.group(1)
    if len(num) >= 5:
        num = num[1:]  # 删除首位
    # 规范到 4 位
    if len(num) < 4:
        num = num.zfill(4)
    elif len(num) > 4:
        num = num[-4:]
    return f"{num}.HK"

def fetch_daily_yf(ticker: str, end_dt: Optional[pd.Timestamp] = None) -> Optional[pd.DataFrame]:
    """
    若 end_dt 提供：使用 start = end_dt - WINDOW_DAYS, end = end_dt + 1 天（确保包含结束日）
    否则：使用 period=PERIOD
    """
    logging.info("Downloading from yfinance: %s (end=%s)", ticker, end_dt.date().isoformat() if isinstance(end_dt, pd.Timestamp) else "latest")
    if end_dt is None:
        df = yf.download(ticker, period=PERIOD, interval=INTERVAL, auto_adjust=False, progress=False)
    else:
        start_dt = end_dt - pd.Timedelta(days=WINDOW_DAYS)
        end_inclusive = end_dt + pd.Timedelta(days=1)  # yfinance 的 end 为开区间
        df = yf.download(
            ticker,
            start=start_dt.strftime("%Y-%m-%d"),
            end=end_inclusive.strftime("%Y-%m-%d"),
            interval=INTERVAL,
            auto_adjust=False,
            progress=False,
        )
    if df is None or df.empty or "Close" not in df.columns:
        logging.warning("Empty/invalid data: %s", ticker)
        return None
    df.index = pd.to_datetime(df.index)
    df.sort_index(inplace=True)
    logging.info("Got %d rows for %s (%s → %s)", len(df), ticker, df.index.min().date(), df.index.max().date())
    return df

def save_raw_csv(ticker_used: str, df: pd.DataFrame) -> str:
    path = os.path.join(RAW_SUBDIR, f"{ticker_used}.csv")
    df.to_csv(path)
    logging.info("Raw saved: %s", path)
    return path

def compute_gainer_from_df(df: pd.DataFrame, currency: str, end_dt: Optional[pd.Timestamp] = None):
    """
    返回: (ret_num, vol_sum_num, gainer_num, ret_str, vol_str, gainer_str, as_of_str, month_ago_str)
    - 若 end_dt 提供，则仅使用 df.index <= end_dt 的数据计算 as_of 与回溯窗口
    """
    if end_dt is not None:
        df = df.loc[df.index <= end_dt]
        if df.empty:
            raise ValueError("No data on/before specified end-date")

    as_of_date = df["Close"].dropna().index.max()
    close_latest = float(df.loc[as_of_date, "Close"])

    target_month_ago = as_of_date - pd.DateOffset(months=1)
    month_ago_candidates = df.loc[:target_month_ago].index
    if len(month_ago_candidates) == 0:
        raise ValueError("Insufficient history to find month-ago close.")
    month_ago_date = month_ago_candidates.max()
    close_month_ago = float(df.loc[month_ago_date, "Close"])

    ret = close_latest - close_month_ago

    vol_series = df["Volume"] if "Volume" in df.columns else pd.Series(0, index=df.index)
    vol_sum = float(vol_series.loc[(df.index > month_ago_date) & (df.index <= as_of_date)].fillna(0).sum())

    gainer = ret * vol_sum

    ret_str = format_thousands_number(ret, decimals=RETURN_DECIMALS)
    vol_str = format_thousands_number(vol_sum, decimals=0)
    gainer_str = format_gainer_billion(gainer, currency)

    logging.info(
        "Latest(as_of): %s Close=%.6f | Target month-ago: %s -> chosen: %s Close=%.6f",
        as_of_date.date(), close_latest, target_month_ago.date(), month_ago_date.date(), close_month_ago
    )
    logging.info(
        "Return = %.6f - %.6f = %.6f | Volumn(%s, %s] = %.0f | Gainer = %.6f",
        close_latest, close_month_ago, ret, month_ago_date.date(), as_of_date.date(), vol_sum, gainer
    )

    return (
        ret, vol_sum, gainer,
        ret_str, vol_str, gainer_str,
        as_of_date.date().isoformat(), month_ago_date.date().isoformat()
    )

# ----------------------- Processing helper ----------------------------
def process_one(code: str, name: str, end_dt: Optional[pd.Timestamp]) -> Tuple[Optional[Dict[str, Any]], Optional[Dict[str, Any]]]:
    """
    处理单只股票：
    - 成功时返回 (row_dict, None)
    - 失败时返回 (None, failure_dict)
    failure_dict: {"symbol","ticker","name","reason"}
    """
    yf_ticker = hk_symbol_to_yf(code)
    if yf_ticker is None:
        reason = "代码格式无法解析"
        logging.error("%s: %s", code, reason)
        return None, {"symbol": code, "ticker": "", "name": name or "", "reason": reason}

    logging.info("==== [%s %s] yfinance ticker = %s | end_date=%s", code, name, yf_ticker, end_dt.date().isoformat() if isinstance(end_dt, pd.Timestamp) else "latest")

    df = fetch_daily_yf(yf_ticker, end_dt=end_dt)
    if df is None:
        reason = "yfinance 无有效数据"
        logging.error("Skip %s %s (%s): %s", code, name, yf_ticker, reason)
        return None, {"symbol": code, "ticker": yf_ticker, "name": name or "", "reason": reason}

    save_raw_csv(yf_ticker, df)

    try:
        ret, vol_sum, gainer, ret_str, vol_str, gainer_str, as_of_str, month_ago_str = compute_gainer_from_df(df, DEFAULT_CURRENCY, end_dt=end_dt)
    except Exception as e:
        reason = f"计算失败: {e}"
        logging.exception("计算失败: %s %s (%s): %s", code, name, yf_ticker, e)
        return None, {"symbol": code, "ticker": yf_ticker, "name": name or "", "reason": reason}

    row = {
        "symbol": code,             # 原始『代號』，如 00001.HK / 01398.HK
        "ticker": yf_ticker,        # yfinance 代码，如 0001.HK / 1398.HK
        "name": name or "",
        "as_of_date": as_of_str,
        "month_ago_date": month_ago_str,
        "Return": ret_str,          # 字符串（千分位）
        "Volumn": vol_str,          # 字符串（千分位）
        "gainer_numeric": gainer,   # 数值型列（未转 B）
        "Gainer": gainer_str,       # 展示列（B + 货币单位）
    }
    return row, None

# ----------------------- Main workflow --------------------------------
def main():
    parser = argparse.ArgumentParser(description="Compute HSI constituents' Gainer from AASTOCKS export.")
    # CLI 默认值取自 Config，便于 Spyder/IDE 调整；命令行传参则覆盖
    parser.add_argument("-d", "--debug", action="store_true", default=DEBUG_DEFAULT,
                        help="Enable console logging (logs are always saved to file).")
    parser.add_argument("--end-date", type=str, default=END_DATE_STR,
                        help="结束日期（YYYY-MM-DD），将使用该日或之前最近交易日作为 as_of。")
    args = parser.parse_args()

    setup_logging(debug=args.debug)

    end_dt = None
    if args.end_date:
        try:
            end_dt = pd.to_datetime(args.end_date).normalize()
        except Exception as e:
            raise SystemExit(f"--end-date 解析失败：{args.end_date}，请用 YYYY-MM-DD 格式。错误：{e}")
    logging.info("Configured end-date: %s", end_dt.date().isoformat() if isinstance(end_dt, pd.Timestamp) else "latest")

    # 1) Constituents
    df = load_hsi_constituents()
    name_col = "名稱" if "名稱" in df.columns else ("名称" if "名称" in df.columns else None)
    code_series = df["代號"].astype(str)

    rows: List[Dict[str, Any]] = []
    failures: List[Dict[str, Any]] = []
    total_gainer_numeric = 0.0

    # 2) 第一轮遍历
    for code in tqdm(code_series, desc="Initial pass for HSI", ncols=100):
        # 取名称（若存在）——按同一行索引
        name = ""
        if name_col:
            try:
                row_idx = df.index[code_series.eq(code)][0]
                name = str(df.loc[row_idx, name_col])
            except Exception:
                name = ""
        row, fail = process_one(str(code), name, end_dt)
        if row:
            rows.append(row)
            total_gainer_numeric += row["gainer_numeric"]
        else:
            failures.append(fail)

    logging.info("First pass completed. successes=%d, failures=%d", len(rows), len(failures))

    # 3) 重试失败的（一次）
    if failures:
        logging.info("Sleeping %d seconds before retry...", RETRY_SLEEP_SECONDS)
        time.sleep(RETRY_SLEEP_SECONDS)

        retry_list = failures
        failures_after_retry: List[Dict[str, Any]] = []
        successes_on_retry = 0

        for item in tqdm(retry_list, desc="Retrying failed for HSI", ncols=100):
            code = item["symbol"]
            name = item.get("name", "")
            row, fail = process_one(code, name, end_dt)
            if row:
                rows.append(row)
                total_gainer_numeric += row["gainer_numeric"]
                successes_on_retry += 1
            else:
                failures_after_retry.append(fail)

        failures = failures_after_retry
        logging.info("Retry completed. successes_on_retry=%d, remaining_failures=%d", successes_on_retry, len(failures))
    else:
        logging.info("No failures in first pass; skip retry.")

    # 4) 汇总与保存结果
    total_gainer_str = format_gainer_billion(total_gainer_numeric, DEFAULT_CURRENCY)
    rows.append({
        "symbol": "HSI Total",
        "ticker": "",
        "name": "",
        "as_of_date": "",
        "month_ago_date": "",
        "Return": "",
        "Volumn": "",
        "gainer_numeric": total_gainer_numeric,
        "Gainer": total_gainer_str,
    })

    result_df = pd.DataFrame(rows)

    # 排序（仅对个股按 gainer_numeric 排序，合计行置底）
    if not result_df.empty:
        stock_mask = result_df["symbol"] != "HSI Total"
        sorted_idx = result_df.loc[stock_mask, :].sort_values("gainer_numeric", ascending=False).index.tolist()
        final_rows = pd.concat([result_df.loc[sorted_idx], result_df.loc[~stock_mask]], ignore_index=True)
    else:
        final_rows = result_df

    out_xlsx = os.path.join(RAW_DIR, "hsi_gainer_results.xlsx")
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        final_rows.to_excel(writer, index=False, sheet_name="Gainer")
        ws = writer.book["Gainer"]
        from openpyxl.utils import get_column_letter
        # 自动列宽
        for col_idx, col_name in enumerate(final_rows.columns, start=1):
            max_len = max(len(str(col_name)), *(len(str(v)) for v in final_rows[col_name].astype(str)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)
        # 设置 gainer_numeric 数字格式，避免科学计数法
        if "gainer_numeric" in final_rows.columns:
            cidx = final_rows.columns.get_loc("gainer_numeric") + 1
            for r_idx in range(2, ws.max_row + 1):
                ws.cell(row=r_idx, column=cidx).number_format = '#,##0.00'

    logging.info("Results saved: %s (rows=%d)", out_xlsx, len(final_rows))

    # 5) 失败清单
    if failures:
        fails_df = pd.DataFrame(failures, columns=["symbol", "ticker", "name", "reason"])
        fails_df.to_csv(FAIL_CSV, index=False, encoding="utf-8-sig")
        logging.info("Failures saved: %s (rows=%d)", FAIL_CSV, len(failures))
        logging.info("==== FAILED SYMBOLS SUMMARY (after retry) ====")
        for f in failures:
            logging.info("FAIL | symbol=%s | ticker=%s | name=%s | reason=%s",
                         f.get("symbol",""), f.get("ticker",""), f.get("name",""), f.get("reason",""))
    else:
        logging.info("No failures remaining after retry.")

    # print("\n已保存结果到：", out_xlsx)
    # print("失败清单（如有）：", FAIL_CSV if failures else "无")
    # print("日志文件：", LOG_PATH)
    # print("原始数据目录：", RAW_SUBDIR)

if __name__ == "__main__":
    main()
