#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
S&P 500 Gainer calculator — supports Config + CLI end-date

数据来源（成分股）:
- https://datahub.io/core/s-and-p-500-companies/r/constituents.csv
  若在线失败，回退读取本地 data/constituents.csv

算法:
- 对每只成分股从 yfinance 下载日线
- 当指定 --end-date 或在 Config 中设置 END_DATE_STR 时：
    * 下载窗口: start = end_date - WINDOW_DAYS天, end = end_date + 1天
    * 只使用 <= end_date 的数据计算
- 未指定 end-date 时使用 period=PERIOD（默认 90d）
- Return = 最新收盘 - 一个月前收盘（自然月回退，选目标日或之前最近交易日）
- Volumn = (month_ago_date, as_of_date] 区间的成交量之和
- gainer_numeric = Return * Volumn
- Gainer（展示列）= "<千分位十亿数> B USD"；若 |gainer_numeric| < 1B 显示 2 位小数，否则 0 位

输出:
- 日志: output/raw_data/sp500_gainer.log
- 失败清单: output/raw_data/sp500_failures.csv
- 原始日线CSV: output/raw_data/Gainer_S&P500/<TICKER>.csv
- 结果Excel: output/sp500_gainer_results.xlsx
- 进度条: tqdm
"""

# ---- 代理 & 忽略告警 ----
import os
import warnings
os.environ['http_proxy'] = 'http://127.0.0.1:7890'
os.environ['https_proxy'] = 'http://127.0.0.1:7890'
warnings.filterwarnings("ignore")

# ============================== Config ===============================
# 在 Spyder/IDE 中可直接修改此处；命令行传参会覆盖这里的设置
END_DATE_STR = "2025-08-15"          # 例如 "2025-08-15"；None 表示用最新交易日
DEBUG_DEFAULT = False        # True 时在控制台同步打印日志
WINDOW_DAYS = 120            # 指定 end-date 时向前抓取的最小天数
PERIOD = "90d"               # 未指定 end-date 时的 yfinance period
INTERVAL = "1d"              # yfinance bar 间隔
DEFAULT_CURRENCY = "USD"     # 展示用货币单位
# ====================================================================

# ---- 标准库 & 三方库 ----
import re
import time
import argparse
import logging
from io import StringIO
from typing import Optional, Tuple, Dict, Any, List

import requests
import pandas as pd
import yfinance as yf
import openpyxl  # noqa: F401
from tqdm import tqdm

# ----------------------------- Paths ---------------------------------
OUTPUT_DIR = "output"
RAW_DIR = os.path.join(OUTPUT_DIR, "raw_data")
RAW_SUBDIR = os.path.join(RAW_DIR, "Gainer_S&P500")  # 原始CSV保存目录
LOG_PATH = os.path.join(RAW_DIR, "sp500_gainer.log")  # 日志放在 output/raw_data
FAIL_CSV = os.path.join(RAW_DIR, "sp500_failures.csv")  # 失败清单也在 output/raw_data

DATA_DIR = "data"             # 成分股本地备份目录
LOCAL_BACKUP = os.path.join(DATA_DIR, "constituents.csv")
SP500_URL = "https://datahub.io/core/s-and-p-500-companies/r/constituents.csv"

BILLION = 1_000_000_000
RETURN_DECIMALS = 2
RETRY_SLEEP_SECONDS = 3  # 重试前暂停秒数

# -------------------------- Utilities ---------------------------------
def ensure_dirs():
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(RAW_DIR, exist_ok=True)
    os.makedirs(RAW_SUBDIR, exist_ok=True)
    os.makedirs(DATA_DIR, exist_ok=True)

def setup_logging(debug: bool = False) -> None:
    ensure_dirs()
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)
    for h in list(logger.handlers):
        logger.removeHandler(h)

    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    fh = logging.FileHandler(LOG_PATH, mode="w", encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    if debug:
        ch = logging.StreamHandler()
        ch.setFormatter(fmt)
        logger.addHandler(ch)

    logging.info("Log file: %s", LOG_PATH)
    logging.info("Console logging enabled: %s", debug)

def format_thousands_number(x: float, decimals: int = 0) -> str:
    if x is None:
        return ""
    return f"{x:,.{decimals}f}"

def format_gainer_billion(value: float, currency: str) -> str:
    if value is None:
        return ""
    b = value / BILLION
    decimals = 2 if abs(value) < BILLION else 0
    return f"{b:,.{decimals}f} B {currency}".strip()

def get_sp500_df(url: str, backup_path: str) -> pd.DataFrame:
    """优先在线获取；失败则读本地备份。"""
    try:
        resp = requests.get(url, timeout=10)
        resp.raise_for_status()
        with open(backup_path, "w", encoding="utf-8") as f:
            f.write(resp.text)
        df = pd.read_csv(StringIO(resp.text))
        logging.info("Fetched S&P500 constituents online: %d rows", len(df))
        return df
    except Exception as e:
        logging.warning("Online fetch failed: %s; trying local backup: %s", e, backup_path)
        if os.path.exists(backup_path):
            df = pd.read_csv(backup_path)
            logging.info("Loaded S&P500 constituents from local backup: %d rows", len(df))
            return df
        raise RuntimeError("无法获取 S&P500 成分股列表（在线与本地备份均失败）")

def yf_symbol_from_symbol(symbol: str) -> str:
    """
    DataHub 的 Symbol 一般已可直接用于 yfinance，但为兼容 BRK.B/BF.B 等，替换 '.' 为 '-'
    """
    s = str(symbol).upper().strip()
    return s.replace(".", "-")

def fetch_daily_yf(ticker: str, end_dt: Optional[pd.Timestamp] = None) -> Optional[pd.DataFrame]:
    """
    若 end_dt 提供：使用 start = end_dt - WINDOW_DAYS, end = end_dt + 1 天（确保包含结束日）
    否则：使用 period=PERIOD
    """
    logging.info("Downloading from yfinance: %s (end=%s)", ticker, end_dt.date().isoformat() if isinstance(end_dt, pd.Timestamp) else "latest")
    if end_dt is None:
        df = yf.download(ticker, period=PERIOD, interval=INTERVAL, auto_adjust=False, progress=False)
    else:
        start_dt = end_dt - pd.Timedelta(days=WINDOW_DAYS)
        end_inclusive = end_dt + pd.Timedelta(days=1)  # yfinance 的 end 为开区间
        df = yf.download(
            ticker,
            start=start_dt.strftime("%Y-%m-%d"),
            end=end_inclusive.strftime("%Y-%m-%d"),
            interval=INTERVAL,
            auto_adjust=False,
            progress=False,
        )
    if df is None or df.empty or "Close" not in df.columns:
        logging.warning("Empty/invalid data: %s", ticker)
        return None
    df.index = pd.to_datetime(df.index)
    df.sort_index(inplace=True)
    logging.info("Got %d rows for %s (%s → %s)", len(df), ticker, df.index.min().date(), df.index.max().date())
    return df

def save_raw_csv(ticker_used: str, df: pd.DataFrame) -> str:
    path = os.path.join(RAW_SUBDIR, f"{ticker_used}.csv")
    df.to_csv(path)
    logging.info("Raw saved: %s", path)
    return path

def compute_gainer_from_df(df: pd.DataFrame, currency: str = DEFAULT_CURRENCY, end_dt: Optional[pd.Timestamp] = None):
    """
    返回: (ret_num, vol_sum_num, gainer_num, ret_str, vol_str, gainer_str, as_of_str, month_ago_str)
    - 若 end_dt 提供，则仅使用 df.index <= end_dt 的数据计算 as_of 与回溯窗口
    """
    if end_dt is not None:
        df = df.loc[df.index <= end_dt]
        if df.empty:
            raise ValueError("No data on/before specified end-date")

    as_of_date = df["Close"].dropna().index.max()
    close_latest = float(df.loc[as_of_date, "Close"])

    target_month_ago = as_of_date - pd.DateOffset(months=1)
    month_ago_candidates = df.loc[:target_month_ago].index
    if len(month_ago_candidates) == 0:
        raise ValueError("Insufficient history to find month-ago close.")
    month_ago_date = month_ago_candidates.max()
    close_month_ago = float(df.loc[month_ago_date, "Close"])

    ret = close_latest - close_month_ago

    vol_series = df["Volume"] if "Volume" in df.columns else pd.Series(0, index=df.index)
    vol_sum = float(vol_series.loc[(df.index > month_ago_date) & (df.index <= as_of_date)].fillna(0).sum())

    gainer = ret * vol_sum

    ret_str = format_thousands_number(ret, decimals=RETURN_DECIMALS)
    vol_str = format_thousands_number(vol_sum, decimals=0)
    gainer_str = format_gainer_billion(gainer, currency)

    logging.info(
        "Latest: %s Close=%.6f | Target month-ago: %s -> chosen: %s Close=%.6f",
        as_of_date.date(), close_latest, target_month_ago.date(), month_ago_date.date(), close_month_ago
    )
    logging.info(
        "Return = %.6f - %.6f = %.6f | Volumn(%s, %s] = %.0f | Gainer = %.6f",
        close_latest, close_month_ago, ret, month_ago_date.date(), as_of_date.date(), vol_sum, gainer
    )

    return (
        ret, vol_sum, gainer,
        ret_str, vol_str, gainer_str,
        as_of_date.date().isoformat(), month_ago_date.date().isoformat()
    )

# ----------------------- Processing helper ----------------------------
def process_one(symbol: str, name_map: Dict[str, str], end_dt: Optional[pd.Timestamp]) -> Tuple[Optional[Dict[str, Any]], Optional[Dict[str, Any]]]:
    """
    处理单只股票：
    - 成功时返回 (row_dict, None)
    - 失败时返回 (None, failure_dict)
    failure_dict: {"symbol","ticker","name","reason"}
    """
    yf_ticker = yf_symbol_from_symbol(symbol)
    stock_name = name_map.get(symbol, "")

    logging.info("==== [%s %s] yfinance ticker = %s | end_date=%s", symbol, stock_name, yf_ticker, end_dt.date().isoformat() if isinstance(end_dt, pd.Timestamp) else "latest")

    df = fetch_daily_yf(yf_ticker, end_dt=end_dt)
    if df is None:
        reason = "yfinance 无有效数据"
        logging.error("Skip %s %s (%s): %s", symbol, stock_name, yf_ticker, reason)
        return None, {"symbol": symbol, "ticker": yf_ticker, "name": stock_name, "reason": reason}

    save_raw_csv(yf_ticker, df)

    try:
        ret, vol_sum, gainer, ret_str, vol_str, gainer_str, as_of_str, month_ago_str = compute_gainer_from_df(df, DEFAULT_CURRENCY, end_dt=end_dt)
    except Exception as e:
        reason = f"计算失败: {e}"
        logging.exception("计算失败: %s %s (%s): %s", symbol, stock_name, yf_ticker, e)
        return None, {"symbol": symbol, "ticker": yf_ticker, "name": stock_name, "reason": reason}

    row = {
        "symbol": symbol,            # 原始 symbol（如 AAPL, BRK.B）
        "ticker": yf_ticker,         # yfinance 代码（如 AAPL, BRK-B）
        "name": stock_name,          # 可为空
        "as_of_date": as_of_str,
        "month_ago_date": month_ago_str,
        "Return": ret_str,           # 字符串（千分位）
        "Volumn": vol_str,           # 字符串（千分位）
        "gainer_numeric": gainer,    # 数值型列（未转 B）
        "Gainer": gainer_str,        # 展示列（B + 货币单位）
    }
    return row, None

# ----------------------- Main workflow --------------------------------
def main():
    parser = argparse.ArgumentParser(description="Compute S&P 500 constituents' Gainer.")
    # CLI 默认值取自 Config（便于 Spyder/IDE 调整）；命令行传参会覆盖
    parser.add_argument("-d", "--debug", action="store_true", default=DEBUG_DEFAULT,
                        help="Enable console logging (logs are always saved to file).")
    parser.add_argument("--end-date", type=str, default=END_DATE_STR,
                        help="结束日期（YYYY-MM-DD），将使用该日或之前最近交易日作为 as_of。")
    args = parser.parse_args()

    setup_logging(debug=args.debug)

    end_dt = None
    if args.end_date:
        try:
            end_dt = pd.to_datetime(args.end_date).normalize()
        except Exception as e:
            raise SystemExit(f"--end-date 解析失败：{args.end_date}，请用 YYYY-MM-DD 格式。错误：{e}")
    logging.info("Configured end-date: %s", end_dt.date().isoformat() if isinstance(end_dt, pd.Timestamp) else "latest")

    # 1) Constituents
    df_cons = get_sp500_df(SP500_URL, LOCAL_BACKUP)
    df_cons.columns = df_cons.columns.str.upper()
    if "SYMBOL" not in df_cons.columns:
        raise RuntimeError("成分股列表缺少 SYMBOL 列")
    symbols_list: List[str] = sorted(set(df_cons["SYMBOL"].astype(str).str.upper().str.strip().tolist()))
    logging.info("S&P500 constituents loaded: %d unique symbols", len(symbols_list))

    # name 映射（若有 NAME 列）
    name_map: Dict[str, str] = {}
    if "NAME" in df_cons.columns:
        name_map = dict(zip(df_cons["SYMBOL"].astype(str).str.upper().str.strip(),
                            df_cons["NAME"].astype(str)))
    elif "SECURITY" in df_cons.columns:  # 备选字段名
        name_map = dict(zip(df_cons["SYMBOL"].astype(str).str.upper().str.strip(),
                            df_cons["SECURITY"].astype(str)))

    rows: List[Dict[str, Any]] = []
    failures: List[Dict[str, Any]] = []
    total_gainer_numeric = 0.0

    # 2) 第一轮遍历
    for symbol in tqdm(symbols_list, desc="Initial pass for S&P500", ncols=100):
        row, fail = process_one(symbol, name_map, end_dt)
        if row:
            rows.append(row)
            total_gainer_numeric += row["gainer_numeric"]
        else:
            failures.append(fail)

    logging.info("First pass completed. successes=%d, failures=%d", len(rows), len(failures))

    # 3) 重试失败的
    if failures:
        logging.info("Sleeping %d seconds before retry...", RETRY_SLEEP_SECONDS)
        time.sleep(RETRY_SLEEP_SECONDS)

        retry_symbols = [f["symbol"] for f in failures]
        failures_after_retry: List[Dict[str, Any]] = []
        successes_on_retry = 0

        for symbol in tqdm(retry_symbols, desc="Retrying failed for S&P500", ncols=100):
            row, fail = process_one(symbol, name_map, end_dt)
            if row:
                rows.append(row)
                total_gainer_numeric += row["gainer_numeric"]
                successes_on_retry += 1
            else:
                failures_after_retry.append(fail)

        failures = failures_after_retry
        logging.info("Retry completed. successes_on_retry=%d, remaining_failures=%d", successes_on_retry, len(failures))
    else:
        logging.info("No failures in first pass; skip retry.")

    # 4) 汇总与保存结果
    total_gainer_str = format_gainer_billion(total_gainer_numeric, DEFAULT_CURRENCY)
    rows.append({
        "symbol": "S&P500 Total",
        "ticker": "",
        "name": "",
        "as_of_date": "",
        "month_ago_date": "",
        "Return": "",
        "Volumn": "",
        "gainer_numeric": total_gainer_numeric,
        "Gainer": total_gainer_str,
    })

    result_df = pd.DataFrame(rows)

    # 排序（仅对个股按 gainer_numeric 排序，合计行置底）
    if not result_df.empty:
        stock_mask = result_df["symbol"] != "S&P500 Total"
        sorted_idx = result_df.loc[stock_mask, :].sort_values("gainer_numeric", ascending=False).index.tolist()
        final_rows = pd.concat([result_df.loc[sorted_idx], result_df.loc[~stock_mask]], ignore_index=True)
    else:
        final_rows = result_df

    out_xlsx = os.path.join(RAW_DIR, "sp500_gainer_results.xlsx")
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        final_rows.to_excel(writer, index=False, sheet_name="Gainer")
        ws = writer.book["Gainer"]
        from openpyxl.utils import get_column_letter
        # 自动列宽
        for col_idx, col_name in enumerate(final_rows.columns, start=1):
            max_len = max(len(str(col_name)), *(len(str(v)) for v in final_rows[col_name].astype(str)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)
        # 设置 gainer_numeric 数字格式，避免科学计数法
        if "gainer_numeric" in final_rows.columns:
            cidx = final_rows.columns.get_loc("gainer_numeric") + 1
            for r_idx in range(2, ws.max_row + 1):
                ws.cell(row=r_idx, column=cidx).number_format = '#,##0.00'

    logging.info("Results saved: %s (rows=%d)", out_xlsx, len(final_rows))

    # 5) 失败清单
    if failures:
        fails_df = pd.DataFrame(failures, columns=["symbol", "ticker", "name", "reason"])
        fails_df.to_csv(FAIL_CSV, index=False, encoding="utf-8-sig")
        logging.info("Failures saved: %s (rows=%d)", FAIL_CSV, len(failures))
        logging.info("==== FAILED SYMBOLS SUMMARY (after retry) ====")
        for f in failures:
            logging.info("FAIL | symbol=%s | ticker=%s | name=%s | reason=%s",
                         f.get("symbol",""), f.get("ticker",""), f.get("name",""), f.get("reason",""))
    else:
        logging.info("No failures remaining after retry.")

    # print("\n已保存结果到：", out_xlsx)
    # print("失败清单（如有）：", FAIL_CSV if failures else "无")
    # print("日志文件：", LOG_PATH)
    # print("原始数据目录：", RAW_SUBDIR)

if __name__ == "__main__":
    main()
