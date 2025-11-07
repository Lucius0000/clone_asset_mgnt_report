"""
债券报表：输出 bonds.xlsx（转置），并拆分“总市值 ($)”为“国债总市值 ($)”与“债券市场总市值 ($)”

实现要点：
- 中国：
  - “国债总市值 ($)”与“中国债券市场总市值 ($)”均来自 akshare 接口
    ak.bond_cash_summary_sse(date)，分别取“国债”“托管市值”和“合计”“托管市值”。
    该接口“托管市值”单位为“亿元”，转换成“B”需 /10。
  - date 参数需在 Spyder 控制台中定义（变量名 ak_bond_date / bond_date / sse_date 其一，格式 YYYYMMDD）。
    也支持环境变量 AK_BOND_DATE；若都缺失则提示并回退到“昨天”。
- 美国：
  - “国债总市值 ($)”来自 data 文件夹中以 MSPD_SumSecty 开头的 CSV，选最新“Record Date”且
    “Security Type Description”为“Total Marketable”的“Total Public Debt Outstanding (in Millions)”。
    该值单位为 Million USD，转换为 Billion USD 需 /1000，取 0 位小数，形如“29,406 B USD”。
  - “债券市场总市值 ($)”沿用原口径：FRED GFDEBTN（Million USD -> /1000 -> B USD）。
- 表格：生成标准表后转置，输出到仓库根目录 bonds.xlsx。
"""

from __future__ import annotations

import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
import pandas_datareader.data as web
from fredapi import Fred
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import akshare as ak


# 路径设置
ROOT = Path(".").resolve()
RAW_DIR = ROOT / "output" / "raw_data"
RAW_DIR.mkdir(parents=True, exist_ok=True)


def _format_billion(value_billion: float, currency: str) -> str:
    return f"{value_billion:,.0f} B {currency}"

def _most_recent_friday(ref_dt: Optional[datetime] = None) -> datetime:
    if ref_dt is None:
        ref_dt = datetime.today()
    # 星期一=0, ..., 星期五=4, 星期日=6
    days_back = (ref_dt.weekday() - 4) % 7
    return ref_dt - timedelta(days=days_back)

def _to_yyyymmdd(d: datetime) -> str:
    return d.strftime("%Y%m%d")

def _resolve_sse_date() -> str:
    """优先从 Spyder/IPython 控制台变量中读取日期；
    支持变量名：ak_bond_date / bond_date / sse_date（格式：YYYYMMDD）。
    回退顺序：环境变量 AK_BOND_DATE -> 交互 input -> 最近一个自然日。
    """
    # 1) Spyder/IPython 控制台
    try:
        from IPython import get_ipython  # type: ignore
        ip = get_ipython()
        if ip is not None:
            ns = getattr(ip, "user_ns", {}) or {}
            for k in ("ak_bond_date", "bond_date", "sse_date"):
                v = ns.get(k)
                if isinstance(v, str) and v.isdigit() and len(v) == 8:
                    return v
    except Exception:
        pass

    # 2) 环境变量
    v = os.getenv("AK_BOND_DATE")
    if isinstance(v, str) and v.isdigit() and len(v) == 8:
        return v

    # 3) 交互输入
    try:
        user_in = input("请输入数据收盘日期，需要周中交易日，回车则选择昨天：").strip()
        if user_in and user_in.isdigit() and len(user_in) == 8:
            return user_in
    except Exception:
        pass

    # 4) 默认用昨天（自然日）
    return (datetime.today() - timedelta(days=1)).strftime("%Y%m%d")


def _get_china_market_caps(date_yyyymmdd: str) -> tuple[str, str, pd.DataFrame]:
    """返回（中国国债总市值(B CNY)，中国债券市场总市值(B CNY)，原始 DataFrame）。
    akshare bond_cash_summary_sse 的‘托管面值’单位是‘亿元’，需要/10 转换为‘十亿’（B）。
    """
    df = ak.bond_cash_summary_sse(date=date_yyyymmdd)
    # 保存原始数据一份（仅保存到 output/raw_data）
    df.to_excel(RAW_DIR / f"国债总市值.xlsx", index=False)

    # 取“国债”、“合计”的托管市值
    df = df.copy()
    if "债券现货" in df.columns:
        key_col = "债券现货"
    else:
        key_col = df.columns[0]

    treasury_row = df[df[key_col].astype(str).str.contains("国债", na=False)]
    total_row = df[df[key_col].astype(str).str.contains("合计", na=False)]

    value_col = "托管市值" if "托管市值" in df.columns else "托管面值"
    treasury_val_b = float(treasury_row[value_col].iloc[0]) / 10 if not treasury_row.empty else np.nan
    total_val_b = float(total_row[value_col].iloc[0]) / 10 if not total_row.empty else np.nan

    return _format_billion(treasury_val_b, "CNY"), _format_billion(total_val_b, "CNY"), df


def _get_china_volumes_30d() -> tuple[Optional[str], Optional[str]]:
    """抓取近30个自然日上交所债券成交汇总：
    - 返回（当日成交金额-记账式国债，30日总成交金额-记账式国债），以 B CNY 表示。
    注意：ak.bond_deal_summary_sse 单位为‘万元’，转换到‘B’需 / 100000。
    """
    end_date = datetime.today()
    start_date = end_date - timedelta(days=30)
    date_range = pd.date_range(start=start_date, end=end_date)

    all_frames = []
    for d in date_range:
        date_str = d.strftime('%Y%m%d')
        try:
            df = ak.bond_deal_summary_sse(date=date_str)
            if df is not None and not df.empty:
                df["数据日期"] = d.strftime('%Y-%m-%d')
                all_frames.append(df)
        except Exception:
            continue

    if not all_frames:
        return None, None

    full_df = pd.concat(all_frames, ignore_index=True)
    try:
        full_df.to_excel(RAW_DIR / "all_bond_deal_data.xlsx", index=False)
    except Exception:
        pass

    # 最新一日“记账式国债”的当日成交金额
    latest_day = full_df["数据日期"].max()
    latest_record = full_df[(full_df["数据日期"] == latest_day) & (full_df["债券类型"] == "记账式国债")]
    day_amt_b = None
    if not latest_record.empty:
        day_amt_b = _format_billion(float(latest_record["当日成交金额"].values[0]) / 100000, "CNY")

    # 30日合计（记账式国债）
    group_sum_df = full_df.groupby("债券类型")["当日成交金额"].sum().reset_index()
    month_amt_b = None
    if not group_sum_df.empty and (group_sum_df["债券类型"] == "记账式国债").any():
        total_amt = float(group_sum_df.loc[group_sum_df["债券类型"] == "记账式国债", "当日成交金额"].values[0])
        month_amt_b = _format_billion(total_amt / 100000, "CNY")

    return day_amt_b, month_amt_b


def _get_cn_us_yield_metrics():
    """返回中国与美国 2Y/10Y 的（月收益率年化, 年收益率, 年化波动率, Sharpe）。"""

    def compute_cn(df: pd.DataFrame, col_name: str):
        x = df[["日期", col_name]].dropna().copy()
        x["日期"] = pd.to_datetime(x["日期"])
        x.set_index("日期", inplace=True)
        x.sort_index(inplace=True)
        x['r_daily'] = (1 + x[col_name] / 100) ** (1 / 252) - 1
        end_date = x.index.max()
        recent_30d = x.loc[end_date - timedelta(days=30): end_date]
        r30 = (1 + recent_30d['r_daily']).prod() ** (252 / max(1, len(recent_30d))) - 1
        recent_1y = x.loc[end_date - timedelta(days=365): end_date]
        r1y = (1 + recent_1y['r_daily']).prod() - 1
        vol_a = x.iloc[-252:]['r_daily'].std() * np.sqrt(252)
        sharpe = (r1y - 0.017) / vol_a if vol_a != 0 else np.nan
        return round(r30 * 100, 2), round(r1y * 100, 2), round(vol_a * 100, 2), round(sharpe, 2)

    def compute_us(df: pd.DataFrame, col: str):
        x = df[[col]].dropna().copy()
        x.index = pd.to_datetime(x.index)
        x.sort_index(inplace=True)
        x['r_daily'] = (1 + x[col]) ** (1 / 252) - 1
        end_date = x.index.max()
        recent_30d = x.loc[end_date - timedelta(days=30): end_date]
        r30 = (1 + recent_30d['r_daily']).prod() ** (252 / max(1, len(recent_30d))) - 1
        recent_1y = x.loc[end_date - timedelta(days=365): end_date]
        r1y = (1 + recent_1y['r_daily']).prod() - 1
        vol_a = recent_1y['r_daily'].std() * np.sqrt(252)
        sharpe = (r1y - 0.045) / vol_a if vol_a != 0 else np.nan
        return round(r30 * 100, 2), round(r1y * 100, 2), round(vol_a * 100, 2), round(sharpe, 2)

    cn_us = ak.bond_zh_us_rate(start_date="20240101")
    # 兼容不同版本列名：模糊匹配中国 2年/10年收益率列
    def _pick_cn_col(df: pd.DataFrame, tenor: str) -> str:
        tenor_candidates = [
            f"中国国债收益率:{tenor}",
            f"中国国债收益率{tenor}",
            f"中国国债收益率 {tenor}",
        ]
        for c in df.columns:
            for t in tenor_candidates:
                if t in str(c):
                    return c
        # 兜底：通过关键字与数字匹配
        for c in df.columns:
            s = str(c)
            if ("中国" in s or "国债" in s) and (tenor.replace("年", "") in s):
                return c
        raise KeyError(f"未找到中国国债收益率列：{tenor}")

    cn2_col = _pick_cn_col(cn_us, "2年")
    cn10_col = _pick_cn_col(cn_us, "10年")
    cn2 = compute_cn(cn_us, cn2_col)
    cn10 = compute_cn(cn_us, cn10_col)

    start_date = "2024-01-01"
    end_date = datetime.today().strftime("%Y-%m-%d")
    codes = {"2Y": "DGS2", "10Y": "DGS10"}
    api_key = os.getenv("FRED_API_KEY")
    data = {}
    for k, code in codes.items():
        d = web.DataReader(code, "fred", start_date, end_date, api_key=api_key)
        data[k] = d[code] / 100.0
    us_yield_df = pd.DataFrame(data)
    us2 = compute_us(us_yield_df, "2Y")
    us10 = compute_us(us_yield_df, "10Y")
    # 保存原始收益率
    try:
        us_yield_df.index.name = "Date"
        us_yield_df.to_excel(RAW_DIR / '美债收益率.xlsx')
    except Exception:
        pass
    return cn2, cn10, us2, us10


def _get_us_market_caps() -> tuple[Optional[str], Optional[str]]:
    """返回（美国国债总市值-可流通Marketable(B USD)，美国债券市场总市值-沿用原口径(B USD)）。"""
    # 1) 国债（Marketable）- 从 MSPD 文件
    treasury_b = None
    csv_files = sorted((ROOT / "data").glob("MSPD_SumSecty*.csv"))
    if csv_files:
        frames = []
        for fp in csv_files:
            try:
                x = pd.read_csv(fp)
                if {"Record Date", "Security Type Description", "Total Public Debt Outstanding (in Millions)"}.issubset(x.columns):
                    frames.append(x)
            except Exception:
                continue
        if frames:
            x = pd.concat(frames, ignore_index=True)
            x["Record Date"] = pd.to_datetime(x["Record Date"], errors="coerce")
            x = x[x["Security Type Description"].astype(str) == "Total Marketable"].dropna(subset=["Record Date"])
            if not x.empty:
                latest_row = x.sort_values("Record Date").iloc[-1]
                millions = float(latest_row["Total Public Debt Outstanding (in Millions)"])
                treasury_b = _format_billion(millions / 1000.0, "USD")

    # 2) 债券市场总市值 - 仍用 GFDEBTN（Million USD）
    market_b = None
    try:
        fred = Fred(api_key=os.environ['FRED_API_KEY'])
        series = fred.get_series('GFDEBTN')  # Million USD
        latest_value = float(series.iloc[-1])
        market_b = _format_billion(latest_value / 1000.0, "USD")
    except Exception:
        pass

    return treasury_b, market_b


def main(debug: bool = False):
    """
    - 当 debug=False（由 main.py 调用的默认情况）：无需输入，自动取最近一个周五；
      通过设置 AK_BOND_DATE 环境变量，复用原有 _resolve_sse_date() 流程。
    - 当 debug=True：要求用户输入日期（清空 AK_BOND_DATE，强制走 input 分支）。
    - 单独运行 bonds.py：等价于 debug=True。
    """

    if debug is False:
        # 自动选择最近周五（含今天为周五也会取今天）
        last_friday = _most_recent_friday()
        os.environ["AK_BOND_DATE"] = _to_yyyymmdd(last_friday)
    else:
        # 强制用户输入：清除可能存在的环境变量，避免被自动覆盖
        if "AK_BOND_DATE" in os.environ:
            os.environ.pop("AK_BOND_DATE", None)

    # 读取中国市值（需用户在 Spyder 控制台设置日期变量）
    sse_date = _resolve_sse_date()
    cn_treasury_b, cn_market_b, _ = _get_china_market_caps(sse_date)

    # 中国近 30 日成交量
    cn_day_vol_b, cn_month_vol_b = _get_china_volumes_30d()

    # 收益率与波动
    (cn2_r30, cn2_r1y, cn2_vol, _), (cn10_r30, cn10_r1y, cn10_vol, _), (us2_r30, us2_r1y, us2_vol, _), (us10_r30, us10_r1y, us10_vol, _) = _get_cn_us_yield_metrics()

    # 美国市值（国债/债券市场）
    us_treasury_b, us_market_b = _get_us_market_caps()

    # 组装行（先标准表，再转置）
    header = [
        '指标类别', '指标', '种类', '月收益率年化 (%)', '年收益率 (%)', '年化波动率 (%)',
        '国债总市值 ($)', '债券市场总市值 ($)', '当日交易量', '月交易量'
    ]

    rows = [
        # 中国 - 记账式 2年、10年
        ['中国', '记账式国债', '2年期', f"{cn2_r30:.2f}%", f"{cn2_r1y:.2f}%", f"{cn2_vol:.2f}%",
         cn_treasury_b, cn_market_b, cn_day_vol_b or '-', cn_month_vol_b or '-'],
        [None, None, '10年期', f"{cn10_r30:.2f}%", f"{cn10_r1y:.2f}%", f"{cn10_vol:.2f}%",
         '-', '-', '-', '-'],
        # 中国 - 储蓄式 3年、5年（留空占位）
        [None, '储蓄式国债', '3年期', '-', '-', '-', '-', '-', '-', '-'],
        [None, None, '5年期',  '-', '-', '-', '-', '-', '-', '-'],
        # 美国 - 记账式 2年、10年
        ['美国', '记账式国债', '2年期', f"{us2_r30:.2f}%", f"{us2_r1y:.2f}%", f"{us2_vol:.2f}%",
         us_treasury_b or '-', us_market_b or '-', '-', '-'],
        [None, None, '10年期', f"{us10_r30:.2f}%", f"{us10_r1y:.2f}%", f"{us10_vol:.2f}%",
         '-', '-', '-', '-'],
        # 美国 - 储蓄式 EE / I（留空占位）
        [None, '储蓄式国债', 'EE bonds', '-', '-', '-', '-', '-', '-', '-'],
        [None, None, 'I bonds',  '-', '-', '-', '-', '-', '-', '-'],
    ]

    df = pd.DataFrame(rows, columns=header)
    df = df.fillna('-')

    # 转置
    df_t = df.T

    # 输出到子文件夹 output/bonds.xlsx，并进行合并单元格
    out_dir = ROOT / 'output'
    out_dir.mkdir(parents=True, exist_ok=True)
    out_file = out_dir / 'bonds.xlsx'
    with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
        df_t.to_excel(writer, sheet_name='bonds', header=False, index=True)
        ws = writer.sheets['bonds']

        # 单元格居中
        max_row = ws.max_row
        max_col = ws.max_column
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                ws.cell(row=r, column=c).alignment = Alignment(horizontal='center', vertical='center')

        # 计算列索引：A=1 为指标名列，数据从 B=2 开始
        # 列布局（B..I）：
        # 中国：B=2(2Y), C=3(10Y), D=4(储蓄3Y), E=5(储蓄5Y)
        # 美国：F=6(2Y), G=7(10Y), H=8(EE), I=9(I)
        cn_start, cn_end = 2, 5
        us_start, us_end = 6, 9

        # 行索引（1-based）：
        # 1: 指标类别, 2: 指标, 3: 种类, 4: 月收益率年化, 5: 年收益率, 6: 年化波动率,
        # 7: 国债总市值, 8: 债券市场总市值, 9: 当日交易量, 10: 月交易量

        # 合并“指标类别”
        ws.merge_cells(start_row=1, start_column=cn_start, end_row=1, end_column=cn_end)
        ws.merge_cells(start_row=1, start_column=us_start, end_row=1, end_column=us_end)

        # 合并“指标”内的小类
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=3)  # 中国记账式国债（2列）
        ws.merge_cells(start_row=2, start_column=4, end_row=2, end_column=5)  # 中国储蓄式国债（2列）
        ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=7)  # 美国记账式国债（2列）
        ws.merge_cells(start_row=2, start_column=8, end_row=2, end_column=9)  # 美国储蓄式国债（2列）

        # 仅对“国债总市值 ($)”与“债券市场总市值 ($)”两行按国家范围合并
        # 行 7、8 分别为这两项
        for row_idx in (7, 8):
            ws.merge_cells(start_row=row_idx, start_column=cn_start, end_row=row_idx, end_column=cn_end)  # 中国 B..E
            ws.merge_cells(start_row=row_idx, start_column=us_start, end_row=row_idx, end_column=us_end)  # 美国 F..I


if __name__ == "__main__":
    main(debug=True)