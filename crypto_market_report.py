"""
加密货币指标计算器
- 计算 BTC/ETH/XRP/USDT 等加密货币的短期（30天）、中期（90天）与长期（365天）指标
- 输出包括：年化波动率、年化收益率、夏普比率、两周变动、过去一年百分点位
- 将结果同时打印到控制台(可选 --debug)并导出到 Excel，绘制 Treemap（面积=市值，颜色=两周变动）

使用方式：
- 直接运行: python crypto_metrics_calculator.py
- 指定结束日: 修改文件顶部 END_DATE_STR 或命令行不传（默认今天）
- 打印控制台报告：加 --debug
"""

import argparse
import os
from typing import List

import matplotlib.pyplot as plt
import matplotlib.patheffects as pe
import numpy as np
import pandas as pd
import squarify
import yfinance as yf
from matplotlib.colors import TwoSlopeNorm
from datetime import datetime

# ========= 全局配置 =========
# 在这里指定结束日期（如 "2025-08-11"），None 表示用今天
END_DATE_STR = None

OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# 代理（如本机未运行代理可临时注释）
os.environ['http_proxy'] = 'http://127.0.0.1:7890'
os.environ['https_proxy'] = 'http://127.0.0.1:7890'

# 常量
ANNUALIZATION_DAYS = 365
MONTH_WINDOW_DAYS = 30
QUARTER_WINDOW_DAYS = 90
YEAR_WINDOW_DAYS = 365
RISK_FREE_RATE = 0.035  # 3.5%

METRICS_TICKERS = ["BTC-USD", "ETH-USD", "XRP-USD", "USDT-USD"]

VIZ_TICKERS = [
    "BTC-USD", "ETH-USD", "USDT-USD", "USDC-USD", "BNB-USD",
    "XRP-USD", "ADA-USD", "SOL-USD", "DOGE-USD", "TON-USD",
    "TRX-USD", "DOT-USD", "AVAX-USD", "LINK-USD", "LTC-USD",
    "BCH-USD", "XLM-USD", "UNI-USD", "ETC-USD", "MATIC-USD"
]

# ================= 抓取 =================
def fetch_price_volume(ticker: str, lookback_days: int = 420,
                       end_date: pd.Timestamp | str | None = None) -> pd.DataFrame:
    """优先使用 Ticker.history 单标的抓取（列规整）；支持指定结束日期"""
    if end_date is None:
        end_dt = pd.Timestamp.today().normalize()
    else:
        end_dt = pd.Timestamp(end_date).normalize()
    start_date = end_dt - pd.Timedelta(days=lookback_days)

    try:
        t = yf.Ticker(ticker)
        hist = t.history(
            start=start_date.strftime("%Y-%m-%d"),
            end=(end_dt + pd.Timedelta(days=1)).strftime("%Y-%m-%d"),
            interval="1d",
            auto_adjust=True,
        )
        if hist is not None and not hist.empty:
            out = hist[["Close", "Volume"]].copy()
            out["Close"] = pd.to_numeric(out["Close"], errors="coerce")
            if "Volume" in out.columns:
                out["Volume"] = pd.to_numeric(out["Volume"], errors="coerce")
            else:
                out["Volume"] = np.nan
            out.index = pd.to_datetime(out.index)
            return out.dropna(subset=["Close"]).sort_index()
        # 回退：download
        df = yf.download(
            ticker,
            start=start_date.strftime("%Y-%m-%d"),
            end=(end_dt + pd.Timedelta(days=1)).strftime("%Y-%m-%d"),
            interval="1d",
            progress=False,
            auto_adjust=True,
            group_by="column",
        )
        if df is None or df.empty:
            print(f"无法获取 {ticker} 的数据")
            return pd.DataFrame()
        out = pd.DataFrame({"Close": pd.to_numeric(df["Close"], errors="coerce")})
        out["Volume"] = pd.to_numeric(df.get("Volume", np.nan), errors="coerce")
        out.index = pd.to_datetime(out.index)
        return out.dropna(subset=["Close"]).sort_index()
    except Exception as e:
        print(f"获取 {ticker} 数据时出错: {e}")
        return pd.DataFrame()

def fetch_market_cap(yf_ticker: str) -> float:
    """总市值：fast_info.market_cap 优先，回退 info['marketCap']"""
    mc = np.nan
    try:
        t = yf.Ticker(yf_ticker)
        fi = getattr(t, "fast_info", None)
        if fi is not None:
            if isinstance(fi, dict):
                mc = fi.get("market_cap", np.nan)
            else:
                mc = getattr(fi, "market_cap", np.nan)
        if mc is None or (isinstance(mc, float) and np.isnan(mc)):
            info = getattr(t, "info", {}) or {}
            mc = info.get("marketCap", np.nan)
    except Exception:
        mc = np.nan
    try:
        return float(mc)
    except Exception:
        return np.nan

# ================= 计算 =================
def compute_annualized_volatility(daily_returns: pd.Series, window_days: int) -> float:
    if daily_returns is None or daily_returns.empty:
        return float("nan")
    sub = daily_returns.dropna()
    if len(sub) < 2:
        return float("nan")
    sub = sub.iloc[-window_days:] if len(sub) >= window_days else sub
    return float(sub.std(ddof=1) * np.sqrt(ANNUALIZATION_DAYS))

def compute_annualized_return(daily_returns: pd.Series, window_days: int, is_long_term: bool = False) -> float:
    if daily_returns is None or daily_returns.empty or len(daily_returns) < 2:
        return float("nan")
    sub = daily_returns.dropna()
    sub = sub.iloc[-window_days:] if len(sub) >= window_days else sub
    if is_long_term and len(sub) >= 30:
        price_index = (1 + sub).cumprod()
        total_return = price_index.iloc[-1] - 1
        years = len(sub) / ANNUALIZATION_DAYS
        return float((1 + total_return) ** (1 / years) - 1)
    return float(sub.mean() * ANNUALIZATION_DAYS)

def compute_sharpe_ratio(annualized_return: float, annualized_volatility: float, risk_free_rate: float = RISK_FREE_RATE) -> float:
    if (pd.isna(annualized_return) or pd.isna(annualized_volatility) or annualized_volatility == 0):
        return float("nan")
    return (annualized_return - risk_free_rate) / annualized_volatility

def compute_relative_change(price: pd.Series, lag_days: int) -> float:
    if price is None or price.empty or len(price.dropna()) <= lag_days:
        return float("nan")
    prev = price.shift(lag_days).dropna().iloc[-1]
    curr = price.dropna().iloc[-1]
    if pd.isna(prev) or pd.isna(curr) or prev == 0:
        return float("nan")
    return float(curr / prev - 1.0)

def compute_percentile_1y(price: pd.Series) -> float:
    """过去365天内当前收盘价的分位（0-1）。"""
    if price is None or price.dropna().empty:
        return float("nan")
    win = price.dropna().iloc[-YEAR_WINDOW_DAYS:]
    if len(win) < 2:
        return float("nan")
    return float(win.rank(pct=True).iloc[-1])

def calculate_metrics(tickers: List[str] = None,
                      risk_free_rate: float = RISK_FREE_RATE,
                      end_date: pd.Timestamp | str | None = END_DATE_STR) -> pd.DataFrame:
    if tickers is None:
        tickers = {"BTC": "BTC-USD", "ETH": "ETH-USD", "XRP": "XRP-USD", "USDT": "USDT-USD"}
    else:
        formatted = {}
        for t in tickers:
            if "-" not in t:
                formatted[t.split("-")[0]] = f"{t}-USD"
            else:
                formatted[t.split("-")[0]] = t
        tickers = formatted

    results = []
    for symbol, yf_ticker in tickers.items():
        print(f"正在处理 {symbol}...")
        # 覆盖 1Y 百分位与 2周变动
        df_pv = fetch_price_volume(yf_ticker, lookback_days=YEAR_WINDOW_DAYS + 60, end_date=end_date)
        if df_pv.empty:
            results.append({"Symbol": symbol})
            continue

        price = df_pv["Close"].copy()
        daily_returns = price.pct_change().dropna()

        # 基础
        current_price = float(price.iloc[-1])
        price_chg_1d = float(daily_returns.iloc[-1]) if len(daily_returns) else float("nan")

        # 波动/收益/夏普
        m_vol = compute_annualized_volatility(daily_returns, MONTH_WINDOW_DAYS)
        m_ret = compute_annualized_return(daily_returns, MONTH_WINDOW_DAYS)
        m_shp = compute_sharpe_ratio(m_ret, m_vol, risk_free_rate)

        q_vol = compute_annualized_volatility(daily_returns, QUARTER_WINDOW_DAYS)
        q_ret = compute_annualized_return(daily_returns, QUARTER_WINDOW_DAYS)
        q_shp = compute_sharpe_ratio(q_ret, q_vol, risk_free_rate)

        y_vol = compute_annualized_volatility(daily_returns, YEAR_WINDOW_DAYS)
        y_ret = compute_annualized_return(daily_returns, YEAR_WINDOW_DAYS, is_long_term=True)
        y_shp = compute_sharpe_ratio(y_ret, y_vol, risk_free_rate)

        # 百分比变化
        mom    = compute_relative_change(price, MONTH_WINDOW_DAYS)
        yoy    = compute_relative_change(price, YEAR_WINDOW_DAYS)
        chg_2w = compute_relative_change(price, 14)

        # 1Y 百分位（0-1）
        pct_1y = compute_percentile_1y(price)

        # 总市值
        mcap = fetch_market_cap(yf_ticker)

        # 近7日交易量（美元额优先）
        weekly_vol_usd = np.nan
        try:
            tail = df_pv.tail(7)
            if "Volume" in df_pv.columns and tail["Volume"].notna().any():
                weekly_vol_usd = float((tail["Close"] * tail["Volume"]).sum())
            else:
                weekly_vol_usd = float("nan")
        except Exception:
            weekly_vol_usd = float("nan")

        results.append({
            "Symbol": symbol,
            "Current_Price": current_price,
            "Price_Change_1D": price_chg_1d,
            "Month_Volatility": m_vol,
            "Month_Return": m_ret,
            "Month_Sharpe": m_shp,
            "Quarter_Volatility": q_vol,
            "Quarter_Return": q_ret,
            "Quarter_Sharpe": q_shp,
            "Year_Volatility": y_vol,
            "Year_Return": y_ret,
            "Year_Sharpe": y_shp,
            "MoM": mom,
            "YoY": yoy,
            "Change_2W": chg_2w,        # 两周变动（小数）
            "Pctile_1Y": pct_1y,        # 1年百分位（0-1）
            "Market_Cap": mcap,
            "Weekly_Volume": weekly_vol_usd,  # 美元额
        })

    return pd.DataFrame(results)

# ================= 控制台打印（原样） =================
def format_decimal(x: float, digits: int = 2) -> str:
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return "NaN"
    return f"{x:.{digits}f}"

def print_metrics_report(df: pd.DataFrame) -> None:
    if df.empty:
        print("没有可用数据"); return
    print("\n======== 加密货币指标报告 ========")
    print(f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 35)
    for _, row in df.iterrows():
        symbol = row["Symbol"]
        print(f"\n{symbol} 当前价格: ${format_decimal(row.get('Current_Price'))}, 日涨跌: {format_decimal(row.get('Price_Change_1D')*100)}%")
        if "Market_Cap" in row:
            print(f"{symbol} 总市值: ${format_decimal(row.get('Market_Cap'), 2)}")
        if "Weekly_Volume" in row:
            print(f"{symbol} 近7日交易量(USD): {format_decimal(row.get('Weekly_Volume'), 0)}")
        if "MoM" in row:
            print(f"{symbol} MoM: {format_decimal(row.get('MoM')*100)}%")
        if "Change_2W" in row:
            print(f"{symbol} 两周变动: {format_decimal(row.get('Change_2W')*100)}%")
        if "Pctile_1Y" in row:
            print(f"{symbol} 百分位(1Y): {format_decimal(row.get('Pctile_1Y')*100)}%")
        if "YoY" in row:
            print(f"{symbol} YoY: {format_decimal(row.get('YoY')*100)}%")
        print(f"\n短期（月，近30天）指标:")
        print(f"  年化波动率: {format_decimal(row.get('Month_Volatility')*100)}%")
        print(f"  年化收益率: {format_decimal(row.get('Month_Return')*100)}%")
        print(f"  夏普比率: {format_decimal(row.get('Month_Sharpe'))}")
        print(f"\n中期（季度，近90天）指标:")
        print(f"  年化波动率: {format_decimal(row.get('Quarter_Volatility')*100)}%")
        print(f"  年化收益率: {format_decimal(row.get('Quarter_Return')*100)}%")
        print(f"  夏普比率: {format_decimal(row.get('Quarter_Sharpe'))}")
        print(f"\n长期（年度，近365天）指标:")
        print(f"  年化波动率: {format_decimal(row.get('Year_Volatility')*100)}%")
        print(f"  年化收益率: {format_decimal(row.get('Year_Return')*100)}%")
        print(f"  夏普比率: {format_decimal(row.get('Year_Sharpe'))}")
        print("-" * 35)
    print("\n注: 波动率和收益率均为年化值；夏普比率无风险收益率为 3.5%")
    print("=" * 35)

# ================= 模板化导出 =================
def export_to_excel_template(df: pd.DataFrame, filename: str = "crypto_metrics_template.xlsx") -> bool:
    """
    生成与模板匹配的表：
    A1:B1 = 指标
    A2:B2 = 总市值 (B $)
    A3:B3 = 收盘价
    A4:B4 = 交易量（B $)
    A5:B5 = 环比 MoM(%)
    A6:B6 = 同比 YoY(%)
    A7:B7 = 百分位(1Y, %)
    A8:B8 = 两周变动(%)
    A9:A11  = 波动率(%),   B9~B11  = 短期/月 | 中期/季 | 长期/年
    A12:A14 = Sharp Ratio, B12~B14 = 短期/月 | 中期/季 | 长期/年
    A15:A17 = 收益率年化(%),B15~B17 = 短期/月 | 中期/季 | 长期/年
    C.. = 币种（按 df['Symbol'] 顺序）
    """
    try:
        if df.empty:
            print("没有数据可导出"); return False

        os.makedirs(os.path.dirname(filename) or ".", exist_ok=True)

        syms = [str(s).strip() for s in df["Symbol"].tolist()]

        price_map = df.set_index("Symbol")["Current_Price"].reindex(syms)
        mom_map   = (df.set_index("Symbol")["MoM"] * 100).reindex(syms)
        yoy_map   = (df.set_index("Symbol")["YoY"] * 100).reindex(syms)
        chg2w_map = (df.set_index("Symbol")["Change_2W"] * 100).reindex(syms)
        pct1y_map = (df.set_index("Symbol")["Pctile_1Y"] * 100).reindex(syms)
        mcap_map  = df.set_index("Symbol")["Market_Cap"].reindex(syms) / 1e9  # 十亿美元
        wv_map    = (df.set_index("Symbol")["Weekly_Volume"].reindex(syms) / 1e9)

        # 波动率(%)
        mvol = (df.set_index("Symbol")["Month_Volatility"] * 100).reindex(syms)
        qvol = (df.set_index("Symbol")["Quarter_Volatility"] * 100).reindex(syms)
        yvol = (df.set_index("Symbol")["Year_Volatility"] * 100).reindex(syms)

        # Sharpe
        msp = df.set_index("Symbol")["Month_Sharpe"].reindex(syms)
        qsp = df.set_index("Symbol")["Quarter_Sharpe"].reindex(syms)
        ysp = df.set_index("Symbol")["Year_Sharpe"].reindex(syms)

        # 收益率年化(%)
        mret = (df.set_index("Symbol")["Month_Return"] * 100).reindex(syms)
        qret = (df.set_index("Symbol")["Quarter_Return"] * 100).reindex(syms)
        yret = (df.set_index("Symbol")["Year_Return"] * 100).reindex(syms)

        from openpyxl import Workbook
        from openpyxl.styles import Alignment
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # 列标题: C.. = 币种
        ws["A1"] = "指标"; ws.merge_cells("A1:B1")
        ws["A2"] = "总市值 (B $)"; ws.merge_cells("A2:B2")
        ws["A3"] = "收盘价"; ws.merge_cells("A3:B3")
        ws["A4"] = "交易量（B $)"; ws.merge_cells("A4:B4")
        ws["A5"] = "环比 MoM(%)"; ws.merge_cells("A5:B5")
        ws["A6"] = "同比 YoY(%)"; ws.merge_cells("A6:B6")
        ws["A7"] = "百分位(1Y, %)"; ws.merge_cells("A7:B7")
        ws["A8"] = "两周变动(%)"; ws.merge_cells("A8:B8")

        ws["A9"]  = "波动率(%)"; ws.merge_cells("A9:A11")
        ws["B9"]  = "短期（月）"
        ws["B10"] = "中期（季度）"
        ws["B11"] = "长期（年）"

        ws["A12"] = "Sharp Ratio"; ws.merge_cells("A12:A14")
        ws["B12"] = "短期（月）"
        ws["B13"] = "中期（季度）"
        ws["B14"] = "长期（年）"

        ws["A15"] = "收益率年化(%)"; ws.merge_cells("A15:A17")
        ws["B15"] = "短期（月）"
        ws["B16"] = "中期（季度）"
        ws["B17"] = "长期（年）"

        # 币种列头
        start_col = 3  # C列
        for i, s in enumerate(syms):
            ws.cell(row=1, column=start_col + i, value=s)

        # 写入工具
        def write_row_values(row_idx, series, integer_thousands=False):
            for i, s in enumerate(syms):
                val = None if pd.isna(series.get(s)) else float(series.get(s))
                cell = ws.cell(row=row_idx, column=start_col + i, value=val)
                cell.number_format = '#,##0' if integer_thousands else '#,##0.00'

        # A2~A8
        write_row_values(2,  mcap_map, integer_thousands=True)   # 总市值(B$)
        write_row_values(3,  price_map)                          # 收盘价
        write_row_values(4,  wv_map, integer_thousands=True)     # 交易量(B$)
        write_row_values(5,  mom_map)                            # MoM %
        write_row_values(6,  yoy_map)                            # YoY %
        write_row_values(7,  pct1y_map)                          # 百分位(1Y)
        write_row_values(8,  chg2w_map)                          # 两周变动 %

        # 波动率
        write_row_values(9,  mvol)
        write_row_values(10, qvol)
        write_row_values(11, yvol)

        # Sharpe
        write_row_values(12, msp)
        write_row_values(13, qsp)
        write_row_values(14, ysp)

        # 收益率年化
        write_row_values(15, mret)
        write_row_values(16, qret)
        write_row_values(17, yret)

        # 对齐与列宽
        for col in range(1, start_col + len(syms)):
            for row in range(1, 18):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 14
        for i in range(len(syms)):
            ws.column_dimensions[chr(ord('C')+i)].width = 14

        wb.save(filename)
        return True
    except Exception as e:
        print(f"导出Excel(模板)时出错: {e}")
        return False

def plot_crypto_treemap(
    df: pd.DataFrame,
    top_n: int = 15,
    title: str | None = "Top 15 Crypto Treemap",   # 主标题（可为 None）
    subtitle: str | None = None,                    # 副标题（可为 None）
    cmap_name: str = "RdYlGn",                      # 绿涨红跌
    output_path: str | None = None,
    min_font: int = 7,
    max_font: int = 20,
    two_line_area_frac: float = 0.03
) -> str:
    """
    绘制加密货币 Treemap：
      - 面积 = Market_Cap；颜色 = Change_2W（[-20%, +20%] 截断）
      - 主/副标题用 fig.text 以固定行距硬编码到图顶部，保证不与色条、图形重叠。
      - 若未传 subtitle 且 title 形如 '主标题 (副标题)'，自动拆分。
    """
    if df is None or df.empty:
        raise ValueError("空的 DataFrame，无法绘图。")

    data = (df.replace([np.inf, -np.inf], np.nan)
              .dropna(subset=["Market_Cap", "Change_2W"])
              .sort_values("Market_Cap", ascending=False)
              .head(top_n))
    if data.empty:
        raise ValueError("没有可用的市值或 2W 变动用于绘图。")

    sizes   = data["Market_Cap"].astype(float).values
    chg2w   = data["Change_2W"].astype(float).values   # 小数
    symbols = data["Symbol"].astype(str).values

    # 颜色映射
    vmin, vmax = -0.20, 0.20
    chg_clip = np.clip(chg2w, vmin, vmax)
    norm = TwoSlopeNorm(vmin=vmin, vcenter=0.0, vmax=vmax)
    cmap = plt.get_cmap(cmap_name)
    colors = [cmap(norm(x)) for x in chg_clip]

    # 布局
    norm_sizes = squarify.normalize_sizes(sizes, 100, 100)
    rects = squarify.squarify(norm_sizes, 0, 0, 100, 100)

    fig = plt.figure(figsize=(12, 8), dpi=150)
    ax = plt.gca()

    # treemap 矩形
    for r, c in zip(rects, colors):
        ax.add_patch(plt.Rectangle((r['x'], r['y']), r['dx'], r['dy'],
                                   facecolor=c, edgecolor="white", linewidth=1))

    # 文本字号随面积
    areas = np.array([r['dx'] * r['dy'] for r in rects], dtype=float)
    total_area = areas.sum() if areas.sum() > 0 else 1.0
    scale = np.sqrt(areas / areas.max()) if areas.max() > 0 else np.ones_like(areas)
    fonts = (min_font + (max_font - min_font) * scale).clip(min_font, max_font)

    fig.canvas.draw()
    renderer = fig.canvas.get_renderer()

    # 标签（自适应黑/白 + 描边）
    for (r, sym, chg, fs, area, c_rgba) in zip(rects, symbols, chg2w, fonts, areas, colors):
        x = r['x'] + r['dx'] / 2
        y = r['y'] + r['dy'] / 2
        r_c, g_c, b_c = c_rgba[:3]
        luminance = 0.299 * r_c + 0.587 * g_c + 0.114 * b_c
        text_color = "black" if luminance > 0.6 else "white"
        stroke_color = "white" if text_color == "black" else "black"

        line1 = sym
        line2 = f"{chg*100:+.2f}%"
        show_two_lines = (area / total_area >= two_line_area_frac)

        text_obj = None
        if show_two_lines:
            text_obj = ax.text(x, y, f"{line1}\n{line2}",
                               ha="center", va="center",
                               fontsize=float(fs), color=text_color, weight="bold",
                               path_effects=[pe.withStroke(linewidth=1.5, foreground=stroke_color)])
            bbox = text_obj.get_window_extent(renderer=renderer).transformed(ax.transData.inverted())
            if (bbox.width > r['dx'] * 0.95) or (bbox.height > r['dy'] * 0.95):
                text_obj.remove()
                text_obj = None

        if text_obj is None:
            ax.text(x, y, line1,
                    ha="center", va="center",
                    fontsize=float(fs), color=text_color, weight="bold",
                    path_effects=[pe.withStroke(linewidth=1.5, foreground=stroke_color)])

    ax.set_xlim(0, 100)
    ax.set_ylim(0, 100)
    ax.set_aspect("equal")
    ax.axis("off")

    # ======= 主/副标题：硬编码法（绝不重合、相对绘图区居中） ======= #
    # 解析主/副标题
    main_title = (title or "").strip() if title else ""
    sub = (subtitle or "").strip() if subtitle else None
    if not sub and title and "(" in title and title.endswith(")"):
        main_title, sub_part = title.split("(", 1)
        main_title = main_title.strip()
        sub = sub_part[:-1].strip()

    # 先画色条，再定位标题，避免位置被色条改变
    sm = plt.cm.ScalarMappable(cmap=cmap, norm=TwoSlopeNorm(vmin=vmin*100, vcenter=0.0, vmax=vmax*100))
    sm.set_array([])
    cbar = plt.colorbar(sm, ax=ax, fraction=0.03, pad=0.02)
    cbar.set_label("2W Change (%)")

    # 计算绘图区在 figure 坐标中的中心 x（不受色条影响）
    fig.canvas.draw()
    axpos = ax.get_position()  # Bbox in figure coords
    cx = axpos.x0 + axpos.width / 2

    # 固定两条 y 线：主标题 y_main；副标题 y_sub（稍微低于主标题）
    y_main = 0.965
    y_sub  = 0.938  # 与主标题相隔约 0.023 的 figure 高度

    # 画主标题
    if main_title:
        fig.text(cx, y_main, main_title, ha="center", va="top",
                 fontsize=18, fontweight="bold")

    # 画副标题
    if sub:
        fig.text(cx, y_sub, sub, ha="center", va="top",
                 fontsize=11, color="0.35")

    # 顶部留白，防止被裁切
    plt.subplots_adjust(top=0.92)

    if output_path is None:
        output_path = os.path.join(OUTPUT_DIR, "crypto_treemap.png")
    plt.savefig(output_path, bbox_inches="tight")
    plt.close()
    return output_path

# ================= 主流程 =================
def main(debug: bool = False) -> None:
    print("正在计算加密货币指标...")

    # 指标表：只统计四个
    df_metrics = calculate_metrics(METRICS_TICKERS, end_date=END_DATE_STR)

    # 仅在 debug=True 时打印到控制台
    if debug:
        print_metrics_report(df_metrics)

    excel_path = os.path.join(OUTPUT_DIR, "crypto_metrics.xlsx")
    export_to_excel_template(df_metrics, filename=excel_path)

    # 可视化：统计至少15个主流加密货币
    df_viz = calculate_metrics(VIZ_TICKERS, end_date=END_DATE_STR)
    png_path = plot_crypto_treemap(
        df_viz,
        top_n=15,
        title="Top 15 Crypto Treemap",
        subtitle="Area = Market Cap · Color = 2W Change",
        cmap_name="RdYlGn",  # 绿涨红跌
        output_path=os.path.join(OUTPUT_DIR, "crypto_treemap.png")
    )
    print(f"已导出：{excel_path}")
    print(f"Treemap 已保存：{png_path}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Crypto metrics and treemap generator")
    parser.add_argument("--debug", action="store_true",
                        help="若指定则打印四个币种的控制台报告")
    args = parser.parse_args()
    main(debug=args.debug)
