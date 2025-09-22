"""
"周报-资产大类表现-整体"表格计算
需要提前把各个资产大类的`月收益率年化 (%)``年收益率 (%)``月波动率年化（%）``总市值 ($)`按照`周报-资产大类表现-整体`的形式整理成`*.xlsx`表格，放在`data`子文件夹
输出整体.xlsx
"""

import os
from openpyxl import load_workbook

# === 路径 ===
input_path = "data/整体.xlsx"
output_path = "output/整体_processed.xlsx"
log_path = "output/raw_data/整体_calculation_steps.txt"

# === 工具函数（Sheet1 百分比的“读/写”统一走 number_format 判断）===
def read_percent_as_fraction(cell):
    """
    从 Sheet1 单元格读取“用于计算的小数”：
    - 若单元格为百分比格式（number_format 含 '%'），其数值本身就是小数，例如 31.05% -> 0.3105
    - 若非百分比格式，数值表示去掉%后的百分数，例如 28.93 -> 0.2893
    """
    v = cell.value
    if v is None:
        return None
    # 字符串兜底（虽然你说都是数值，这里仅为稳妥）
    if isinstance(v, str):
        s = v.strip().replace("％", "%")
        if s.endswith("%"):
            s = s[:-1]
            return float(s) / 100.0
        return float(s) / 100.0  # 无%时按百分数处理
    # 数值
    fmt = (cell.number_format or "").lower()
    if "%" in fmt:
        return float(v)        # 已是 fraction
    else:
        return float(v) / 100.0  # 28.93 -> 0.2893

def plain_number_from_percent_cell(cell):
    """
    将 Sheet1 百分比列的单元格，转换成“去掉%后的数值”以写回：
    - 若为百分比格式：0.3105 -> 31.05
    - 若为普通数字：28.93 -> 28.93
    """
    v = cell.value
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip().replace("％", "%")
        if s.endswith("%"):
            s = s[:-1]
        return round(float(s), 2)
    fmt = (cell.number_format or "").lower()
    if "%" in fmt:
        return round(float(v) * 100.0, 2)
    else:
        return round(float(v), 2)

# === 工具函数（Sheet2 YoY：无%百分比，需 /100 转 fraction）===
def find_yoy(sheet, pair_name):
    """
    在 Sheet2 里找到指定货币对的 YoY，并返回“用于计算的小数”（fraction）。
    约定：Sheet2 的 YoY 是“无%符号的百分比”，例如
      - USD_HKD YoY = 0.708182587945502  表示 0.708182587945502%  ->  0.00708182587945502
      - USD_CNH YoY = 0.0584559284054055 表示 0.0584559284054055% -> 0.000584559284054055
    因此：fraction = (读取到的数值) / 100
    """
    header = [sheet.cell(1, c).value for c in range(1, sheet.max_column + 1)]
    yoy_cols = [i for i, v in enumerate(header, start=1) if v and "yoy" in str(v).lower()]

    def _to_float(v):
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip().replace("％", "%")
        if s.endswith("%"):
            s = s[:-1]
        try:
            return float(s)
        except ValueError:
            return None

    for r in range(2, sheet.max_row + 1):
        row_has_pair = any(sheet.cell(r, c).value == pair_name for c in range(1, sheet.max_column + 1))
        if not row_has_pair:
            continue

        yoy_val = None
        for yc in yoy_cols:
            val = sheet.cell(r, yc).value
            if val is not None:
                yoy_val = val
                break
        if yoy_val is None:
            row_vals = [sheet.cell(r, cc).value for cc in range(1, sheet.max_column + 1)]
            nums = [v for v in row_vals if isinstance(v, (int, float))]
            if nums:
                yoy_val = nums[-1]

        y = _to_float(yoy_val)
        if y is None:
            return None

        return y / 100.0  # 无%百分比 -> fraction

def fmt_frac(x, d=6):
    return "None" if x is None else f"{x:.{d}f}"

def fmt_pctnum(x, d=2):
    """把 fraction 小数转为“去掉%后的数值”字符串，例如 0.3286 -> '32.86'"""
    if x is None:
        return "None"
    return f"{x*100:.{d}f}"

# === 读取工作簿 ===
wb = load_workbook(input_path)
ws = wb["Sheet1"]
ws2 = wb["Sheet2"]

# === 构建 区域 -> 本地无风险利率 映射（债券固收行；处理合并单元格：前向填充区域）===
rf_by_region = {}
current_region = None
for r in range(2, ws.max_row + 1):
    region_cell_val = ws.cell(r, 1).value
    if region_cell_val not in (None, ""):
        current_region = str(region_cell_val).strip()

    asset = ws.cell(r, 2).value
    ret_frac = read_percent_as_fraction(ws.cell(r, 3))  # <=== 用新逻辑

    if asset == "债券固收" and current_region and ret_frac is not None:
        rf_by_region[current_region] = ret_frac

# 抓取美国与中国的无风险，香港固定为 0.0062（题设）
us_rf = rf_by_region.get("美国", None)
cn_rf = rf_by_region.get("中国", None)
hk_rf = 0.0062

# === Sheet2 汇率 YoY（fraction）===
usd_cnh_yoy = find_yoy(ws2, "USD_CNH")
usd_hkd_yoy = find_yoy(ws2, "USD_HKD")

# 调试输出（必要时可注释掉）
print("rf_by_region:", rf_by_region)
print("us_rf:", us_rf, "cn_rf:", cn_rf, "hk_rf:", hk_rf)
print("usd_cnh_yoy (fraction):", usd_cnh_yoy, "-> YoY(%) plain:", fmt_pctnum(usd_cnh_yoy, 12))
print("usd_hkd_yoy (fraction):", usd_hkd_yoy, "-> YoY(%) plain:", fmt_pctnum(usd_hkd_yoy, 12))

# 若美国无风险未读到，给出提示并兜底为 0 防止报错
if us_rf is None:
    print("警告：未在表中识别到美国‘债券固收’无风险利率，将临时使用 0.0。请检查 Sheet1 的区域/合并单元格。")
    us_rf = 0.0

# === 准备日志 ===
os.makedirs(os.path.dirname(output_path), exist_ok=True)
os.makedirs(os.path.dirname(log_path), exist_ok=True)
log_lines = []
def log(s: str):
    print(s)
    log_lines.append(s)

log("=== Calculation Steps ===")

# === 逐行计算 Sharpe / Adjusted Sharpe ===
current_region = None
for r in range(2, ws.max_row + 1):
    region_cell_val = ws.cell(r, 1).value
    if region_cell_val not in (None, ""):
        current_region = str(region_cell_val).strip()
    region = current_region

    asset = ws.cell(r, 2).value
    ret = read_percent_as_fraction(ws.cell(r, 3))  # <=== 用新逻辑
    vol = read_percent_as_fraction(ws.cell(r, 5))  # <=== 用新逻辑

    if ret is None or vol is None:
        continue

    hdr = f"[Row {r}] 区域={region or '-'} | 资产={asset or '-'}"
    log(hdr)

    if asset == "债券固收":
        log(f"  规则：债券固收作为无风险资产，Sharpe=0，Adjusted Sharpe=0")
        log(f"  数据：return={fmt_frac(ret)} ({fmt_pctnum(ret)}), vol={fmt_frac(vol)} ({fmt_pctnum(vol)})")
        ws.cell(r, 7).value = 0.00
        ws.cell(r, 7).number_format = "0.00"
        ws.cell(r, 8).value = 0.00
        ws.cell(r, 8).number_format = "0.00"
        log("")
        continue

    # Sharpe: 本地无风险
    if region == "美国":
        rf_local = us_rf
    elif region == "中国":
        rf_local = cn_rf
    elif region == "香港":
        rf_local = hk_rf
    else:
        rf_local = us_rf

    if rf_local is None:
        log("  注意：未取到本地无风险 rf_local，临时使用美国无风险 us_rf")
        rf_local = us_rf

    numerator = (ret - rf_local)
    sharpe = (numerator / vol) if vol != 0 else 0.0
    log(f"  Sharpe 计算： (return - rf_local) / vol = ({fmt_frac(ret)} - {fmt_frac(rf_local)}) / {fmt_frac(vol)} = {fmt_frac(sharpe)} -> {round(sharpe,2):.2f}")
    log(f"    其中：return={fmt_frac(ret)} ({fmt_pctnum(ret)})，rf_local={fmt_frac(rf_local)} ({fmt_pctnum(rf_local)})，vol={fmt_frac(vol)} ({fmt_pctnum(vol)})")

    currency_effect = 0.0
    fx_note = "FX: 无（USD 计价或不适用）"
    if region == "中国" and usd_cnh_yoy is not None:
        currency_effect = -usd_cnh_yoy
        fx_note = f"FX: 使用 USD_CNH YoY(%)={fmt_pctnum(usd_cnh_yoy, 12)} → effect={fmt_frac(currency_effect, 12)}"
    elif region == "香港" and usd_hkd_yoy is not None:
        currency_effect = -usd_hkd_yoy
        fx_note = f"FX: 使用 USD_HKD YoY(%)={fmt_pctnum(usd_hkd_yoy, 12)} → effect={fmt_frac(currency_effect, 12)}"

    adj_sharpe = ((ret + currency_effect - us_rf) / vol) if vol != 0 else 0.0

    if asset in ("商品与贵金属（黄金）", "数字货币（BTC）"):
        log("  说明：该资产以 USD 计价/无汇率调整，Adjusted Sharpe = Sharpe")
        adj_sharpe = sharpe
        fx_note = "FX: 无（黄金/BTC 不做汇率调整）"

    log(f"  Adjusted Sharpe 计算： (return + FX - us_rf) / vol = ({fmt_frac(ret)} + {fmt_frac(currency_effect, 12)} - {fmt_frac(us_rf)}) / {fmt_frac(vol)} = {fmt_frac(adj_sharpe)} -> {round(adj_sharpe,2):.2f}")
    log(f"    {fx_note}；us_rf={fmt_frac(us_rf)} ({fmt_pctnum(us_rf)})")

    ws.cell(r, 7).value = round(sharpe, 2)
    ws.cell(r, 7).number_format = "0.00"
    ws.cell(r, 8).value = round(adj_sharpe, 2)
    ws.cell(r, 8).number_format = "0.00"
    log("")

# === 去掉百分号：把 C(3)/D(4)/E(5) 列转为“无%数值”，保留两位小数 ===
for r in range(2, ws.max_row + 1):
    for c in (3, 4, 5):
        cell = ws.cell(r, c)
        plain = plain_number_from_percent_cell(cell)  # <=== 用新逻辑
        if plain is not None:
            cell.value = plain
            cell.number_format = "0.00"

# === 保存文件与日志 ===
wb.save(output_path)
with open(log_path, "w", encoding="utf-8") as f:
    f.write("\n".join(log_lines))