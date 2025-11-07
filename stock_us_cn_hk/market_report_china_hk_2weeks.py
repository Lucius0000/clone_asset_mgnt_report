#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import yfinance as yf
import numpy as np
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from dateutil.relativedelta import relativedelta  # 需要安装: pip install python-dateutil
import time

os.environ['http_proxy'] = 'http://127.0.0.1:7890'
os.environ['https_proxy'] = 'http://127.0.0.1:7890'


def adjust_date_for_market(date, symbol):
    """
    根据不同市场调整日期，处理时区差异
    
    时区问题说明：
    - 用户在北京时间输入日期（UTC+8）
    - yfinance默认按照各市场的本地时间处理：
      * 美股：美国东部时间（UTC-5/-4）
      * 港股：香港时间（UTC+8，与北京时间相同）
      * A股：北京时间（UTC+8）
    
    解决方案：
    - 美股：向前推1天，确保获取到正确的美国交易日数据
    - 港股/A股：不需要调整，时区相同
    """
    if symbol.endswith('.HK'):
        # 港股：使用香港时间，与北京时间相同 (UTC+8)
        return date
    elif symbol.endswith(('.SS', '.SZ')):
        # A股：使用北京时间 (UTC+8)
        return date
    else:
        # 美股：需要考虑时区差异
        # 北京时间比美国东部时间快12-13小时
        # 当用户输入北京时间5号时，美国时间可能还是4号
        # 为确保获取正确数据，向前推1天
        return date + timedelta(days=1)


def get_market_adjusted_date_range(start_date, end_date, symbol):
    """
    获取针对特定市场调整后的日期范围
    """
    adjusted_start = adjust_date_for_market(start_date, symbol)
    adjusted_end = adjust_date_for_market(end_date, symbol)

    # 为了确保获取足够的数据，稍微扩大范围
    buffer_start = adjusted_start - timedelta(days=3)
    buffer_end = adjusted_end + timedelta(days=3)

    return buffer_start, buffer_end


# 建议通过环境变量来获取数据库连接信息
conn_info = {
    'dbname': os.getenv('DB_NAME', 'ibkr_data'),
    'user': os.getenv('DB_USER', 'postgres'),
    'password': os.getenv('DB_PASSWORD', '441322191139'),
    'host': os.getenv('DB_HOST', 'localhost'),
    'port': os.getenv('DB_PORT', '5432')
}

# 美股标的分类（从原market_report_custom.py）
us_market_symbols = {
    # 股票-大盘
    'SPY': 'SPY',

    # 股票-行业-科技
    'QQQ': 'QQQ',

    # 股票-个股-七巨头
    'AMZN': 'AMZN',
    'GOOG': 'GOOG',
    'AAPL': 'AAPL',
    'MSFT': 'MSFT',
    'META': 'META',
    'TSLA': 'TSLA',
    'NVDA': 'NVDA',

    # 股票-个股-半导体
    'TSM': 'TSM',
    'AVGO': 'AVGO',
    'AMD': 'AMD',
    'INTC': 'INTC',
    'QCOM': 'QCOM',

    # 股票-个股-Data & AI Mgnt
    'SNOW': 'SNOW',
    'ORCL': 'ORCL',
    'MDB': 'MDB',
    'PLTR': 'PLTR',
    'DDOG': 'DDOG',

    # 股票-个股-SaaS
    'CRM': 'CRM',
    'APP': 'APP',
    'ADBE': 'ADBE',
    'NOW': 'NOW',
    'WDAY': 'WDAY',

    # 股票-个股-中概
    'BABA': 'BABA',
    'PDD': 'PDD',
    'JD': 'JD',
    'BIDU': 'BIDU',
    'NTES': 'NTES',

    # 债券-大盘
    'BND': 'BND',
    'AGG': 'AGG'
}

us_categories = {
    '股票-大盘': ['SPY'],
    '股票-行业-科技': ['QQQ'],
    '股票-个股-七巨头': ['AMZN', 'GOOG', 'AAPL', 'MSFT', 'META', 'TSLA', 'NVDA'],
    '股票-个股-半导体': ['NVDA', 'TSM', 'AVGO', 'AMD', 'INTC', 'QCOM'],
    '股票-个股-Data & AI Mgnt': ['SNOW', 'ORCL', 'MDB', 'PLTR', 'DDOG'],
    '股票-个股-SaaS': ['CRM', 'APP', 'ADBE', 'NOW', 'WDAY'],
    '股票-个股-中概': ['BABA', 'PDD', 'JD', 'BIDU', 'NTES'],
    '债券-大盘': ['BND', 'AGG']
}

# 中港股标的分类（按照用户要求的严格分类）
china_hk_market_symbols = {
    # 中国A股
    '华泰柏瑞沪深300ETF': '510300.SS',
    '华夏科创50ETF': '588000.SS',
    '易方达蓝筹精选': '005827.SS',
    '贵州茅台': '600519.SS',
    '工商银行': '601398.SS',
    '比亚迪A': '002594.SZ',
    '中芯国际A': '688981.SS',

    # 港股
    '恒生指数盈富基金': '2800.HK',
    # '中证香港300本地股ETF': '',  # 需要确认具体代码
    '华夏恒生科技ETF': '3032.HK',  # 使用港股代码
    # '惠理高息股票基金': '',  # 需要确认具体代码
    '汇丰控股': '0005.HK',
    '友邦保险': '1299.HK',
    '新鸿基地产': '0016.HK',
    '领展房产基金': '0823.HK',
    '中电控股': '0002.HK',
    '香港中华煤气': '0003.HK',
    '中广核电力': '1816.HK',
    '中海油': '0883.HK',  # 中海油有限公司
    '腾讯控股': '0700.HK',
    '阿里巴巴': '9988.HK',
    '美团': '3690.HK',
    '比亚迪H': '1211.HK',
    '中芯国际H': '0981.HK'
}

china_hk_categories = {
    # 中国A股分类
    '股票-大盘': ['华泰柏瑞沪深300ETF', '华夏科创50ETF'],
    '股票-策略': ['易方达蓝筹精选'],
    '股票-个股-消费': ['贵州茅台'],
    '股票-个股-金融': ['工商银行'],
    '股票-个股-新能源': ['比亚迪A'],
    '股票-个股-半导体': ['中芯国际A'],

    # 港股分类
    '港股-股票-大盘': ['恒生指数盈富基金'],
    '港股-股票-行业': ['华夏恒生科技ETF'],
    '港股-股票-个股-金融': ['汇丰控股', '友邦保险'],
    '港股-股票-个股-地产': ['新鸿基地产', '领展房产基金'],
    '港股-股票-个股-公用事业': ['中电控股', '香港中华煤气'],
    '港股-股票-个股-能源': ['中广核电力', '中海油'],
    '港股-股票-个股-互联网': ['腾讯控股', '阿里巴巴', '美团'],
    '港股-股票-个股-新能源': ['比亚迪H'],
    '港股-股票-个股-半导体': ['中芯国际H']
}


def safe_get_first(value):
    """统一处理可能为列表、Series等类型的数据，返回第一个有效值。"""
    if isinstance(value, (list, tuple)):
        return value[1] if len(value) > 1 else value[0]
    elif isinstance(value, pd.Series):
        return value.iloc[0]
    return value


def get_date_input(prompt):
    """获取用户输入的日期，并进行格式和逻辑校验"""
    while True:
        try:
            print(f"\n📅 输入{prompt}:")
            year = int(input("请输入年份 (如2024): "))

            # 年份合理性检查
            if year < 1900 or year > datetime.now().year:
                print(f"年份应该在1900到{datetime.now().year}之间，请重新输入。")
                continue

            month = int(input("请输入月份 (1-12): "))
            if month < 1 or month > 12:
                print("月份应该在1到12之间，请重新输入。")
                continue

            day = int(input("请输入日期 (1-31): "))
            if day < 1 or day > 31:
                print("日期应该在1到31之间，请重新输入。")
                continue

            date = datetime(year, month, day)

            # 检查日期不能是未来日期
            if date > datetime.now():
                print("❌ 错误：不能输入未来的日期，请重新输入。")
                continue

            # 检查日期不能过于久远（超过20年）
            twenty_years_ago = datetime.now() - timedelta(days=365 * 20)
            if date < twenty_years_ago:
                print(f"⚠️  警告：输入的日期过于久远（{date.strftime('%Y-%m-%d')}），可能无法获取到准确的金融数据。")
                confirm = input("是否继续使用此日期？(y/n): ").strip().lower()
                if confirm not in ['y', 'yes', '是']:
                    continue

            print(
                f"✅ {prompt}已确认：{date.strftime('%Y-%m-%d')}({['周一', '周二', '周三', '周四', '周五', '周六', '周日'][date.weekday()]})")
            return date

        except ValueError as e:
            print(f"❌ 输入的日期无效：{e}，请重新输入。")


def get_default_dates():
    """
    获取默认的金融分析日期范围
    
    逻辑：
    - end_date: 本周六（确保包含本周五的交易数据）
    - start_date: 上上周五（标准的两周分析起点）
    
    这样设计的原因：
    1. 金融市场交易周是周一到周五
    2. 周六作为结束日期确保包含本周五的数据
    3. 与用户验证的模式一致：周五→周六的数据能计算正确的两周变动
    4. 不受运行脚本具体时间影响，结果稳定
    """
    today = datetime.now()
    current_weekday = today.weekday()  # 0=周一, 1=周二, ..., 5=周六, 6=周日

    # 计算本周六作为 end_date
    if current_weekday == 5:  # 今天是周六
        end_date = today
    elif current_weekday == 6:  # 今天是周日
        end_date = today - timedelta(days=1)  # 昨天是周六
    else:  # 周一到周五
        days_until_saturday = 5 - current_weekday
        end_date = today + timedelta(days=days_until_saturday)

    # 计算上上周五作为 start_date（从本周六向前推15天）
    start_date = end_date - timedelta(days=15)

    # 验证start_date确实是周五，如果不是则调整
    while start_date.weekday() != 4:  # 4 = 周五
        start_date -= timedelta(days=1)

    return start_date, end_date


# 选择市场模式
print("选择市场模式：")
print("1. 美股模式")
print("2. 中港股模式")
print("3. 混合模式（美股+中港股）")

market_choice = input("请选择 (1、2 或 3): ").strip()

market_type = ""
if market_choice == "1":
    market_symbols = us_market_symbols
    categories = us_categories
    market_type = "美股"
    report_prefix = "us_market_report"
elif market_choice == "2":
    market_symbols = china_hk_market_symbols
    categories = china_hk_categories
    market_type = "中港股"
    report_prefix = "china_hk_market_report"
else:
    # 混合模式
    market_symbols = {**us_market_symbols, **china_hk_market_symbols}
    categories = {**us_categories, **china_hk_categories}
    market_type = "混合"
    report_prefix = "mixed_market_report"

print(f"已选择{market_type}模式")

# 获取日期
print("选择日期输入方式：")
print("1. 使用智能默认日期（上周五→本周六）")
print("2. 手动输入日期")

choice = input("请选择 (1 或 2): ").strip()

if choice == "1":
    start_date, end_date = get_default_dates()
    weekday_names = ['周一', '周二', '周三', '周四', '周五', '周六', '周日']
    start_weekday = weekday_names[start_date.weekday()]
    end_weekday = weekday_names[end_date.weekday()]
    print(
        f"使用智能默认日期：{start_date.strftime('%Y-%m-%d')}({start_weekday}) 到 {end_date.strftime('%Y-%m-%d')}({end_weekday})")
else:
    print("请手动输入日期")
    start_date = get_date_input("开始日期")
    end_date = get_date_input("结束日期")

    # 改进的日期验证逻辑：更加用户友好
    while start_date >= end_date:
        print("错误：开始日期必须早于结束日期！")
        print(f"当前输入：开始日期 {start_date.strftime('%Y-%m-%d')}，结束日期 {end_date.strftime('%Y-%m-%d')}")

        # 询问用户要修改哪个日期
        print("请选择要修改的日期：")
        print("1. 修改开始日期")
        print("2. 修改结束日期")
        print("3. 重新输入全部日期")

        modify_choice = input("请选择 (1、2 或 3): ").strip()

        if modify_choice == "1":
            start_date = get_date_input("开始日期")
        elif modify_choice == "2":
            end_date = get_date_input("结束日期")
        else:
            # 选择3或其他：重新输入全部
            start_date = get_date_input("开始日期")
            end_date = get_date_input("结束日期")

    # 显示最终确认的日期
    print(f"✓ 确认日期范围：{start_date.strftime('%Y-%m-%d')} 到 {end_date.strftime('%Y-%m-%d')}")
    print(f"  分析时间跨度：{(end_date - start_date).days} 天")


def get_weekly_data(symbols, start_date, end_date):
    """下载指定符号的历史数据（修复版本）"""
    data = {}
    total_symbols = len(symbols)

    print(f"\n开始下载 {total_symbols} 个符号的数据...")

    for i, (name, symbol) in enumerate(symbols.items(), 1):
        try:
            print(f"[{i}/{total_symbols}] 正在获取 {name} ({symbol}) 的数据...")

            # 修复 FutureWarning - 明确设置 auto_adjust 参数
            df = yf.download(
                symbol,
                start=start_date,
                end=end_date,
                auto_adjust=True,  # 明确设置以避免警告
                progress=False  # 关闭进度条以减少输出
            )

            if not df.empty:
                data[name] = df
                print(f"  ✓ 成功获取 {len(df)} 条数据")
            else:
                print(f"  ✗ 警告: {name}({symbol}) 在指定日期范围内无数据")

            # 添加短暂延迟以避免请求过于频繁
            time.sleep(0.1)

        except Exception as e:
            print(f"  ✗ 获取 {name}({symbol}) 数据时出错: {e}")

    print(f"数据下载完成！成功获取 {len(data)} 个符号的数据。\n")
    return data


# 添加新的标准化指标计算函数
def get_standardized_two_week_change(symbol: str, df=None, start_date=None, end_date=None) -> float:
    """
    计算标准的两周变动率

    参数:
    - symbol: 股票代码
    - df: 已下载的数据，如果提供则优先使用
    - start_date: 开始日期，如果提供则计算从 start_date 到 end_date 的变动率
    - end_date: 结束日期

    逻辑:
    1. 如果提供了 df 和 start_date，即手动选择日期，使用指定日期范围计算变动率
    2. 否则（即智能日期选择）按两周前的交易日回溯计算
    """
    try:
        if df is not None and start_date is not None:
            if df.empty or len(df) < 2:
                print(f"警告：{symbol} 的数据不足，无法计算变动率")
                return 0

            start_price = df.iloc[0]['Close']
            end_price = df.iloc[-1]['Close']

            change = ((end_price - start_price) / start_price) * 100
            return change

        ticker = yf.Ticker(symbol)
        if end_date is None:
            end_date = datetime.now()

        start_date_auto = end_date - timedelta(days=31)
        buffer_start, buffer_end = get_market_adjusted_date_range(start_date_auto, end_date, symbol)

        data = ticker.history(start=buffer_start, end=buffer_end, auto_adjust=True)

        if data.empty or len(data) < 2:
            print(f"警告：无法获取 {symbol} 的两周变动数据")
            return 0

        current_price = data.iloc[-1]['Close']
        current_date = data.index[-1]

        two_weeks_ago_price = None
        target_trading_days = [10, 11, 9, 12, 13]

        for trading_days in target_trading_days:
            if len(data) >= trading_days + 1:
                candidate_price = data.iloc[-(trading_days + 1)]['Close']
                candidate_date = data.index[-(trading_days + 1)]

                days_diff = (current_date.date() - candidate_date.date()).days
                if 9 <= days_diff <= 16:
                    two_weeks_ago_price = candidate_price
                    break

        if two_weeks_ago_price is None:
            if len(data) >= 13:
                two_weeks_ago_price = data.iloc[-13]['Close']
            elif len(data) >= 11:
                two_weeks_ago_price = data.iloc[-11]['Close']
            elif len(data) >= 2:
                two_weeks_ago_price = data.iloc[0]['Close']
            else:
                return 0

        two_week_change = ((current_price - two_weeks_ago_price) / two_weeks_ago_price) * 100

        return two_week_change

    except Exception as e:
        print(f"计算两周变动时出错：{symbol}，错误：{e}")
        return 0

def get_standardized_ytd(symbol: str, end_date: datetime = None) -> float:
    """
    计算标准的年初至今收益率 (YTD)。
    该方法使用上一年最后一个交易日的收盘价作为计算基准，这是最标准的做法。
    """
    if end_date is None:
        end_date = datetime.now()

    # 1. 确定获取数据的日期范围
    # 为了确保能获取到上一年最后一个交易日，我们将开始日期设为上一年的12月20日左右
    start_of_fetch_range = datetime(end_date.year - 1, 12, 20)

    try:
        # 2. 获取数据
        ticker = yf.Ticker(symbol)
        # 我们获取从去年年底到指定结束日期的数据
        data = ticker.history(start=start_of_fetch_range, end=end_date + timedelta(days=1), auto_adjust=True)

        if data.empty:
            print(f"警告：无法获取 {symbol} 在指定日期范围的历史数据。")
            return 0.0

        # 3. 找到上一年最后一个交易日的收盘价 (起始价格)
        prev_year_data = data[data.index.year == end_date.year - 1]
        if prev_year_data.empty:
            print(f"警告：找不到 {symbol} 在 {end_date.year - 1} 年的交易数据。")
            return 0.0

        # 上一年最后一个交易日的收盘价即为我们的起始价格
        start_price = prev_year_data['Close'].iloc[-1]
        start_price_date = prev_year_data.index[-1].date()

        # 4. 找到 end_date 或之前最近的交易日收盘价 (当前价格)
        current_year_data = data[data.index.year == end_date.year]
        # 筛选出不晚于 end_date 的数据
        current_year_data = current_year_data[current_year_data.index.date <= end_date.date()]
        if current_year_data.empty:
            print(f"警告：找不到 {symbol} 在 {end_date.year} 年截至 {end_date.date()} 的交易数据。")
            return 0.0

        current_price = current_year_data['Close'].iloc[-1]
        current_price_date = current_year_data.index[-1].date()

        print(f"计算 {symbol} YTD:")
        print(f" - 起始日期 (上年收盘): {start_price_date}, 价格: {start_price:.2f}")
        print(f" - 结束日期 (当前): {current_price_date}, 价格: {current_price:.2f}")

        # 5. 计算 YTD 收益率
        if start_price == 0:
            print(f"警告：起始价格为0，无法计算YTD。")
            return 0.0

        ytd_return = ((current_price - start_price) / start_price) * 100
        return ytd_return

    except Exception as e:
        print(f"计算YTD时出错：{symbol}，错误：{e}")
        return 0.0


def get_standardized_mom(symbol: str, end_date=None) -> float:
    """
    计算基于日历的精确月环比收益率 (MoM)。
    """
    if end_date is None:
        end_date = datetime.now().date()
    else:
        end_date = end_date.date()

    # 1. 定义查找范围
    # 获取额外的数据以确保能找到目标日期
    fetch_start_date = end_date - timedelta(days=45)

    try:
        # 2. 获取历史数据
        ticker = yf.Ticker(symbol)
        data = ticker.history(start=fetch_start_date, end=end_date + timedelta(days=1), auto_adjust=True)

        if data.empty or len(data) < 2:
            print(f"警告：无法获取 {symbol} 的MoM数据")
            return 0.0

        # 将索引转换为日期，方便比较
        data.index = data.index.date

        # 3. 确定当前价格和目标日期
        current_price_series = data[data.index <= end_date]
        if current_price_series.empty:
            print(f"警告：找不到 {symbol} 在 {end_date} 或之前的价格。")
            return 0.0
        current_price = current_price_series['Close'].iloc[-1]
        current_date = current_price_series.index[-1]

        # 计算一个月前的目标日期
        target_date_1m_ago = current_date - relativedelta(months=1)

        # 4. 寻找一个月前最接近的实际交易日价格
        # 从数据中筛选出所有早于或等于目标日期的记录
        past_data = data[data.index <= target_date_1m_ago]
        if past_data.empty:
            print(f"警告：找不到 {symbol} 在 {target_date_1m_ago} 或之前的足够历史数据。")
            return 0.0  # 或者返回 None 表示无法计算

        month_ago_price = past_data['Close'].iloc[-1]
        month_ago_date = past_data.index[-1]

        print(f"计算 {symbol} MoM:")
        print(f" - 起始日期: {month_ago_date}, 价格: {month_ago_price:.2f}")
        print(f" - 结束日期: {current_date}, 价格: {current_price:.2f}")

        # 5. 计算收益率
        return ((current_price - month_ago_price) / month_ago_price) * 100

    except Exception as e:
        print(f"计算MoM时出错：{symbol}，错误：{e}")
        return 0.0


def get_standardized_yoy(symbol: str, end_date: datetime = None) -> float:
    #    """
    #     计算基于日历的精确年同比收益率 (YoY)。
    #     """
    if end_date is None:
        end_date = datetime.now().date()
    else:
        end_date = end_date.date()

    # 1. 定义查找范围 (一年大约365天，加一些缓冲)
    fetch_start_date = end_date - timedelta(days=380)

    try:
        # 2. 获取历史数据
        ticker = yf.Ticker(symbol)
        data = ticker.history(start=fetch_start_date, end=end_date + timedelta(days=1), auto_adjust=True)

        if data.empty or len(data) < 2:
            print(f"警告：无法获取 {symbol} 的YoY数据")
            return 0.0

        data.index = data.index.date

        # 3. 确定当前价格和目标日期
        current_price_series = data[data.index <= end_date]
        if current_price_series.empty:
            print(f"警告：找不到 {symbol} 在 {end_date} 或之前的价格。")
            return 0.0
        current_price = current_price_series['Close'].iloc[-1]
        current_date = current_price_series.index[-1]

        # 计算一年前的目标日期
        target_date_1y_ago = current_date - relativedelta(years=1)

        # 4. 寻找一年前最接近的实际交易日价格
        past_data = data[data.index <= target_date_1y_ago]
        if past_data.empty:
            print(f"警告：找不到 {symbol} 在 {target_date_1y_ago} 或之前的足够历史数据。")
            return 0.0

        year_ago_price = past_data['Close'].iloc[-1]
        year_ago_date = past_data.index[-1]

        print(f"计算 {symbol} YoY:")
        print(f" - 起始日期: {year_ago_date}, 价格: {year_ago_price:.2f}")
        print(f" - 结束日期: {current_date}, 价格: {current_price:.2f}")

        # 5. 计算收益率
        return ((current_price - year_ago_price) / year_ago_price) * 100

    except Exception as e:
        print(f"计算YoY时出错：{symbol}，错误：{e}")
        return 0.0


def get_standardized_market_cap(symbol, end_date=None):
    """计算标准市值（基于指定日期的收盘价）"""
    try:
        stock = yf.Ticker(symbol)
        info = stock.info
        shares = info.get('sharesOutstanding')

        if end_date:
            # 使用时区调整的日期范围
            start_date = end_date - timedelta(days=7)
            buffer_start, buffer_end = get_market_adjusted_date_range(start_date, end_date, symbol)

            hist = stock.history(start=buffer_start, end=buffer_end, auto_adjust=True)
            if not hist.empty:
                price = hist.iloc[-1]['Close']
            else:
                price = info.get('currentPrice')
        else:
            price = info.get('currentPrice')

        if shares and price:
            # 根据不同市场计算市值
            if symbol.endswith('.HK'):
                return shares * price / 1e8  # 转换为亿港币
            elif symbol.endswith(('.SS', '.SZ')):
                return shares * price / 1e8  # 转换为亿人民币
            else:
                return shares * price / 1e8  # 美股转换为亿美元
    except Exception as e:
        print(f"获取市值时出错：{symbol}，错误：{e}")
    return None

def get_standardized_dividend_yield(symbol):
    """获取标准化的股息率（百分比）"""
    try:
        info = yf.Ticker(symbol).info
        dy = info.get('dividendYield')
        if dy is not None:
            return round(dy, 2)
    except Exception as e:
        print(f"获取股息率时出错：{symbol}，错误：{e}")
    return None


def get_standardized_annualized_volatility(symbol, end_date=None, period_days=252):
    """计算标准年化波动率（基于指定日期向前推算）"""
    try:
        stock = yf.Ticker(symbol)
        if end_date is None:
            end_date = datetime.now()

        # 使用时区调整的日期范围
        start_date = end_date - timedelta(days=period_days + 50)  # 多加50天确保有足够交易日
        buffer_start, buffer_end = get_market_adjusted_date_range(start_date, end_date, symbol)

        hist = stock.history(start=buffer_start, end=buffer_end, auto_adjust=True)

        if not hist.empty and len(hist) >= 20:  # 至少需要20个交易日
            # 取最近的交易日数据，最多取period_days天
            if len(hist) > period_days:
                hist = hist.tail(period_days)

            daily_return = hist['Close'].pct_change().dropna()
            if len(daily_return) > 0:
                daily_vol = daily_return.std()
                return daily_vol * np.sqrt(252)  # 年化波动率
    except Exception as e:
        print(f"计算年化波动率时出错：{symbol}，错误：{e}")
    return None


def get_standardized_sharpe_ratio(symbol, end_date=None, risk_free_rate=0.02):
    """计算标准夏普比率 - 使用1年数据"""
    try:
        stock = yf.Ticker(symbol)
        if end_date is None:
            end_date = datetime.now()

        # 获取1年的历史数据（约252个交易日）
        start_date = end_date - timedelta(days=365 + 50)  # 多加50天确保有足够交易日
        buffer_start, buffer_end = get_market_adjusted_date_range(start_date, end_date, symbol)

        hist = stock.history(start=buffer_start, end=buffer_end, auto_adjust=True)

        if not hist.empty and len(hist) >= 252:  # 至少需要1年数据
            # 计算1年年化收益率
            start_price = hist.iloc[0]['Close']
            end_price = hist.iloc[-1]['Close']
            total_return = (end_price - start_price) / start_price

            # 计算实际年数（基于交易日）
            actual_days = len(hist)
            years = actual_days / 252  # 转换为年数

            # 年化收益率
            annualized_return = (1 + total_return) ** (1 / years) - 1

            # 计算年化波动率
            daily_returns = hist['Close'].pct_change().dropna()
            if len(daily_returns) > 0:
                annualized_vol = daily_returns.std() * np.sqrt(252)

                if annualized_vol != 0:
                    return (annualized_return - risk_free_rate) / annualized_vol
        elif not hist.empty and len(hist) >= 60:  # 如果数据不足1年但至少有60个交易日
            # 使用可用数据计算，但给出警告
            print(f"警告：{symbol} 的历史数据不足1年（{len(hist)}个交易日），夏普比率可能不够准确")

            start_price = hist.iloc[0]['Close']
            end_price = hist.iloc[-1]['Close']
            total_return = (end_price - start_price) / start_price

            actual_days = len(hist)
            years = actual_days / 252

            annualized_return = (1 + total_return) ** (1 / years) - 1

            daily_returns = hist['Close'].pct_change().dropna()
            if len(daily_returns) > 0:
                annualized_vol = daily_returns.std() * np.sqrt(252)

                if annualized_vol != 0:
                    return (annualized_return - risk_free_rate) / annualized_vol
        else:
            print(f"警告：{symbol} 的历史数据不足，无法计算夏普比率")

    except Exception as e:
        print(f"计算夏普比率时出错：{symbol}，错误：{e}")
    return None


def calculate_indicators(df, symbol, market_symbols, market_type, start_date=None, end_date=None):
    """计算指标：根据市场类型计算不同的指标"""
    # 获取最新收盘价（用于显示）
    latest_close = safe_get_first(df['Close'].iloc[-1])

    market_symbol = market_symbols.get(symbol)
    if not market_symbol:
        raise ValueError(f"符号 {symbol} 不在市场符号列表中。")

    # 计算两周变动：在自定义日期模式下直接使用已下载的数据
    if start_date is not None:
        # 自定义日期模式：计算从start_date到end_date的变动率
        first_close = safe_get_first(df['Close'].iloc[0])
        if first_close == 0:
            raise ValueError("初始收盘价为 0，无法计算涨跌幅。")
        two_week_return_value = (latest_close - first_close) / first_close * 100
    else:
        # 智能默认日期模式：使用标准化函数计算"一周前"的变动率
        two_week_return_value = get_standardized_two_week_change(market_symbol, None, None, end_date)

    # 其他指标使用标准化函数（这些需要特定时间基准，合理重新下载数据）
    market_cap = get_standardized_market_cap(market_symbol, end_date)
    ytd_rate = get_standardized_ytd(market_symbol, end_date)
    mom_rate = get_standardized_mom(market_symbol, end_date)
    yoy_rate = get_standardized_yoy(market_symbol, end_date)
    sharpe_ratio = get_standardized_sharpe_ratio(market_symbol, end_date)

    if market_type == "美股":
        # 美股模式：包含两周变动和年化波动率
        annual_vol = get_standardized_annualized_volatility(market_symbol, end_date)

        return {
            'symbol': symbol,
            'market_region': '美股',
            'two_week_return': round(two_week_return_value, 2),
            'ytd_rate': round(ytd_rate, 2),
            'mom_rate': round(mom_rate, 2),
            'yoy_rate': round(yoy_rate, 2),
            'latest_close': round(latest_close, 2),
            'market_cap': round(market_cap, 2) if market_cap is not None else None,
            'sharp_ratio': round(sharpe_ratio, 2) if sharpe_ratio is not None else None,
            'dividend_yield': round(get_standardized_dividend_yield(market_symbol), 2) if get_standardized_dividend_yield(market_symbol) is not None else None,
            'annualized_volatility': round(annual_vol, 2) if annual_vol is not None else None
        }
    elif market_type == "中港股":
        # 中港股模式：包含两周变动和年化波动率
        annual_vol = get_standardized_annualized_volatility(market_symbol, end_date)

        return {
            'symbol': symbol,
            'market_region': '中港股',
            'two_week_return': round(two_week_return_value, 2),
            'ytd_rate': round(ytd_rate, 2),
            'mom_rate': round(mom_rate, 2),
            'yoy_rate': round(yoy_rate, 2),
            'latest_close': round(latest_close, 2),
            'market_cap': round(market_cap, 2) if market_cap is not None else None,
            'sharp_ratio': round(sharpe_ratio, 2) if sharpe_ratio is not None else None,
            'dividend_yield': round(get_standardized_dividend_yield(market_symbol), 2) if get_standardized_dividend_yield(market_symbol) is not None else None,
            'annualized_volatility': round(annual_vol, 2) if annual_vol is not None else None
        }
    else:
        # 混合模式：需要检测symbol属于哪个市场
        is_us_stock = symbol in us_market_symbols

        if is_us_stock:
            # 美股标的：包含所有指标
            annual_vol = get_standardized_annualized_volatility(market_symbol, end_date)
            return {
                'symbol': symbol,
                'market_region': '美股',
                'two_week_return': round(two_week_return_value, 2),
                'ytd_rate': round(ytd_rate, 2),
                'mom_rate': round(mom_rate, 2),
                'yoy_rate': round(yoy_rate, 2),
                'latest_close': round(latest_close, 2),
                'market_cap': round(market_cap, 2) if market_cap is not None else None,
                'sharp_ratio': round(sharpe_ratio, 2) if sharpe_ratio is not None else None,
                'dividend_yield': round(get_standardized_dividend_yield(market_symbol), 2) if get_standardized_dividend_yield(market_symbol) is not None else None,
                'annualized_volatility': round(annual_vol, 2) if annual_vol is not None else None
            }
        else:
            # 中港股标的：现在也计算所有指标
            annual_vol = get_standardized_annualized_volatility(market_symbol, end_date)
            return {
                'symbol': symbol,
                'market_region': '中港股',
                'two_week_return': round(two_week_return_value, 2),
                'ytd_rate': round(ytd_rate, 2),
                'mom_rate': round(mom_rate, 2),
                'yoy_rate': round(yoy_rate, 2),
                'latest_close': round(latest_close, 2),
                'market_cap': round(market_cap, 2) if market_cap is not None else None,
                'sharp_ratio': round(sharpe_ratio, 2) if sharpe_ratio is not None else None,
                'dividend_yield': round(get_standardized_dividend_yield(market_symbol), 2) if get_standardized_dividend_yield(market_symbol) is not None else None,
                'annualized_volatility': round(annual_vol, 2) if annual_vol is not None else None
            }


def get_gradient_fill(value, max_value, min_value):
    """根据数值大小生成渐变色填充"""
    if np.isnan(value) or max_value == min_value:
        color = 'FFFFFF'
    elif value > 0:
        intensity = int(150 * (value / max_value))
        color = f'FF{(210 - intensity):02X}{(210 - intensity):02X}'
    elif value < 0:
        intensity = int(210 * (abs(value) / abs(min_value)))
        color = f'{(210 - intensity):02X}FF{(210 - intensity):02X}'
    else:
        color = 'FFFFFF'
    return PatternFill(start_color=color, end_color=color, fill_type='solid')


def apply_gradient_fill(ws, skip_rows=1, skip_columns=0):
    """给工作表应用渐变色填充（跳过表头和指定列）"""
    start_row = ws.min_row + skip_rows
    start_col = ws.min_column + skip_columns

    for col_idx in range(start_col + 1, ws.max_column + 1):
        col_values = []
        for row_idx in range(start_row, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if isinstance(value, (int, float)) and not np.isnan(value):
                col_values.append(float(value))
        if col_values:
            max_val, min_val = max(col_values), min(col_values)
            for row_idx in range(start_row, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)) and not np.isnan(cell.value):
                    fill = get_gradient_fill(float(cell.value), max_val, min_val)
                    cell.fill = fill


def get_category_for_symbol(symbol, categories):
    """获取符号所属的类别"""
    for category, symbols in categories.items():
        if symbol in symbols:
            return category
    return "其他"


def export_to_excel_by_category(data, categories, report_prefix, market_type):
    """将数据按分类导出为Excel文件，并设置单元格样式和颜色填充"""

    # 创建工作簿
    wb = Workbook()

    # 先删除默认工作表，我们要创建按分类的工作表
    wb.remove(wb.active)

    # 字体和样式设置
    header_font_ch = Font(name='SimSun', size=11, bold=True)
    header_font_en = Font(name='New Times Roman', size=11, bold=True)
    data_font_ch = Font(name='SimSun', size=11)
    data_font_en = Font(name='New Times Roman', size=11)
    center_align = Alignment(horizontal="center")
    right_align = Alignment(horizontal="right")
    grey_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # 按分类创建工作表
    for category, symbols in categories.items():
        # 筛选该分类的数据
        category_data = {symbol: data[symbol] for symbol in symbols if symbol in data}

        if not category_data:
            continue

        # 创建工作表
        ws = wb.create_sheet(title=category.replace('-', '_'))

        # 创建DataFrame，按照不同市场类型重新排列列顺序
        df = pd.DataFrame(category_data).T

        if market_type == "美股":
            # 美股模式列顺序：Symbol, 两周变动(%), YTD(%), MoM(%), YoY(%), 收盘, 市值(亿美元), 夏普比率, 年化波动率
            df = df[['symbol', 'two_week_return', 'ytd_rate', 'mom_rate', 'yoy_rate', 'latest_close',
                     'market_cap', 'dividend_yield', 'sharp_ratio', 'annualized_volatility']]
            df.columns = ['Symbol', '两周变动(%)', 'YTD(%)', 'MoM(%)', 'YoY(%)', '收盘', '市值(亿)', '股息率(%)', '夏普比率', '年化波动率']

        elif market_type == "中港股":
            # 中港股模式列顺序：现在和美股保持一致
            df = df[['symbol', 'two_week_return', 'ytd_rate', 'mom_rate', 'yoy_rate', 'latest_close',
                     'market_cap', 'dividend_yield', 'sharp_ratio', 'annualized_volatility']]
            df.columns = ['Symbol', '两周变动(%)', 'YTD(%)', 'MoM(%)', 'YoY(%)', '收盘', '市值(亿)', '股息率(%)', '夏普比率', '年化波动率']

        else:
            # 混合模式列顺序：市场区域, Symbol, 两周变动(%), YTD(%), MoM(%), YoY(%), 收盘, 市值(亿), 夏普比率, 年化波动率
            df = df[['market_region', 'symbol', 'two_week_return', 'ytd_rate', 'mom_rate', 'yoy_rate',
                     'latest_close', 'market_cap', 'dividend_yield', 'sharp_ratio', 'annualized_volatility']]
            df.columns = ['市场区域', 'Symbol', '两周变动(%)', 'YTD(%)', 'MoM(%)', 'YoY(%)', '收盘',
                          '市值(亿)', '股息率(%)', '夏普比率', '年化波动率']

        numeric_cols = df.select_dtypes(include=['number']).columns
        df[numeric_cols] = df[numeric_cols].applymap(lambda x: round(x, 2) if pd.notna(x) else x)
        df.replace({0: 'n/a'}, inplace=True)
        df.fillna('n/a', inplace=True)

        # 写入表头
        for col_idx, header in enumerate(df.columns.tolist(), start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font_en if header.isascii() else header_font_ch
            cell.alignment = center_align
            if header == 'n/a':
                cell.fill = grey_fill

        # 写入数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if isinstance(value, str):
                    cell.font = data_font_en if value.isascii() else data_font_ch
                else:
                    cell.font = data_font_ch

                # 根据模式调整对齐方式
                if market_type == "混合":
                    # 混合模式：市场区域和Symbol列居中，其他右对齐
                    cell.alignment = center_align if c_idx <= 2 else right_align
                else:
                    # 单一模式：只有Symbol列居中，其他右对齐
                    cell.alignment = center_align if c_idx == 1 else right_align

                if value == 'n/a':
                    cell.fill = grey_fill

        # 应用渐变填充
        if market_type == "混合":
            apply_gradient_fill(ws, skip_rows=1, skip_columns=1)  # 跳过市场区域列
        else:
            apply_gradient_fill(ws, skip_rows=1, skip_columns=0)

    # 创建汇总工作表
    ws_summary = wb.create_sheet(title="汇总", index=0)

    # 创建汇总DataFrame，按照不同市场类型重新排列列顺序
    df_summary = pd.DataFrame(data).T

    if market_type == "美股":
        # 美股模式列顺序
        df_summary = df_summary[
            ['symbol', 'two_week_return', 'ytd_rate', 'mom_rate', 'yoy_rate', 'latest_close',
             'market_cap', 'dividend_yield', 'sharp_ratio', 'annualized_volatility']]
        # 添加分类列
        df_summary.insert(0, 'category', [get_category_for_symbol(symbol, categories) for symbol in df_summary.index])
        # 设置表头
        df_summary.columns = ['分类', 'Symbol', '两周变动(%)', 'YTD(%)', 'MoM(%)', 'YoY(%)', '收盘',
                              '市值(亿美元)', '股息率(%)', '夏普比率', '年化波动率']
    elif market_type == "中港股":
        # 中港股模式列顺序：现在和美股保持一致
        df_summary = df_summary[
            ['symbol', 'two_week_return', 'ytd_rate', 'mom_rate', 'yoy_rate', 'latest_close',
             'market_cap', 'dividend_yield', 'sharp_ratio', 'annualized_volatility']]
        # 添加分类列
        df_summary.insert(0, 'category', [get_category_for_symbol(symbol, categories) for symbol in df_summary.index])
        # 设置表头
        df_summary.columns = ['分类', 'Symbol', '两周变动(%)', 'YTD(%)', 'MoM(%)', 'YoY(%)', '收盘',
                              '市值(亿美元)', '股息率(%)', '夏普比率', '年化波动率']
    else:
        # 混合模式列顺序
        df_summary = df_summary[
            ['market_region', 'symbol', 'two_week_return', 'ytd_rate', 'mom_rate', 'yoy_rate',
             'latest_close', 'market_cap', 'dividend_yield', 'sharp_ratio', 'annualized_volatility']]
        # 添加分类列
        df_summary.insert(0, 'category', [get_category_for_symbol(symbol, categories) for symbol in df_summary.index])
        # 设置表头
        df_summary.columns = ['分类', '市场区域', 'Symbol', '两周变动(%)', 'YTD(%)', 'MoM(%)', 'YoY(%)',
                              '收盘', '市值(亿)', '股息率(%)', '夏普比率', '年化波动率']

    numeric_cols = df_summary.select_dtypes(include=['number']).columns
    df_summary[numeric_cols] = df_summary[numeric_cols].applymap(lambda x: round(x, 2) if pd.notna(x) else x)
    df_summary.replace({0: 'n/a'}, inplace=True)
    df_summary.fillna('n/a', inplace=True)
    # 按要求重排汇总sheet列顺序（以中文表头为准，避免列名不匹配被误裁剪）
    # 说明：此处 df_summary 的列名已在上方按不同市场类型设置为中文
    if market_type == "混合":
        desired_headers = ['分类', '市场区域', 'Symbol', '收盘', '两周变动(%)', 'MoM(%)', 'YoY(%)', 'YTD(%)', '夏普比率', '市值(亿)', '股息率(%)']
    else:
        # 单一市场模式无“市场区域”列
        desired_headers = ['分类', 'Symbol', '收盘', '两周变动(%)', 'MoM(%)', 'YoY(%)', 'YTD(%)', '夏普比率', '市值(亿)', '股息率(%)']

    # 仅选择实际存在的列，按期望顺序排列
    ordered_headers = [h for h in desired_headers if h in df_summary.columns]
    if ordered_headers:
        df_summary = df_summary[ordered_headers]

    # 写入汇总表头
    for col_idx, header in enumerate(df_summary.columns.tolist(), start=1):
        cell = ws_summary.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_en if header.isascii() else header_font_ch
        cell.alignment = center_align

    # 写入汇总数据
    for r_idx, row in enumerate(dataframe_to_rows(df_summary, index=False, header=False), start=2):
        for c_idx, value in enumerate(row, start=1):
            cell = ws_summary.cell(row=r_idx, column=c_idx, value=value)
            if isinstance(value, str):
                cell.font = data_font_en if value.isascii() else data_font_ch
            else:
                cell.font = data_font_ch

            # 根据模式调整对齐方式
            if market_type == "混合":
                # 混合模式：分类、市场区域、Symbol列居中，其他右对齐
                cell.alignment = center_align if c_idx <= 3 else right_align
            else:
                # 单一模式：分类和Symbol列居中，其他右对齐
                cell.alignment = center_align if c_idx <= 2 else right_align

            if value == 'n/a':
                cell.fill = grey_fill

    # 应用渐变填充到汇总表
    if market_type == "混合":
        apply_gradient_fill(ws_summary, skip_rows=1, skip_columns=2)  # 跳过分类列和市场区域列
    else:
        apply_gradient_fill(ws_summary, skip_rows=1, skip_columns=1)  # 跳过分类列

    filename = f'{report_prefix}_{end_date.strftime("%Y%m%d")}.xlsx'
    wb.save(filename)
    print(f"{market_type}分类报告已保存为: {filename}")


if __name__ == "__main__":
    print(f"数据日期范围：{start_date.strftime('%Y-%m-%d')} 到 {end_date.strftime('%Y-%m-%d')}")
    print(f"将分析以下 {len(market_symbols)} 个{market_type}标的：")

    # 按分类显示标的
    for category, symbols in categories.items():
        print(f"  {category}: {', '.join(symbols)}")

    # 下载数据
    weekly_data = get_weekly_data(market_symbols, start_date, end_date)

    # 计算指标
    results = {}

    print("正在计算指标...")
    for name, df in weekly_data.items():
        try:
            results[name] = calculate_indicators(df, name, market_symbols, market_type, start_date, end_date)
        except Exception as e:
            print(f"计算 {name} 指标时出错: {e}")

    # 导出Excel文件
    if results:
        export_to_excel_by_category(results, categories, report_prefix, market_type)
        print("数据处理完毕！")
        print(f"成功分析了 {len(results)} 个标的")

        # 显示各分类的标的数量
        for category, symbols in categories.items():
            count = sum(1 for symbol in symbols if symbol in results)
            print(f"  {category}: {count}/{len(symbols)} 个标的")
    else:
        print("没有获取到任何数据，请检查网络连接和符号列表。")
